"""Tests du module Tableau (secrétariat).

Couvre :
  - detect_anomalies / compute_analyse_stats (règles à jour : seuil 120h,
    critères par catégorie)
  - verdict.evaluer : INSCRIT / NON INSCRIT / À DÉLIBÉRER par section,
    exemptions (nouveau membre, âge) et réglages configurables
  - excel_import.parse_excel_bytes : multi-feuilles, en-tête décalé, mapping
  - exporter.build_workbook : sortie par section, numérotation, sociétés
    sans colonne formation
  - comparison.compare_exercices
"""
from __future__ import annotations

import io

import openpyxl

from app.modules.secretariat.tableau import verdict as V
from app.modules.secretariat.tableau.analyzer import (
    HEURES_FORCO_MIN,
    compute_analyse_stats,
    detect_anomalies,
)
from app.modules.secretariat.tableau.comparison import compare_exercices
from app.modules.secretariat.tableau.excel_import import parse_excel_bytes
from app.modules.secretariat.tableau.exporter import build_workbook
from app.modules.secretariat.tableau.service import _valider_lignes


def _dossier(**overrides) -> dict:
    base = {
        "id": 1,
        "numero_ordre": "EC/16.00001",
        "nom": "Dupont",
        "prenom": "Jean",
        "categorie": "EC Cabinet",
        "cotisation_payee": True,
        "heures_forco": 120.0,
        "assurance": True,
        "chiffre_affaires": True,
        "anciennete": "Ancien",
        "age": 45,
    }
    base.update(overrides)
    return base


class TestDetectAnomalies:
    def test_dossier_complet_ne_genere_aucune_anomalie(self):
        assert detect_anomalies([_dossier()]) == []

    def test_doublon_nom_prenom_insensible_a_la_casse(self):
        dossiers = [
            _dossier(id=1, nom="Dupont", prenom="Jean"),
            _dossier(id=2, nom="DUPONT", prenom="jean"),
        ]
        doublons = [a for a in detect_anomalies(dossiers) if a["type_anomalie"] == "doublon"]
        assert len(doublons) == 1
        assert doublons[0]["dossier_id"] == 2

    def test_nom_manquant_signale_dossier_incomplet(self):
        types = {a["type_anomalie"] for a in detect_anomalies([_dossier(nom="")])}
        assert "dossier_incomplet" in types

    def test_categorie_inconnue(self):
        anomalies = detect_anomalies([_dossier(categorie="Autre")])
        assert any(a["type_anomalie"] == "categorie_inconnue" for a in anomalies)

    def test_cotisation_non_payee(self):
        anomalies = detect_anomalies([_dossier(cotisation_payee=False)])
        assert any(a["type_anomalie"] == "cotisation_non_payee" and a["gravite"] == "high" for a in anomalies)

    def test_heures_insuffisantes_sous_120(self):
        anomalies = detect_anomalies([_dossier(heures_forco=50.0)])
        assert any(a["type_anomalie"] == "heures_forco_insuffisantes" for a in anomalies)

    def test_heures_au_seuil_120_pas_anomalie(self):
        anomalies = detect_anomalies([_dossier(heures_forco=HEURES_FORCO_MIN)])
        assert not any(a["type_anomalie"] == "heures_forco_insuffisantes" for a in anomalies)

    def test_nouveau_membre_exempte_de_formation(self):
        # 0h mais nouveau -> pas d'anomalie de formation
        anomalies = detect_anomalies([_dossier(heures_forco=0.0, anciennete="Nouveau")])
        assert not any("heures_forco" in a["type_anomalie"] for a in anomalies)

    def test_assurance_requise_seulement_pour_independant(self):
        # Cabinet sans assurance : pas d'anomalie assurance
        cab = detect_anomalies([_dossier(categorie="EC Cabinet", assurance=False)])
        assert not any("assurance" in a["type_anomalie"] for a in cab)
        # Indépendant sans assurance : anomalie
        indep = detect_anomalies([_dossier(categorie="EC Indépendant", assurance=False)])
        assert any(a["type_anomalie"] == "assurance_manquante" for a in indep)

    def test_chiffre_affaires_non_declare_independant(self):
        anomalies = detect_anomalies([_dossier(categorie="EC Indépendant", chiffre_affaires=False)])
        assert any(a["type_anomalie"] == "chiffre_affaires_non_declare" for a in anomalies)


class TestVerdict:
    def test_cabinet_inscrit(self):
        assert V.evaluer(_dossier())["conclusion"] == V.INSCRIT

    def test_cabinet_non_inscrit_formation(self):
        assert V.evaluer(_dossier(heures_forco=50.0))["conclusion"] == V.NON_INSCRIT

    def test_cotisation_impayee_non_inscrit(self):
        assert V.evaluer(_dossier(cotisation_payee=False))["conclusion"] == V.NON_INSCRIT

    def test_independant_tous_criteres(self):
        d = _dossier(categorie="EC Indépendant")
        assert V.evaluer(d)["conclusion"] == V.INSCRIT
        assert V.evaluer({**d, "chiffre_affaires": False})["conclusion"] == V.NON_INSCRIT

    def test_societe_sans_formation(self):
        # société : pas de critère formation, 0h ne bloque pas
        d = _dossier(categorie="Société", heures_forco=0.0, chiffre_affaires=True, assurance=True)
        assert V.evaluer(d)["conclusion"] == V.INSCRIT

    def test_nouveau_membre_par_numero_ordre(self):
        # exercice 2026, ordre 2025 -> nouveau -> exempté formation
        d = _dossier(numero_ordre="EC/25.00604", anciennete="Ancien", heures_forco=0.0)
        assert V.evaluer(d, 2026)["conclusion"] == V.INSCRIT

    def test_stagiaire_non_applicable(self):
        assert V.evaluer(_dossier(categorie="Stagiaire"))["conclusion"] == V.NON_APPLICABLE

    def test_age_action_a_deliberer(self):
        d = _dossier(heures_forco=0.0, age=65)
        assert V.evaluer(d, 2026)["conclusion"] == V.A_DELIBERER

    def test_age_action_inscrit_direct(self):
        d = _dossier(heures_forco=0.0, age=65)
        reg = V.TableauReglages(age_action="inscrit")
        assert V.evaluer(d, 2026, reg)["conclusion"] == V.INSCRIT

    def test_age_seuil_configurable(self):
        # seuil 99 : un membre de 65 ans n'est plus exempté
        d = _dossier(heures_forco=0.0, age=65)
        reg = V.TableauReglages(age_seuil=99)
        assert V.evaluer(d, 2026, reg)["conclusion"] == V.NON_INSCRIT

    def test_seuil_heures_configurable(self):
        d = _dossier(heures_forco=80.0)
        reg = V.TableauReglages(heures_formation_min=60)
        assert V.evaluer(d, 2026, reg)["conclusion"] == V.INSCRIT


def _make_workbook_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EC EN CABINET 25"
    ws.append([None] * 3)
    ws.append([None, "CONSEIL PROVINCIAL DE KINSHASA"])
    ws.append([None, "SECTION A1 : EXPERTS-COMPTABLES EN CABINET"])
    ws.append([None, "N°", "N° D'ORDRE", "NOM, POST NOMS ET PRENOMS", "Sexe",
               "N° TELEPHONE", "E-MAIL", "CABINET D'ATTACHE", "Ancienneté",
               "Cotisation", "H.Formation", "Conclusion", "NHV"])
    ws.append([None, 1, "EC/16.00001", "ABEDI ASSAD", "M", "(+243)895873021",
               "a@b.cd", "FICADEX", "Ancien", "OUI", "OUI", "INSCRIT", 120])
    ws.append([None, 2, "EC/18.00003", "ADRUPIAKO Emmanuel", "M", "(+243)818112782",
               "e@d.cd", "AUDIGEC", "Ancien", "OUI", "NON", "NON INSCRIT", 0])

    ws2 = wb.create_sheet("EC Indépendant 25")
    ws2.append([None] * 3)
    ws2.append([None, "CONSEIL PROVINCIAL"])
    ws2.append([None, "SECTION A2"])
    ws2.append([None, "N°", "N° D'ORDRE", "NOM, POST NOMS ET PRENOMS", "Sexe",
                "N° TELEPHONE", "E-MAIL", "NIF", "Ancienneté", "Cotisation",
                "Assurance Valide", "H.Formation", "C d'affaire", "Conclusion", "NHV"])
    ws2.append([None, 1, "EC/17.00002", "ABISA Lydie", "F", "(+243)859437431",
                "l@m.cd", "A1810696X", "Ancien", "OUI", "OUI", "OUI", "OUI", "INSCRIT", 122])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestExcelImport:
    def test_parse_multi_feuilles_et_categorie_par_feuille(self):
        rows, errors = parse_excel_bytes(_make_workbook_bytes(), "2026")
        assert errors == []
        assert len(rows) == 3
        cats = {r["categorie"] for r in rows}
        assert cats == {"EC Cabinet", "EC Indépendant"}

    def test_mapping_colonnes(self):
        rows, _ = parse_excel_bytes(_make_workbook_bytes(), "2026")
        r0 = next(r for r in rows if r["numero_ordre"] == "EC/16.00001")
        assert r0["nom"] == "ABEDI ASSAD"
        assert r0["cotisation_payee"] is True
        assert r0["heures_forco"] == 120
        assert r0["cabinet"] == "FICADEX"

    def test_champs_delibération_en_colonnes(self):
        rows, _ = parse_excel_bytes(_make_workbook_bytes(), "2026")
        indep = next(r for r in rows if r["categorie"] == "EC Indépendant")
        assert indep["assurance"] is True
        assert indep["chiffre_affaires"] is True
        assert indep["nif"] == "A1810696X"


class TestExporter:
    def test_export_par_section_avec_numerotation(self):
        rows, _ = parse_excel_bytes(_make_workbook_bytes(), "2026")
        content = build_workbook(rows, "2026", V.TableauReglages(), 2026)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        assert "EC EN CABINET" in wb.sheetnames
        ws = wb["EC EN CABINET"]
        # en-tête en ligne 5, première donnée ligne 6, N° séquentiel = 1
        assert ws.cell(5, 1).value == "N°"
        assert ws.cell(6, 1).value == 1

    def test_societe_sans_colonne_formation(self):
        wb = openpyxl.Workbook()
        content = build_workbook(
            [{"categorie": "Société", "numero_ordre": "SEC/18.00001", "nom": "ABN SAS",
              "cotisation_payee": True, "assurance": True,
              "raw_data": {"chiffre_affaires": True, "gerant": "X"}}],
            "2026", V.TableauReglages(), 2026,
        )
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb["SOCIETES"]
        headers = [ws.cell(5, c).value for c in range(1, ws.max_column + 1)]
        assert not any(h and "formation" in str(h).lower() for h in headers)
        assert "Chiffre d'affaires" in headers


class TestImportValidation:
    def test_ligne_valide_sans_avertissement(self):
        rows = [{"numero_ordre": "EC/16.00001", "nom": "Dupont", "categorie": "EC Cabinet"}]
        assert _valider_lignes(rows) == []

    def test_numero_ordre_manquant(self):
        rows = [{"numero_ordre": None, "nom": "Dupont", "categorie": "EC Cabinet"}]
        errs = _valider_lignes(rows)
        assert any(e["champ"] == "numero_ordre" for e in errs)

    def test_categorie_inconnue(self):
        rows = [{"numero_ordre": "EC/16.00001", "nom": "Dupont", "categorie": "Inconnu"}]
        errs = _valider_lignes(rows)
        assert any(e["champ"] == "categorie" for e in errs)


class TestComputeAnalyseStats:
    def test_stats_sur_dossiers_complets(self):
        dossiers = [_dossier(id=1), _dossier(id=2, nom="Martin", prenom="Alice")]
        stats = compute_analyse_stats(dossiers, detect_anomalies(dossiers))
        assert stats["total_dossiers"] == 2
        assert stats["anomalies_count"] == 0

    def test_stats_conclusions_dans_stats_json(self):
        dossiers = [_dossier(id=1), _dossier(id=2, nom="M", cotisation_payee=False)]
        verdicts = {d["id"]: V.evaluer(d, 2026) for d in dossiers}
        stats = compute_analyse_stats(dossiers, detect_anomalies(dossiers), verdicts)
        concl = stats["stats_json"]["conclusions"]
        assert concl["inscrits"] == 1
        assert concl["non_inscrits"] == 1


class TestCompareExercices:
    def test_dossiers_identiques_sont_en_commun(self):
        dossiers = [_dossier(id=1)]
        result = compare_exercices(dossiers, dossiers, "2025", "2026")
        assert result["dossiers_en_commun"] == 1
        assert result["nouveaux_dans_b"] == 0

    def test_nouveau_dossier_dans_exercice_b(self):
        a = [_dossier(id=1, nom="Dupont", prenom="Jean")]
        b = [_dossier(id=1, nom="Dupont", prenom="Jean"), _dossier(id=2, nom="Martin", prenom="Alice")]
        result = compare_exercices(a, b, "2025", "2026")
        assert result["nouveaux_dans_b"] == 1

    def test_changement_de_categorie_est_detecte(self):
        a = [_dossier(id=1, nom="Dupont", prenom="Jean", categorie="EC Salarié")]
        b = [_dossier(id=1, nom="Dupont", prenom="Jean", categorie="EC Indépendant")]
        result = compare_exercices(a, b, "2025", "2026")
        assert result["changements_categorie"] == 1

    def test_matching_ignore_casse_et_espaces(self):
        a = [_dossier(id=1, nom="  Dupont ", prenom=" Jean")]
        b = [_dossier(id=1, nom="DUPONT", prenom="jean")]
        result = compare_exercices(a, b, "2025", "2026")
        assert result["dossiers_en_commun"] == 1
