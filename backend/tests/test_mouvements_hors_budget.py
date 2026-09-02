from __future__ import annotations

import uuid
from decimal import Decimal

import pytest
from fastapi import BackgroundTasks, HTTPException
from sqlalchemy import select

from app.models.budget import BudgetExercice, BudgetPoste, StatutBudget
from app.models.caisse_centrale import CaisseCentrale
from app.models.compte_bancaire import CompteBancaire
from app.models.fonds_tiers_operation import FondsTiersOperation
from app.models.mouvement_budget_imputation import MouvementBudgetImputation
from app.models.ligne_requisition import LigneRequisition
from app.models.organisation import Organisation
from app.models.requisition import Requisition
from app.models.user import User
from app.models.regularisation_budgetaire import RegularisationBudgetaire
from app.services.fonds_tiers import resolve_fonds_tiers_display_name
from app.schemas.payment import AffecterBudgetPayload, BudgetAffectationLine, EncaissementCancelPayload, EncaissementCreate, FondsTiersCreate
from app.schemas.sortie_fonds import SortieFondsCreate, SortieFondsStatusUpdate

# Les écritures comptables sont générées dans le flux : sans cet import, les
# tables `compta_*` manquent aux métadonnées quand ce fichier tourne seul.
from app.modules.comptabilite import models as _compta_models  # noqa: F401


class _FakeRequest:
    headers: dict = {}
    client = None


async def _org(db, slug_prefix="hb"):
    org = Organisation(nom="Hors Budget", slug=f"{slug_prefix}-{uuid.uuid4().hex[:8]}", is_active=True)
    db.add(org)
    await db.flush()
    return org


async def _other_org(db, name="Conseil Provincial Test", active=True):
    org = Organisation(nom=name, slug=f"tiers-{uuid.uuid4().hex[:8]}", is_active=active)
    db.add(org)
    await db.flush()
    return org


async def _admin(db, org):
    user = User(id=uuid.uuid4(), email=f"{uuid.uuid4().hex[:8]}@ex.com", role="admin", organisation_id=org.id)
    db.add(user)
    await db.flush()
    return user


async def _poste(db, org, type_="RECETTE", montant_prevu=Decimal("10000")):
    exercice = BudgetExercice(organisation_id=org.id, annee=2026, statut=StatutBudget.BROUILLON)
    db.add(exercice)
    await db.flush()
    poste = BudgetPoste(
        organisation_id=org.id,
        exercice_id=exercice.id,
        code=f"{type_[:3]}-{uuid.uuid4().hex[:6]}",
        libelle=f"Poste {type_}",
        type=type_,
        active=True,
        montant_prevu=montant_prevu,
        montant_engage=0,
        montant_paye=0,
        is_deleted=False,
    )
    db.add(poste)
    await db.flush()
    return poste


async def _caisse(db, org, usd=Decimal("0")):
    caisse = CaisseCentrale(organisation_id=org.id, solde_usd=usd, solde_cdf=0, est_ouverte=True)
    db.add(caisse)
    await db.flush()
    return caisse


async def _banque(db, org, solde=Decimal("0")):
    compte = CompteBancaire(
        organisation_id=org.id,
        intitule="Equity",
        numero_compte=f"EQ-{uuid.uuid4().hex[:10]}",
        devise="USD",
        solde_initial=solde,
        solde_actuel=solde,
        is_active=True,
        account_type="BANK",
    )
    db.add(compte)
    await db.flush()
    return compte


async def _requisition_approuvee(
    db,
    org,
    *,
    nature="BUDGETAIRE",
    montant=Decimal("0"),
    mode_paiement="cash",
    poste=None,
    tiers_organisation_id=None,
    tiers_nom_libre=None,
    beneficiaire=None,
):
    """Source autorisée d'une sortie de fonds.

    Depuis la refonte des natures, la caisse n'ouvre plus de mouvement de sa
    propre initiative : toute sortie descend d'une réquisition approuvée (ou
    d'un ordre direct). Les tests doivent donc poser la source, comme le fait
    le circuit réel.
    """
    req = Requisition(
        organisation_id=org.id,
        numero_requisition=f"REQ-{uuid.uuid4().hex[:8]}",
        objet="Source de test",
        mode_paiement=mode_paiement,
        type_requisition="classique",
        nature_requisition=nature,
        status="APPROUVEE",
        montant_total=montant,
        devise="USD",
        beneficiaire=beneficiaire,
        tiers_organisation_id=tiers_organisation_id,
        tiers_nom_libre=tiers_nom_libre,
    )
    db.add(req)
    await db.flush()
    if poste is not None:
        # Une réquisition budgétaire porte l'imputation dans ses lignes : c'est
        # d'elles que la sortie tire son poste, jamais du payload de la caisse.
        db.add(
            LigneRequisition(
                requisition_id=req.id,
                organisation_id=org.id,
                rubrique="Test",
                description="Ligne de test",
                quantite=1,
                montant_unitaire=montant,
                montant_total=montant,
                budget_poste_id=poste.id,
            )
        )
        await db.flush()
    return req


@pytest.mark.asyncio
async def test_encaissement_hors_budget_affecte_tresorerie_pas_budget_puis_regularise(db_session, monkeypatch):
    db = db_session
    org = await _org(db)
    user = await _admin(db, org)
    poste = await _poste(db, org, "RECETTE")
    caisse = await _caisse(db, org, Decimal("0"))
    await db.commit()

    async def fake_recu(**_kwargs):
        return "REC-HB-1"

    monkeypatch.setattr("app.api.v1.endpoints.encaissements._generate_numero_recu", fake_recu)
    from app.api.v1.endpoints.encaissements import affecter_encaissement_budget, create_encaissement

    enc = await create_encaissement(
        payload=EncaissementCreate(
            type_client="client_externe",
            client_nom="CPK",
            libelle="Recette non prévue",
            montant=Decimal("2000"),
            montant_total=Decimal("2000"),
            montant_paye=Decimal("2000"),
            nature_mouvement="HORS_BUDGET_A_REGULARISER",
            mode_paiement="cash",
            canal="CAISSE",
        ),
        background_tasks=BackgroundTasks(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    await db.refresh(caisse)
    await db.refresh(poste)
    assert caisse.solde_usd == Decimal("2000")
    assert poste.montant_paye == Decimal("0")

    await affecter_encaissement_budget(
        encaissement_id=str(enc["id"]),
        payload=AffecterBudgetPayload(
            lignes=[BudgetAffectationLine(budget_poste_id=poste.id, montant=Decimal("1200"))],
            justification="Affectation validée",
            idempotency_key="reg-hb-1",
        ),
        request=_FakeRequest(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    await db.refresh(poste)
    assert poste.montant_paye == Decimal("1200")

    await affecter_encaissement_budget(
        encaissement_id=str(enc["id"]),
        payload=AffecterBudgetPayload(
            lignes=[BudgetAffectationLine(budget_poste_id=poste.id, montant=Decimal("1200"))],
            justification="Retry",
            idempotency_key="reg-hb-1",
        ),
        request=_FakeRequest(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    await db.refresh(poste)
    assert poste.montant_paye == Decimal("1200")


@pytest.mark.asyncio
async def test_fonds_tiers_remboursements_partiels_et_surremboursement_rejete(db_session, monkeypatch):
    db = db_session
    org = await _org(db, "ft")
    tiers_org = await _other_org(db, "CP Sud")
    user = await _admin(db, org)
    banque = await _banque(db, org, Decimal("0"))
    await db.commit()

    async def fake_recu(**_kwargs):
        return "REC-FT-1"

    monkeypatch.setattr("app.api.v1.endpoints.encaissements._generate_numero_recu", fake_recu)

    async def fake_num(*_args, **_kwargs):
        return f"PAY-{uuid.uuid4().hex[:8]}"

    monkeypatch.setattr("app.api.v1.endpoints.sorties_fonds.generate_document_number", fake_num)

    from app.api.v1.endpoints.encaissements import create_encaissement
    from app.api.v1.endpoints.sorties_fonds import create_sortie_fonds

    await create_encaissement(
        payload=EncaissementCreate(
            type_client="autre",
            client_nom=None,
            client_id=None,
            libelle="Fonds reçus pour compte de tiers",
            montant=Decimal("500"),
            montant_total=Decimal("500"),
            montant_paye=Decimal("500"),
            nature_mouvement="FONDS_DE_TIERS",
            mode_paiement="virement",
            canal="BANQUE",
            compte_bancaire_id=banque.id,
            fonds_tiers=FondsTiersCreate(tiers_organisation_id=tiers_org.id),
        ),
        background_tasks=BackgroundTasks(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    op = (await db.execute(select(FondsTiersOperation).where(FondsTiersOperation.organisation_id == org.id))).scalar_one()
    assert op.tiers_organisation_id == tiers_org.id
    assert op.tiers_nom_libre is None
    await db.refresh(banque)
    assert banque.solde_actuel == Decimal("500")

    # Un reversement partiel se rattache à la même réquisition : le reste dû y
    # est décompté paiement après paiement, exactement comme en budgétaire.
    req = await _requisition_approuvee(
        db,
        org,
        nature="FONDS_DE_TIERS",
        montant=Decimal("500"),
        mode_paiement="virement",
        tiers_organisation_id=tiers_org.id,
    )
    for amount in (Decimal("200"), Decimal("300")):
        await create_sortie_fonds(
            payload=SortieFondsCreate(
                type_sortie="autre",
                requisition_id=req.id,
                nature_mouvement="FONDS_DE_TIERS",
                fonds_tiers_operation_id=op.id,
                montant_paye=amount,
                mode_paiement="virement",
                devise="USD",
                canal="BANQUE",
                compte_bancaire_id=banque.id,
                motif="Remboursement tiers",
                beneficiaire="CP Sud",
            ),
            request=_FakeRequest(),
            background_tasks=BackgroundTasks(),
            user=user,
            tenant_id=org.id,
            db=db,
        )
    await db.refresh(op)
    await db.refresh(banque)
    assert op.statut == "REGULARISE"
    assert banque.solde_actuel == Decimal("0")

    # Source neuve et non soldée : le refus doit venir du fonds de tiers épuisé,
    # pas du reste dû de la réquisition précédente.
    req_trop = await _requisition_approuvee(
        db,
        org,
        nature="FONDS_DE_TIERS",
        montant=Decimal("1"),
        mode_paiement="virement",
        tiers_organisation_id=tiers_org.id,
    )
    with pytest.raises(HTTPException):
        await create_sortie_fonds(
            payload=SortieFondsCreate(
                type_sortie="autre",
                requisition_id=req_trop.id,
                nature_mouvement="FONDS_DE_TIERS",
                fonds_tiers_operation_id=op.id,
                montant_paye=Decimal("1"),
                mode_paiement="virement",
                devise="USD",
                canal="BANQUE",
                compte_bancaire_id=banque.id,
                motif="Trop",
                beneficiaire="CP Sud",
            ),
            request=_FakeRequest(),
            background_tasks=BackgroundTasks(),
            user=user,
            tenant_id=org.id,
            db=db,
        )


@pytest.mark.asyncio
async def test_fonds_tiers_identite_tenant_externe_et_legacy(db_session, monkeypatch):
    db = db_session
    org = await _org(db, "ft-ident")
    tiers_org = await _other_org(db, "Conseil Provincial du Kongo Central")
    user = await _admin(db, org)
    banque = await _banque(db, org, Decimal("0"))
    await db.commit()

    async def fake_recu(**_kwargs):
        return f"REC-FT-{uuid.uuid4().hex[:8]}"

    monkeypatch.setattr("app.api.v1.endpoints.encaissements._generate_numero_recu", fake_recu)

    from app.api.v1.endpoints.encaissements import create_encaissement

    await create_encaissement(
        payload=EncaissementCreate(
            type_client="client_externe",
            client_nom="Client",
            libelle="Fonds tenant",
            montant=Decimal("500"),
            montant_total=Decimal("500"),
            montant_paye=Decimal("500"),
            nature_mouvement="FONDS_DE_TIERS",
            mode_paiement="virement",
            canal="BANQUE",
            compte_bancaire_id=banque.id,
            fonds_tiers=FondsTiersCreate(tiers_organisation_id=tiers_org.id),
        ),
        background_tasks=BackgroundTasks(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    tenant_op = (
        await db.execute(
            select(FondsTiersOperation).where(FondsTiersOperation.organisation_id == org.id)
        )
    ).scalar_one()
    assert tenant_op.tiers_organisation_id == tiers_org.id
    assert tenant_op.tiers_nom_libre is None
    assert tenant_op.tiers_concerne is None
    assert await resolve_fonds_tiers_display_name(db, tenant_op) == (
        "Conseil Provincial du Kongo Central",
        "ORGANISATION",
    )

    await create_encaissement(
        payload=EncaissementCreate(
            type_client="client_externe",
            client_nom="Client",
            libelle="Fonds externe",
            montant=Decimal("300"),
            montant_total=Decimal("300"),
            montant_paye=Decimal("300"),
            nature_mouvement="FONDS_DE_TIERS",
            mode_paiement="virement",
            canal="BANQUE",
            compte_bancaire_id=banque.id,
            fonds_tiers=FondsTiersCreate(tiers_nom_libre=" Association ABC "),
        ),
        background_tasks=BackgroundTasks(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    external_op = (
        await db.execute(
            select(FondsTiersOperation)
            .where(FondsTiersOperation.organisation_id == org.id, FondsTiersOperation.tiers_nom_libre.is_not(None))
        )
    ).scalar_one()
    assert external_op.tiers_organisation_id is None
    assert external_op.tiers_nom_libre == "Association ABC"
    assert await resolve_fonds_tiers_display_name(db, external_op) == ("Association ABC", "EXTERNE")

    external_op.tiers_nom_libre = None
    external_op.tiers_concerne = "CP Kin legacy"
    await db.flush()
    assert await resolve_fonds_tiers_display_name(db, external_op) == ("CP Kin legacy", "LEGACY")


@pytest.mark.asyncio
async def test_fonds_tiers_reference_autre_tenant_sans_acces_aux_comptes(db_session, monkeypatch):
    from app.api.v1.endpoints.banques import list_comptes_bancaires
    from app.api.v1.endpoints.encaissements import create_encaissement

    db = db_session
    org_a = await _org(db, "iso-a")
    org_b = await _other_org(db, "Conseil Provincial isolé")
    user_a = await _admin(db, org_a)
    compte_a = await _banque(db, org_a, Decimal("0"))
    compte_b = await _banque(db, org_b, Decimal("0"))
    await db.commit()

    async def fake_recu(**_kwargs):
        return f"REC-ISO-{uuid.uuid4().hex[:8]}"

    monkeypatch.setattr("app.api.v1.endpoints.encaissements._generate_numero_recu", fake_recu)

    await create_encaissement(
        payload=EncaissementCreate(
            type_client="client_externe",
            client_nom="Client",
            libelle="Fonds tenant B",
            montant=Decimal("500"),
            montant_total=Decimal("500"),
            montant_paye=Decimal("500"),
            nature_mouvement="FONDS_DE_TIERS",
            mode_paiement="virement",
            canal="BANQUE",
            compte_bancaire_id=compte_a.id,
            fonds_tiers=FondsTiersCreate(tiers_organisation_id=org_b.id),
        ),
        background_tasks=BackgroundTasks(),
        user=user_a,
        tenant_id=org_a.id,
        db=db,
    )
    op = (
        await db.execute(select(FondsTiersOperation).where(FondsTiersOperation.organisation_id == org_a.id))
    ).scalar_one()

    assert op.organisation_id == org_a.id
    assert op.tiers_organisation_id == org_b.id

    comptes_visibles = await list_comptes_bancaires(
        active=True,
        banque_id=None,
        devise=None,
        account_type=None,
        tenant_id=org_a.id,
        user=user_a,
        db=db,
    )
    compte_ids = {compte.id for compte in comptes_visibles}

    assert compte_a.id in compte_ids
    assert compte_b.id not in compte_ids


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "fonds_tiers,expected_detail",
    [
        (FondsTiersCreate(tiers_organisation_id=999999), "Organisation tiers introuvable"),
        (FondsTiersCreate(), "tiers_organisation_id ou tiers_nom_libre requis"),
    ],
)
async def test_fonds_tiers_identite_rejets_simples(db_session, monkeypatch, fonds_tiers, expected_detail):
    db = db_session
    org = await _org(db, "ft-reject")
    user = await _admin(db, org)
    banque = await _banque(db, org, Decimal("0"))
    await db.commit()

    async def fake_recu(**_kwargs):
        return f"REC-FT-{uuid.uuid4().hex[:8]}"

    monkeypatch.setattr("app.api.v1.endpoints.encaissements._generate_numero_recu", fake_recu)

    from app.api.v1.endpoints.encaissements import create_encaissement

    with pytest.raises(HTTPException) as exc_info:
        await create_encaissement(
            payload=EncaissementCreate(
                type_client="client_externe",
                client_nom="Client",
                libelle="Fonds invalide",
                montant=Decimal("100"),
                montant_total=Decimal("100"),
                montant_paye=Decimal("100"),
                nature_mouvement="FONDS_DE_TIERS",
                mode_paiement="virement",
                canal="BANQUE",
                compte_bancaire_id=banque.id,
                fonds_tiers=fonds_tiers,
            ),
            background_tasks=BackgroundTasks(),
            user=user,
            tenant_id=org.id,
            db=db,
        )
    assert expected_detail in str(exc_info.value.detail)


@pytest.mark.asyncio
async def test_fonds_tiers_identite_rejette_double_inactive_et_tenant_courant(db_session, monkeypatch):
    db = db_session
    org = await _org(db, "ft-reject2")
    active_tiers = await _other_org(db, "Conseil actif")
    inactive_tiers = await _other_org(db, "Conseil inactif", active=False)
    user = await _admin(db, org)
    banque = await _banque(db, org, Decimal("0"))
    await db.commit()

    async def fake_recu(**_kwargs):
        return f"REC-FT-{uuid.uuid4().hex[:8]}"

    monkeypatch.setattr("app.api.v1.endpoints.encaissements._generate_numero_recu", fake_recu)

    from app.api.v1.endpoints.encaissements import create_encaissement

    async def create_with(fonds_tiers):
        return await create_encaissement(
            payload=EncaissementCreate(
                type_client="client_externe",
                client_nom="Client",
                libelle="Fonds invalide",
                montant=Decimal("100"),
                montant_total=Decimal("100"),
                montant_paye=Decimal("100"),
                nature_mouvement="FONDS_DE_TIERS",
                mode_paiement="virement",
                canal="BANQUE",
                compte_bancaire_id=banque.id,
                fonds_tiers=fonds_tiers,
            ),
            background_tasks=BackgroundTasks(),
            user=user,
            tenant_id=org.id,
            db=db,
        )

    with pytest.raises(HTTPException) as double_exc:
        await create_with(FondsTiersCreate(tiers_organisation_id=active_tiers.id, tiers_nom_libre="Association ABC"))
    assert "exclusifs" in str(double_exc.value.detail)

    with pytest.raises(HTTPException) as inactive_exc:
        await create_with(FondsTiersCreate(tiers_organisation_id=inactive_tiers.id))
    assert "inactive" in str(inactive_exc.value.detail)

    with pytest.raises(HTTPException) as current_exc:
        await create_with(FondsTiersCreate(tiers_organisation_id=org.id))
    assert "autre organisation" in str(current_exc.value.detail)


@pytest.mark.asyncio
async def test_remboursement_fonds_tiers_force_beneficiaire_depuis_operation(db_session, monkeypatch):
    db = db_session
    org = await _org(db, "ft-remb-benef")
    tiers_org = await _other_org(db, "Conseil Provincial du Haut-Katanga")
    user = await _admin(db, org)
    banque = await _banque(db, org, Decimal("0"))
    await db.commit()

    async def fake_recu(**_kwargs):
        return f"REC-FT-{uuid.uuid4().hex[:8]}"

    async def fake_num(*_args, **_kwargs):
        return f"PAY-{uuid.uuid4().hex[:8]}"

    monkeypatch.setattr("app.api.v1.endpoints.encaissements._generate_numero_recu", fake_recu)
    monkeypatch.setattr("app.api.v1.endpoints.sorties_fonds.generate_document_number", fake_num)

    from app.api.v1.endpoints.encaissements import create_encaissement
    from app.api.v1.endpoints.sorties_fonds import create_sortie_fonds

    await create_encaissement(
        payload=EncaissementCreate(
            type_client="client_externe",
            client_nom="Client",
            libelle="Fonds tenant",
            montant=Decimal("500"),
            montant_total=Decimal("500"),
            montant_paye=Decimal("500"),
            nature_mouvement="FONDS_DE_TIERS",
            mode_paiement="virement",
            canal="BANQUE",
            compte_bancaire_id=banque.id,
            fonds_tiers=FondsTiersCreate(tiers_organisation_id=tiers_org.id),
        ),
        background_tasks=BackgroundTasks(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    op = (await db.execute(select(FondsTiersOperation).where(FondsTiersOperation.organisation_id == org.id))).scalar_one()

    req = await _requisition_approuvee(
        db,
        org,
        nature="FONDS_DE_TIERS",
        montant=Decimal("200"),
        mode_paiement="virement",
        tiers_organisation_id=tiers_org.id,
    )
    sortie = await create_sortie_fonds(
        payload=SortieFondsCreate(
            type_sortie="autre",
            requisition_id=req.id,
            nature_mouvement="FONDS_DE_TIERS",
            fonds_tiers_operation_id=op.id,
            montant_paye=Decimal("200"),
            mode_paiement="virement",
            devise="USD",
            canal="BANQUE",
            compte_bancaire_id=banque.id,
            motif="Remboursement tiers",
            beneficiaire="Bénéficiaire falsifié",
        ),
        request=_FakeRequest(),
        background_tasks=BackgroundTasks(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    assert sortie.beneficiaire == "Conseil Provincial du Haut-Katanga"
    assert sortie.fonds_tiers_operation_id == op.id
    assert op.tiers_organisation_id == tiers_org.id
    assert op.tiers_nom_libre is None


@pytest.mark.asyncio
async def test_annulation_sortie_utilise_imputations_persistantes(db_session, monkeypatch):
    db = db_session
    org = await _org(db, "imp")
    user = await _admin(db, org)
    poste = await _poste(db, org, "DEPENSE")
    caisse = await _caisse(db, org, Decimal("500"))
    await db.commit()

    async def fake_num(*_args, **_kwargs):
        return f"PAY-{uuid.uuid4().hex[:8]}"

    monkeypatch.setattr("app.api.v1.endpoints.sorties_fonds.generate_document_number", fake_num)
    from app.api.v1.endpoints.sorties_fonds import create_sortie_fonds, update_sortie_statut

    req = await _requisition_approuvee(
        db, org, montant=Decimal("120"), poste=poste
    )
    sortie = await create_sortie_fonds(
        payload=SortieFondsCreate(
            type_sortie="autre",
            requisition_id=req.id,
            montant_paye=Decimal("120"),
            mode_paiement="cash",
            devise="USD",
            canal="CAISSE",
            motif="Dépense",
            beneficiaire="Fournisseur",
            budget_poste_id=poste.id,
        ),
        request=_FakeRequest(),
        background_tasks=BackgroundTasks(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    imps = (await db.execute(select(MouvementBudgetImputation).where(MouvementBudgetImputation.sortie_fonds_id == sortie.id))).scalars().all()
    assert len(imps) == 1

    await update_sortie_statut(
        sortie_id=str(sortie.id),
        payload=SortieFondsStatusUpdate(statut="ANNULEE", motif_annulation="Erreur"),
        request=_FakeRequest(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    await db.refresh(poste)
    await db.refresh(caisse)
    assert poste.montant_paye == Decimal("0")
    assert caisse.solde_usd == Decimal("500")


@pytest.mark.asyncio
async def test_sortie_hors_budget_est_affectee_puis_reprise_a_l_annulation(db_session, monkeypatch):
    db = db_session
    org = await _org(db, "sfhb")
    user = await _admin(db, org)
    poste = await _poste(db, org, "DEPENSE")
    caisse = await _caisse(db, org, Decimal("500"))
    await db.commit()

    async def fake_num(*_args, **_kwargs):
        return f"PAY-{uuid.uuid4().hex[:8]}"

    monkeypatch.setattr("app.api.v1.endpoints.sorties_fonds.generate_document_number", fake_num)
    from app.api.v1.endpoints.sorties_fonds import affecter_sortie_budget, create_sortie_fonds, update_sortie_statut

    # Le hors budget est désormais une nature de réquisition : c'est elle qui
    # autorise la dépense, la régularisation venant ensuite.
    req = await _requisition_approuvee(
        db, org, nature="HORS_BUDGET", montant=Decimal("120")
    )
    sortie = await create_sortie_fonds(
        payload=SortieFondsCreate(
            type_sortie="autre",
            requisition_id=req.id,
            nature_mouvement="HORS_BUDGET_A_REGULARISER",
            montant_paye=Decimal("120"),
            mode_paiement="cash",
            devise="USD",
            canal="CAISSE",
            motif="Urgence non budgétée",
            beneficiaire="Fournisseur",
        ),
        request=_FakeRequest(),
        background_tasks=BackgroundTasks(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    await db.refresh(poste)
    await db.refresh(caisse)
    # La caisse a bougé, le budget non.
    assert caisse.solde_usd == Decimal("380")
    assert poste.montant_paye == Decimal("0")
    assert sortie.hors_budget_status == "A_REGULARISER"

    await affecter_sortie_budget(
        sortie_id=str(sortie.id),
        payload=AffecterBudgetPayload(
            lignes=[BudgetAffectationLine(budget_poste_id=poste.id, montant=Decimal("120"))],
            justification="Imputée après coup sur le poste dépense",
            idempotency_key="reg-sf-1",
        ),
        request=_FakeRequest(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    await db.refresh(poste)
    assert poste.montant_paye == Decimal("120")
    reg = (
        await db.execute(select(RegularisationBudgetaire).where(RegularisationBudgetaire.sortie_fonds_id == sortie.id))
    ).scalar_one()
    assert reg.ancien_nature_mouvement == "HORS_BUDGET_A_REGULARISER"

    # Rejouer la même clé n'impute pas deux fois.
    await affecter_sortie_budget(
        sortie_id=str(sortie.id),
        payload=AffecterBudgetPayload(
            lignes=[BudgetAffectationLine(budget_poste_id=poste.id, montant=Decimal("120"))],
            justification="Retry",
            idempotency_key="reg-sf-1",
        ),
        request=_FakeRequest(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    await db.refresh(poste)
    assert poste.montant_paye == Decimal("120")

    await update_sortie_statut(
        sortie_id=str(sortie.id),
        payload=SortieFondsStatusUpdate(statut="ANNULEE", motif_annulation="Erreur"),
        request=_FakeRequest(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    await db.refresh(poste)
    await db.refresh(caisse)
    assert poste.montant_paye == Decimal("0")
    assert caisse.solde_usd == Decimal("500")


@pytest.mark.asyncio
async def test_annulation_encaissement_regularise_rend_le_budget(db_session, monkeypatch):
    """Un encaissement hors budget régularisé puis annulé ne doit rien laisser au poste.

    L'imputation vient de la régularisation (portée par l'encaissement), pas du
    paiement : elle doit être reprise même quand l'encaissement a des paiements
    actifs à annuler.
    """
    db = db_session
    org = await _org(db, "regann")
    user = await _admin(db, org)
    poste = await _poste(db, org, "RECETTE")
    caisse = await _caisse(db, org, Decimal("0"))
    await db.commit()

    async def fake_recu(**_kwargs):
        return f"REC-{uuid.uuid4().hex[:6]}"

    monkeypatch.setattr("app.api.v1.endpoints.encaissements._generate_numero_recu", fake_recu)
    from app.api.v1.endpoints.encaissements import (
        affecter_encaissement_budget,
        cancel_encaissement_operation,
        create_encaissement,
    )

    enc = await create_encaissement(
        payload=EncaissementCreate(
            type_client="client_externe",
            client_nom="CPK",
            libelle="Recette non prévue",
            montant=Decimal("800"),
            montant_total=Decimal("800"),
            montant_paye=Decimal("800"),
            nature_mouvement="HORS_BUDGET_A_REGULARISER",
            mode_paiement="cash",
            canal="CAISSE",
        ),
        background_tasks=BackgroundTasks(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    await affecter_encaissement_budget(
        encaissement_id=str(enc["id"]),
        payload=AffecterBudgetPayload(
            lignes=[BudgetAffectationLine(budget_poste_id=poste.id, montant=Decimal("800"))],
            justification="Affectation validée",
            idempotency_key="reg-ann-1",
        ),
        request=_FakeRequest(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    await db.refresh(poste)
    assert poste.montant_paye == Decimal("800")

    await cancel_encaissement_operation(
        encaissement_id=str(enc["id"]),
        payload=EncaissementCancelPayload(motif_annulation="Encaissement erroné"),
        request=_FakeRequest(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    await db.refresh(poste)
    await db.refresh(caisse)
    assert poste.montant_paye == Decimal("0")
    assert caisse.solde_usd == Decimal("0")
    restantes = (
        await db.execute(
            select(MouvementBudgetImputation).where(
                MouvementBudgetImputation.organisation_id == org.id,
                MouvementBudgetImputation.statut == "ACTIVE",
            )
        )
    ).scalars().all()
    assert restantes == []


@pytest.mark.asyncio
async def test_dashboard_separe_tresorerie_et_budget(db_session, monkeypatch):
    """Le hors budget gonfle la trésorerie, jamais l'exécution budgétaire.

    C'est toute la raison d'être de la séparation : un encaissement hors budget
    est bien de l'argent reçu — il doit apparaître en trésorerie — mais il n'a
    alimenté aucun poste, et le dire dans les mêmes chiffres ferait croire à
    une recette budgétaire qui n'existe pas.
    """
    db = db_session
    org = await _org(db, "dash")
    user = await _admin(db, org)
    poste = await _poste(db, org, "RECETTE")
    await _caisse(db, org, Decimal("0"))
    await db.commit()

    async def fake_recu(**_kwargs):
        return f"REC-{uuid.uuid4().hex[:6]}"

    monkeypatch.setattr("app.api.v1.endpoints.encaissements._generate_numero_recu", fake_recu)
    # Le cache Redis servirait la réponse d'un autre test : on le neutralise.
    monkeypatch.setattr("app.api.v1.endpoints.dashboard.cache_get", lambda *_a, **_k: _none())
    monkeypatch.setattr("app.api.v1.endpoints.dashboard.cache_set", lambda *_a, **_k: _none())

    from app.api.v1.endpoints.dashboard import stats as dashboard_stats
    from app.api.v1.endpoints.encaissements import create_encaissement

    async def _encaisser(nature: str, montant: Decimal, poste_id: int | None):
        return await create_encaissement(
            payload=EncaissementCreate(
                type_client="client_externe",
                client_nom="CPK",
                libelle="Recette",
                montant=montant,
                montant_total=montant,
                montant_paye=montant,
                nature_mouvement=nature,
                budget_poste_id=poste_id,
                mode_paiement="cash",
                canal="CAISSE",
            ),
            background_tasks=BackgroundTasks(),
            user=user,
            tenant_id=org.id,
            db=db,
        )

    await _encaisser("BUDGETAIRE", Decimal("300"), poste.id)
    await _encaisser("HORS_BUDGET_A_REGULARISER", Decimal("700"), None)
    await db.commit()

    res = await dashboard_stats(
        period_type="month",
        tenant_id=org.id,
        user=user,
        db=db,
    )
    # Trésorerie : tout l'argent reçu.
    assert Decimal(res.stats.total_encaissements_period) == Decimal("1000.00")
    # Budget : seulement ce qui a alimenté un poste.
    assert Decimal(res.stats.total_recettes_budgetaires_period) == Decimal("300.00")
    assert Decimal(res.stats.total_recettes_hors_budget_period) == Decimal("700.00")
    # Et l'encours qui appelle une décision.
    assert Decimal(res.stats.hors_budget_a_regulariser_montant) == Decimal("700.00")
    assert res.stats.hors_budget_a_regulariser_count == 1


async def _none():
    return None


@pytest.mark.asyncio
async def test_export_encaissements_porte_la_nature_budgetaire(db_session, monkeypatch):
    """L'export nomme ce que chaque ligne fait au budget.

    Sans cette colonne, une recette budgétaire et un fonds de tiers se
    ressemblent trait pour trait dans le classeur : même montant, même date,
    même client.
    """
    from io import BytesIO

    from openpyxl import load_workbook

    db = db_session
    org = await _org(db, "exp")
    user = await _admin(db, org)
    banque = await _banque(db, org, Decimal("0"))
    await db.commit()

    async def fake_recu(**_kwargs):
        return f"REC-{uuid.uuid4().hex[:6]}"

    monkeypatch.setattr("app.api.v1.endpoints.encaissements._generate_numero_recu", fake_recu)
    from app.api.v1.endpoints.encaissements import create_encaissement
    from app.api.v1.endpoints.exports import construire_classeur_encaissements

    await create_encaissement(
        payload=EncaissementCreate(
            type_client="client_externe",
            client_nom="CP Sud",
            libelle="Fonds pour compte de tiers",
            montant=Decimal("250"),
            montant_total=Decimal("250"),
            montant_paye=Decimal("250"),
            nature_mouvement="FONDS_DE_TIERS",
            mode_paiement="virement",
            canal="BANQUE",
            compte_bancaire_id=banque.id,
            fonds_tiers=FondsTiersCreate(tiers_nom_libre="CP Sud"),
        ),
        background_tasks=BackgroundTasks(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    await db.commit()

    classeur, _nom = await construire_classeur_encaissements(db, org.id)
    buffer = BytesIO()
    classeur.save(buffer)
    buffer.seek(0)
    ws = load_workbook(buffer, data_only=False)["Encaissements"]
    lignes = list(ws.iter_rows(values_only=True))

    entete = next(row for row in lignes if "Nature budgétaire" in [str(v) for v in row if v])
    index_nature = [str(v) for v in entete].index("Nature budgétaire")
    index_impact = [str(v) for v in entete].index("Impact budgétaire")
    index_tiers = [str(v) for v in entete].index("Tiers / Conseil concerné")
    ligne = next(row for row in lignes if row[index_nature] == "Fonds de tiers")
    assert ligne[index_nature] == "Fonds de tiers"
    assert ligne[index_impact] == "Non"
    assert ligne[index_tiers] == "CP Sud"


@pytest.mark.asyncio
async def test_export_budget_ignore_fonds_tiers_meme_avec_poste_residuel(db_session):
    """L'export budget ne doit pas transformer un poste résiduel en recette réalisée."""
    from io import BytesIO

    from openpyxl import load_workbook

    from app.api.v1.endpoints.exports import construire_classeur_budget
    from app.models.encaissement import Encaissement

    db = db_session
    org = await _org(db, "exp-budget-ft")
    poste = await _poste(db, org, "RECETTE", Decimal("1000"))
    db.add(
        Encaissement(
            organisation_id=org.id,
            type_client="client_externe",
            client_nom="Tiers",
            libelle="Fonds de tiers legacy avec poste",
            montant=Decimal("500"),
            montant_total=Decimal("500"),
            montant_paye=Decimal("500"),
            montant_percu=Decimal("500"),
            devise_perception="USD",
            canal="CAISSE",
            statut_paiement="complet",
            mode_paiement="cash",
            est_proforma=False,
            is_deleted=False,
            statut_operation="ACTIVE",
            nature_mouvement="FONDS_DE_TIERS",
            impact_budgetaire=False,
            budget_poste_id=poste.id,
            budget_poste_code=poste.code,
            budget_poste_libelle=poste.libelle,
        )
    )
    await db.commit()

    classeur, _nom = await construire_classeur_budget(db, org.id, annee=2026, type="RECETTE")
    buffer = BytesIO()
    classeur.save(buffer)
    buffer.seek(0)
    ws = load_workbook(buffer, data_only=False)["Budget 2026"]
    rows = list(ws.iter_rows(values_only=True))
    row = next(r for r in rows if r[0] == poste.code)

    assert row[9] == 0


@pytest.mark.asyncio
async def test_export_sorties_fonds_affiche_tiers_sans_impact_budgetaire(db_session, monkeypatch):
    from io import BytesIO

    from openpyxl import load_workbook

    from app.api.v1.endpoints.encaissements import create_encaissement
    from app.api.v1.endpoints.exports import construire_classeur_sorties_fonds
    from app.api.v1.endpoints.sorties_fonds import create_sortie_fonds

    db = db_session
    org = await _org(db, "exp-sort-ft")
    user = await _admin(db, org)
    banque = await _banque(db, org, Decimal("0"))
    await db.commit()

    async def fake_recu(**_kwargs):
        return f"REC-FT-{uuid.uuid4().hex[:8]}"

    async def fake_num(*_args, **_kwargs):
        return f"PAY-{uuid.uuid4().hex[:8]}"

    monkeypatch.setattr("app.api.v1.endpoints.encaissements._generate_numero_recu", fake_recu)
    monkeypatch.setattr("app.api.v1.endpoints.sorties_fonds.generate_document_number", fake_num)

    await create_encaissement(
        payload=EncaissementCreate(
            type_client="client_externe",
            client_nom="Client",
            libelle="Fonds externe",
            montant=Decimal("500"),
            montant_total=Decimal("500"),
            montant_paye=Decimal("500"),
            nature_mouvement="FONDS_DE_TIERS",
            mode_paiement="virement",
            canal="BANQUE",
            compte_bancaire_id=banque.id,
            fonds_tiers=FondsTiersCreate(tiers_nom_libre="Association ABC"),
        ),
        background_tasks=BackgroundTasks(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    op = (await db.execute(select(FondsTiersOperation).where(FondsTiersOperation.organisation_id == org.id))).scalar_one()
    req = await _requisition_approuvee(
        db,
        org,
        nature="FONDS_DE_TIERS",
        montant=Decimal("200"),
        mode_paiement="virement",
        tiers_organisation_id=op.tiers_organisation_id,
        tiers_nom_libre=op.tiers_nom_libre,
    )
    await create_sortie_fonds(
        payload=SortieFondsCreate(
            type_sortie="autre",
            requisition_id=req.id,
            nature_mouvement="FONDS_DE_TIERS",
            fonds_tiers_operation_id=op.id,
            montant_paye=Decimal("200"),
            mode_paiement="virement",
            devise="USD",
            canal="BANQUE",
            compte_bancaire_id=banque.id,
            motif="Remboursement tiers",
            beneficiaire="Ignoré",
        ),
        request=_FakeRequest(),
        background_tasks=BackgroundTasks(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    await db.commit()

    classeur, _nom = await construire_classeur_sorties_fonds(db, org.id)
    buffer = BytesIO()
    classeur.save(buffer)
    buffer.seek(0)
    ws = load_workbook(buffer, data_only=False)["Sorties"]
    lignes = list(ws.iter_rows(values_only=True))

    entete = next(row for row in lignes if "Nature budgétaire" in [str(v) for v in row if v])
    header = [str(v) for v in entete]
    index_nature = header.index("Nature budgétaire")
    index_impact = header.index("Impact budgétaire")
    index_tiers = header.index("Tiers / Conseil concerné")
    ligne = next(row for row in lignes if row[index_nature] == "Fonds de tiers")

    assert ligne[index_impact] == "Non"
    assert ligne[index_tiers] == "Association ABC"


@pytest.mark.asyncio
async def test_dashboard_separe_hors_budget_et_fonds_de_tiers(db_session, monkeypatch):
    """Un remboursement de fonds de tiers sort de trésorerie sans devenir hors budget."""
    from app.api.v1.endpoints.dashboard import stats as dashboard_stats
    from app.api.v1.endpoints.encaissements import create_encaissement
    from app.api.v1.endpoints.sorties_fonds import create_sortie_fonds

    db = db_session
    org = await _org(db, "dash-ft")
    user = await _admin(db, org)
    banque = await _banque(db, org, Decimal("0"))
    await db.commit()

    async def fake_recu(**_kwargs):
        return f"REC-FT-{uuid.uuid4().hex[:8]}"

    async def fake_num(*_args, **_kwargs):
        return f"PAY-{uuid.uuid4().hex[:8]}"

    monkeypatch.setattr("app.api.v1.endpoints.encaissements._generate_numero_recu", fake_recu)
    monkeypatch.setattr("app.api.v1.endpoints.sorties_fonds.generate_document_number", fake_num)

    await create_encaissement(
        payload=EncaissementCreate(
            type_client="client_externe",
            client_nom="Client",
            libelle="Fonds externe",
            montant=Decimal("500"),
            montant_total=Decimal("500"),
            montant_paye=Decimal("500"),
            nature_mouvement="FONDS_DE_TIERS",
            mode_paiement="virement",
            canal="BANQUE",
            compte_bancaire_id=banque.id,
            fonds_tiers=FondsTiersCreate(tiers_nom_libre="Association ABC"),
        ),
        background_tasks=BackgroundTasks(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    op = (await db.execute(select(FondsTiersOperation).where(FondsTiersOperation.organisation_id == org.id))).scalar_one()
    req = await _requisition_approuvee(
        db,
        org,
        nature="FONDS_DE_TIERS",
        montant=Decimal("200"),
        mode_paiement="virement",
        tiers_organisation_id=op.tiers_organisation_id,
        tiers_nom_libre=op.tiers_nom_libre,
    )
    await create_sortie_fonds(
        payload=SortieFondsCreate(
            type_sortie="autre",
            requisition_id=req.id,
            nature_mouvement="FONDS_DE_TIERS",
            fonds_tiers_operation_id=op.id,
            montant_paye=Decimal("200"),
            mode_paiement="virement",
            devise="USD",
            canal="BANQUE",
            compte_bancaire_id=banque.id,
            motif="Remboursement tiers",
            beneficiaire="Ignoré",
        ),
        request=_FakeRequest(),
        background_tasks=BackgroundTasks(),
        user=user,
        tenant_id=org.id,
        db=db,
    )
    await db.commit()

    res = await dashboard_stats(
        period_type="month",
        tenant_id=org.id,
        user=user,
        db=db,
    )

    assert Decimal(res.stats.total_encaissements_period) == Decimal("500.00")
    assert Decimal(res.stats.total_sorties_period) == Decimal("200.00")
    assert Decimal(res.stats.total_recettes_hors_budget_period) == Decimal("0.00")
    assert Decimal(res.stats.total_depenses_hors_budget_period) == Decimal("0.00")
    assert Decimal(res.stats.fonds_tiers_solde) == Decimal("300.00")
    assert res.stats.fonds_tiers_count == 1


@pytest.mark.asyncio
async def test_requisition_sans_impact_budgetaire_franchit_le_circuit_sans_lignes(db_session):
    """Le circuit de validation ne doit pas exiger de lignes là où il n'y en a pas.

    Hors budget et fonds de tiers autorisent un montant en bloc, sans imputation
    initiale : leur réclamer une ligne les bloquerait en BROUILLON, donc les
    rendrait impossibles à approuver — et sans réquisition approuvée, la sortie
    correspondante n'existe pas non plus. C'est le montant autorisé qui prend le
    relais comme garde-fou.
    """
    db = db_session
    org = await _org(db, "req-sans-lignes")
    from app.services.requisition_service import require_requisition_lines

    for nature in ("HORS_BUDGET", "FONDS_DE_TIERS"):
        req = await _requisition_approuvee(db, org, nature=nature, montant=Decimal("120"))
        await require_requisition_lines(db, req)  # ne lève pas

    # Sans montant, la réquisition n'autorise rien : le garde-fou reprend la main.
    req_vide = await _requisition_approuvee(db, org, nature="HORS_BUDGET", montant=Decimal("0"))
    with pytest.raises(HTTPException) as exc:
        await require_requisition_lines(db, req_vide)
    assert "Montant autorisé" in str(exc.value.detail)

    # La réquisition budgétaire, elle, reste portée par ses lignes.
    req_budget = await _requisition_approuvee(db, org, montant=Decimal("120"))
    with pytest.raises(HTTPException) as exc_budget:
        await require_requisition_lines(db, req_budget)
    assert "Aucune ligne" in str(exc_budget.value.detail)


@pytest.mark.asyncio
async def test_sortie_de_fonds_exige_une_source_autorisee(db_session, monkeypatch):
    """La caisse exécute, elle n'ouvre pas le mouvement.

    Sans réquisition approuvée ni ordre direct, la sortie n'a pas d'autorité
    derrière elle : elle est refusée, même parfaitement formée par ailleurs.
    """
    db = db_session
    org = await _org(db, "sortie-sans-source")
    poste = await _poste(db, org, type_="DEPENSE", montant_prevu=Decimal("1000"))
    await _caisse(db, org, usd=Decimal("500"))
    user = await _admin(db, org)
    await db.commit()

    async def fake_num(*a, **k):
        return f"PAY-{uuid.uuid4().hex[:8]}"

    monkeypatch.setattr("app.api.v1.endpoints.sorties_fonds.generate_document_number", fake_num)
    from app.api.v1.endpoints.sorties_fonds import create_sortie_fonds

    with pytest.raises(HTTPException) as exc:
        await create_sortie_fonds(
            payload=SortieFondsCreate(
                type_sortie="autre",
                montant_paye=Decimal("120"),
                mode_paiement="cash",
                devise="USD",
                canal="CAISSE",
                motif="Dépense sans source",
                beneficiaire="Fournisseur",
                budget_poste_id=poste.id,
            ),
            request=_FakeRequest(),
            background_tasks=BackgroundTasks(),
            user=user,
            tenant_id=org.id,
            db=db,
        )
    assert exc.value.status_code == 400
    assert "Réquisition approuvée ou ordre direct" in str(exc.value.detail)
