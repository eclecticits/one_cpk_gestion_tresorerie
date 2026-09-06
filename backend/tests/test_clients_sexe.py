"""Le sexe du client : saisie, fiche, export.

Le champ est demandé à la saisie d'un encaissement (personne physique et
client externe) mais il ne se range pas sur l'encaissement : il vit sur la
fiche client, comme l'email et le téléphone.
"""

import uuid
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from app.api.v1.endpoints.exports import MONEY, export_encaissements
from app.api.v1.endpoints.encaissements import _resolve_or_create_client
from app.models.caisse_centrale import CaisseCentrale
from app.models.client import Client
from app.models.encaissement import Encaissement
from app.models.expert_comptable import ExpertComptable
from app.models.organisation import Organisation
from app.models.user import User
from app.schemas.payment import EncaissementCreate


def _payload(**overrides) -> EncaissementCreate:
    """Charge utile minimale acceptée par EncaissementCreate."""
    base = dict(
        type_client="personne_physique",
        client_nom="Mwamba Kabila",
        libelle="Frais de dossier",
        montant=Decimal("50"),
        montant_total=Decimal("50"),
    )
    base.update(overrides)
    return EncaissementCreate(**base)


# --------------------------------------------------------------------------
# Normalisation
# --------------------------------------------------------------------------

@pytest.mark.parametrize(
    "saisi,attendu",
    [
        ("M", "M"),
        ("f", "F"),
        ("Masculin", "M"),
        ("FÉMININ", "F"),
        ("  m  ", "M"),
        ("", None),
        (None, None),
    ],
)
def test_le_sexe_se_ramene_a_m_ou_f(saisi, attendu):
    assert _payload(client_sexe=saisi).client_sexe == attendu


def test_un_sexe_hors_nomenclature_est_refuse_a_la_porte():
    # Plutôt qu'une violation de la contrainte de table, illisible pour
    # l'appelant, l'erreur remonte du schéma.
    with pytest.raises(ValidationError, match="Sexe invalide"):
        _payload(client_sexe="Non binaire")


# --------------------------------------------------------------------------
# La fiche client
# --------------------------------------------------------------------------

async def _organisation(db_session) -> Organisation:
    org = Organisation(
        nom="Sexe Client",
        slug=f"sexe-client-{uuid.uuid4().hex[:8]}",
        is_active=True,
    )
    db_session.add(org)
    await db_session.flush()
    return org


@pytest.mark.asyncio
async def test_la_fiche_creee_porte_le_sexe(db_session):
    org = await _organisation(db_session)
    payload = _payload(client_sexe="F", client_nom=f"Nouvelle Cliente {uuid.uuid4().hex[:6]}")

    client_id = await _resolve_or_create_client(db_session, org.id, payload, None)

    client = await db_session.get(Client, client_id)
    assert client.sexe == "F"


@pytest.mark.asyncio
async def test_une_fiche_sans_sexe_se_complete_au_passage_suivant(db_session):
    org = await _organisation(db_session)
    nom = f"Client Incomplet {uuid.uuid4().hex[:6]}"
    db_session.add(Client(organisation_id=org.id, nom=nom, active=True))
    await db_session.flush()

    client_id = await _resolve_or_create_client(
        db_session, org.id, _payload(client_nom=nom, client_sexe="M"), None
    )

    client = await db_session.get(Client, client_id)
    assert client.sexe == "M"


@pytest.mark.asyncio
async def test_un_sexe_deja_su_ne_se_fait_pas_ecraser(db_session):
    """Même règle que l'email et le téléphone : on complète, on n'écrase pas.

    Une faute de frappe au guichet ne doit pas réécrire une fiche établie ;
    la correction passe par l'écran du référentiel clients, qui, lui, écrase.
    """
    org = await _organisation(db_session)
    nom = f"Client Etabli {uuid.uuid4().hex[:6]}"
    db_session.add(Client(organisation_id=org.id, nom=nom, sexe="F", active=True))
    await db_session.flush()

    await _resolve_or_create_client(
        db_session, org.id, _payload(client_nom=nom, client_sexe="M"), None
    )

    client = (await db_session.execute(
        Client.__table__.select().where(Client.nom == nom)
    )).first()
    assert client.sexe == "F"


# --------------------------------------------------------------------------
# L'export
# --------------------------------------------------------------------------

async def _streaming_response_bytes(response) -> bytes:
    body = b""
    async for chunk in response.body_iterator:
        body += chunk if isinstance(chunk, bytes) else chunk.encode()
    return body


@pytest.mark.asyncio
async def test_l_export_porte_le_sexe_et_garde_ses_colonnes_de_montants(db_session):
    org = await _organisation(db_session)
    user = User(
        id=uuid.uuid4(),
        email=f"sexe-export-{uuid.uuid4().hex[:6]}@example.com",
        role="admin",
        organisation_id=org.id,
    )
    client = Client(
        organisation_id=org.id,
        nom=f"Kabeya Ngoy {uuid.uuid4().hex[:6]}",
        type_client="personne_physique",
        sexe="F",
        active=True,
    )
    expert = ExpertComptable(
        numero_ordre=f"EC-{uuid.uuid4().hex[:8]}",
        nom_denomination="Cabinet Sans Sexe",
        active=True,
    )
    db_session.add_all([
        user,
        client,
        expert,
        CaisseCentrale(organisation_id=org.id, solde_usd=Decimal("1000"), est_ouverte=True),
    ])
    await db_session.flush()

    commun = dict(
        organisation_id=org.id,
        libelle="Frais",
        montant=Decimal("250.00"),
        montant_total=Decimal("250.00"),
        montant_paye=Decimal("100.00"),
        montant_percu=Decimal("100.00"),
        devise_perception="USD",
        canal="CAISSE",
        statut_paiement="partiel",
        mode_paiement="cash",
        est_proforma=False,
        is_deleted=False,
        statut_operation="ACTIVE",
        date_encaissement=datetime(2026, 8, 28, 13, 0, tzinfo=timezone.utc),
    )
    db_session.add_all([
        Encaissement(
            type_client="personne_physique",
            client_nom=client.nom,
            client_id=client.id,
            **commun,
        ),
        Encaissement(type_client="expert_comptable", expert_comptable_id=expert.id, **commun),
    ])
    await db_session.commit()

    response = await export_encaissements(
        date_debut="2026-01-01",
        date_fin="2026-08-28",
        statut_paiement=None,
        numero_recu=None,
        client=None,
        budget_poste_id=None,
        type_client=None,
        mode_paiement=None,
        expert_comptable_id=None,
        deleted_status="all",
        est_proforma=False,
        user=user,
        db=db_session,
    )
    workbook = load_workbook(BytesIO(await _streaming_response_bytes(response)), data_only=False)
    ws = workbook["Encaissements"]
    rows = list(ws.iter_rows(values_only=True))

    header_row = next(row for row in rows if "N° Note de débit" in row)
    col = {nom: index for index, nom in enumerate(header_row) if nom}
    assert "Sexe" in col

    ligne = next(row for row in rows if row[col["Client"]] == client.nom)
    assert ligne[col["Sexe"]] == "F"

    # L'expert-comptable relève d'un autre référentiel, qui ne porte pas le
    # sexe : la case reste vide plutôt que de mentir.
    ligne_expert = next(
        row for row in rows
        if isinstance(row[col["Client"]], str) and row[col["Client"]].endswith("Cabinet Sans Sexe")
    )
    # openpyxl relit une cellule vide en None, jamais en "" : les deux disent
    # la même chose ici.
    assert ligne_expert[col["Sexe"]] in (None, "")

    # Le garde-fou du décalage : money_cols et total_values reperent leurs
    # colonnes par INDEX, et « Sexe » s'insère avant elles. Si le report est
    # oublié, le format monétaire glisse d'un cran — ce que ces deux
    # assertions attrapent, là où une comparaison de valeurs ne verrait rien.
    entete_excel = next(
        idx for idx, row in enumerate(ws.iter_rows(values_only=True), start=1)
        if "N° Note de débit" in row
    )
    ligne_excel = next(
        idx for idx, row in enumerate(ws.iter_rows(values_only=True), start=1)
        if row[col["Client"]] == client.nom
    )
    montant_total = ws.cell(row=ligne_excel, column=col["Montant total (USD)"] + 1)
    mode_paiement = ws.cell(row=ligne_excel, column=col["Mode de paiement"] + 1)
    assert ws.cell(row=entete_excel, column=col["Montant total (USD)"] + 1).value == "Montant total (USD)"
    assert montant_total.number_format == MONEY
    assert mode_paiement.number_format != MONEY
