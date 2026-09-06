import uuid
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.api.v1.endpoints.exports import export_encaissements
from app.models.banque import Banque
from app.models.caisse_centrale import CaisseCentrale
from app.models.compte_bancaire import CompteBancaire
from app.models.client import Client
from app.models.encaissement import Encaissement
from app.models.expert_comptable import ExpertComptable
from app.models.organisation import Organisation
from app.models.sortie_fonds import SortieFonds
from app.models.user import User


async def _streaming_response_bytes(response) -> bytes:
    body = b""
    async for chunk in response.body_iterator:
        body += chunk if isinstance(chunk, bytes) else chunk.encode()
    return body


@pytest.mark.asyncio
async def test_export_encaissements_affiche_versement_banque_hors_totaux(db_session):
    org = Organisation(
        nom="Export Transferts",
        slug=f"export-transferts-{uuid.uuid4().hex[:8]}",
        is_active=True,
    )
    db_session.add(org)
    await db_session.flush()
    user = User(
        id=uuid.uuid4(),
        email=f"export-transferts-{uuid.uuid4().hex[:6]}@example.com",
        role="admin",
        organisation_id=org.id,
    )
    banque = Banque(organisation_id=org.id, nom="Equity BCDC", code="EQUITY", is_active=True)
    client = Client(
        organisation_id=org.id,
        nom="Client réel",
        email="client@example.com",
        telephone="+243 900 000 001",
        active=True,
    )
    expert = ExpertComptable(
        numero_ordre=f"EC-{uuid.uuid4().hex[:8]}",
        nom_denomination="Cabinet Expert",
        email="expert@example.com",
        telephone="+243 900 000 002",
        active=True,
    )
    db_session.add_all([
        user,
        banque,
        client,
        expert,
        CaisseCentrale(organisation_id=org.id, solde_usd=Decimal("1000"), est_ouverte=True),
    ])
    await db_session.flush()
    compte = CompteBancaire(
        organisation_id=org.id,
        banque_id=banque.id,
        intitule="Compte Equity",
        numero_compte="EQ-USD-001",
        devise="USD",
        solde_initial=Decimal("5000"),
        solde_actuel=Decimal("8000"),
        is_active=True,
        account_type="BANK",
    )
    db_session.add(compte)
    await db_session.flush()
    enc = Encaissement(
        organisation_id=org.id,
        type_client="client_externe",
        client_nom="Client réel",
        client_id=client.id,
        libelle="Recette économique",
        montant=Decimal("500.00"),
        montant_total=Decimal("500.00"),
        montant_paye=Decimal("500.00"),
        montant_percu=Decimal("500.00"),
        devise_perception="USD",
        canal="CAISSE",
        statut_paiement="complet",
        mode_paiement="cash",
        est_proforma=False,
        is_deleted=False,
        statut_operation="ACTIVE",
        date_encaissement=datetime(2026, 8, 28, 13, 0, tzinfo=timezone.utc),
    )
    sortie = SortieFonds(
        organisation_id=org.id,
        type_sortie="versement_banque",
        montant_paye=Decimal("3000.00"),
        mode_paiement="cash",
        devise="USD",
        canal="CAISSE",
        compte_bancaire_id=compte.id,
        motif="Dépôt des recettes journalières à la banque",
        beneficiaire="Equity BCDC",
        statut="VALIDE",
        date_paiement=datetime(2026, 8, 28, 15, 54, tzinfo=timezone.utc),
        created_by=user.id,
        reference_numero="PAY-CENTRAL-2026-00003",
    )
    enc_expert = Encaissement(
        organisation_id=org.id,
        type_client="expert_comptable",
        expert_comptable_id=expert.id,
        libelle="Cotisation expert",
        montant=Decimal("100.00"),
        montant_total=Decimal("100.00"),
        montant_paye=Decimal("100.00"),
        montant_percu=Decimal("100.00"),
        devise_perception="USD",
        canal="CAISSE",
        statut_paiement="complet",
        mode_paiement="cash",
        est_proforma=False,
        is_deleted=False,
        statut_operation="ACTIVE",
        date_encaissement=datetime(2026, 8, 27, 13, 0, tzinfo=timezone.utc),
    )
    db_session.add_all([enc, enc_expert, sortie])
    await db_session.commit()

    response = await export_encaissements(
        date_debut="2026-01-01",
        date_fin="2026-08-28",
        statut_paiement=None,
        numero_recu=None,
        client=None,
        budget_poste_id=None,
        type_client=None,
        mode_paiement=None,
        expert_comptable_id=None,
        deleted_status="all",
        est_proforma=False,
        user=user,
        db=db_session,
    )
    workbook = load_workbook(BytesIO(await _streaming_response_bytes(response)), data_only=False)
    ws = workbook["Encaissements"]
    rows = list(ws.iter_rows(values_only=True))

    # Repérage par nom d'en-tête plutôt que par position : ajouter une colonne
    # au classeur ne doit pas faire échouer un test qui ne parle pas d'elle.
    header_row = next(row for row in rows if "N° Note de débit" in row)
    col = {nom: index for index, nom in enumerate(header_row) if nom}

    versement_row = next(
        row for row in rows
        if any(value == "PAY-CENTRAL-2026-00003" for value in row)
    )
    assert versement_row[col["Type d'opération"]] == "Versement banque"
    assert versement_row[col["Source / Mode"]] == "Caisse"
    assert versement_row[col["Banque source"]] == "Equity BCDC"
    assert versement_row[col["Compte bancaire"]] == "EQ-USD-001"
    # La colonne « Nature budgétaire » suit le poste : elle dit d'un transfert
    # interne qu'il ne consomme aucun budget, ce que le libellé seul laissait
    # deviner.
    assert versement_row[col["Nature budgétaire"]] == "Transfert interne"
    assert versement_row[col["Description"]] == "Transfert interne caisse → banque (entrée bancaire)"
    assert Decimal(str(versement_row[col["Montant perçu"]])) == Decimal("3000")

    enc_row = next(row for row in rows if row[col["Client"]] == "Client réel")
    assert enc_row[col["Email client"]] == "client@example.com"
    assert enc_row[col["Téléphone client"]] == "+243 900 000 001"
    expert_row = next(
        row for row in rows
        if isinstance(row[col["Client"]], str) and row[col["Client"]].endswith("Cabinet Expert")
    )
    assert expert_row[col["Email client"]] == "expert@example.com"
    assert expert_row[col["Téléphone client"]] == "+243 900 000 002"
    assert versement_row[col["Mode de paiement"]] == "Transfert interne"

    total_row = next(row for row in rows if row[0] == "TOTAL")
    assert Decimal(str(total_row[col["Montant total (USD)"]])) == Decimal("600")
    assert Decimal(str(total_row[col["Montant payé (USD)"]])) == Decimal("600")
    assert Decimal(str(total_row[col["Reste à payer (USD)"]])) == Decimal("0")
