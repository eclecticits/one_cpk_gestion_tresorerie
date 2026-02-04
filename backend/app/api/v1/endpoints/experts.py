from __future__ import annotations

import logging
import re
import uuid
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from sqlalchemy import func, or_, select, update
from sqlalchemy.exc import DataError, IntegrityError, ProgrammingError
from sqlalchemy.ext.asyncio import AsyncSession
from io import BytesIO

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.expert_comptable import ExpertComptable
from app.models.category_changes_history import CategoryChangesHistory
from app.models.imports_history import ImportsHistory
from app.models.user import User
from app.schemas.expert import (
    CategoryChangeRequest,
    CategoryChangeResponse,
    ExpertComptableCreate,
    ExpertComptableResponse,
    ExpertComptableUpdate,
    ExpertsListResponse,
    ExpertImportRow,
    ExpertImportRequest,
    ExpertImportResponse,
)

router = APIRouter()
logger = logging.getLogger(__name__)


def _normalize_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _coerce_json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, dict):
        return {str(k): _coerce_json_value(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_coerce_json_value(v) for v in value]
    return str(value)


def _normalize_phone(raw: str) -> str | None:
    if not raw:
        return None
    raw = raw.strip()
    if not raw:
        return None
    has_plus = raw.startswith("+")
    digits = "".join(ch for ch in raw if ch.isdigit())
    if not digits:
        return None
    if has_plus:
        return f"+{digits}"
    if digits.startswith("0") and len(digits) == 10:
        return f"+243{digits[1:]}"
    if len(digits) == 9:
        return f"+243{digits}"
    if digits.startswith("243"):
        return f"+{digits}"
    return None


def _normalize_email(raw: str) -> str | None:
    if not raw:
        return None
    value = raw.strip().lower()
    return value or None


def _is_valid_email(value: str) -> bool:
    return re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", value) is not None


def _read_excel_rows(file_bytes: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook  # lazy import
    except Exception:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="openpyxl not installed")
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows: list[dict] = []
    for row in rows[1:]:
        row_dict = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row_dict[header] = row[idx] if idx < len(row) else None
        data_rows.append(row_dict)
    return data_rows


def _row_to_import_row(category: str, row: dict) -> ExpertImportRow:
    base = {
        "numero_ordre": _normalize_value(row.get("N° d'ordre")),
        "email": _normalize_value(row.get("E-mail")),
        "telephone": _normalize_value(row.get("N° de téléphone")),
    }

    if category == "sec":
        return ExpertImportRow(
            **base,
            nom_denomination=_normalize_value(row.get("Dénomination")),
            type_ec="SEC",
            categorie_personne="Personne Morale",
            statut_professionnel="Cabinet",
            raison_sociale=_normalize_value(row.get("Raison sociale")),
            associe_gerant=_normalize_value(row.get("Associé gérant")),
        )
    if category == "en_cabinet":
        return ExpertImportRow(
            **base,
            nom_denomination=_normalize_value(row.get("Noms")),
            type_ec="EC",
            categorie_personne="Personne Physique",
            statut_professionnel="En Cabinet",
            sexe=_normalize_value(row.get("Sexe")).upper(),
            cabinet_attache=_normalize_value(row.get("Cabinet d'attache")),
        )
    if category == "independant":
        return ExpertImportRow(
            **base,
            nom_denomination=_normalize_value(row.get("Noms")),
            type_ec="EC",
            categorie_personne="Personne Physique",
            statut_professionnel="Indépendant",
            sexe=_normalize_value(row.get("Sexe")).upper(),
            nif=_normalize_value(row.get("NIF")),
        )
    # salarie
    return ExpertImportRow(
        **base,
        nom_denomination=_normalize_value(row.get("Noms")),
        type_ec="EC",
        categorie_personne="Personne Physique",
        statut_professionnel="Salarié",
        sexe=_normalize_value(row.get("Sexe")).upper(),
        nom_employeur=_normalize_value(row.get("Nom de l'employeur")),
    )


def _get_category_from_expert(expert: ExpertComptable) -> str | None:
    """Détermine la catégorie à partir des champs de l'expert."""
    if expert.type_ec == "SEC":
        return "sec"
    if expert.statut_professionnel == "En Cabinet":
        return "en_cabinet"
    if expert.statut_professionnel == "Indépendant":
        return "independant"
    if expert.statut_professionnel == "Salarié":
        return "salarie"
    return None


def _expert_to_response(expert: ExpertComptable) -> dict[str, Any]:
    """Convertit un modèle Expert en dict pour la réponse."""
    return {
        "id": str(expert.id),
        "numero_ordre": expert.numero_ordre,
        "nom_denomination": expert.nom_denomination,
        "type_ec": expert.type_ec,
        "categorie_personne": expert.categorie_personne,
        "statut_professionnel": expert.statut_professionnel,
        "sexe": expert.sexe,
        "telephone": expert.telephone,
        "email": expert.email,
        "nif": expert.nif,
        "cabinet_attache": expert.cabinet_attache,
        "nom_employeur": expert.nom_employeur,
        "raison_sociale": expert.raison_sociale,
        "associe_gerant": expert.associe_gerant,
        "import_id": str(expert.import_id) if expert.import_id else None,
        "active": expert.active,
        "created_at": expert.created_at,
    }


def _parse_order(order: str | None):
    if not order:
        return ExpertComptable.numero_ordre.asc()
    parts = order.split(".")
    field = parts[0]
    direction = parts[1] if len(parts) > 1 else "asc"
    column_map = {
        "numero_ordre": ExpertComptable.numero_ordre,
        "nom_denomination": ExpertComptable.nom_denomination,
        "created_at": ExpertComptable.created_at,
        "statut_professionnel": ExpertComptable.statut_professionnel,
    }
    col = column_map.get(field)
    if col is None:
        return ExpertComptable.numero_ordre.asc()
    return col.desc() if direction.lower() == "desc" else col.asc()


@router.get("", response_model=list[ExpertComptableResponse] | ExpertsListResponse)
async def list_experts(
    numero_ordre: str | None = Query(default=None, description="Recherche exacte par numéro d'ordre"),
    nom: str | None = Query(default=None, description="Recherche partielle par nom"),
    q: str | None = Query(default=None, description="Recherche globale (nom, numéro, email, cabinet)"),
    type_ec: str | None = Query(default=None, description="Filtrer par type (EC ou SEC)"),
    active: bool | None = Query(default=True, description="Filtrer par statut actif"),
    include_inactive: bool = Query(default=False, description="Inclure les experts inactifs"),
    statut_professionnel: str | None = Query(default=None),
    order: str | None = Query(default=None, description="Ex: numero_ordre.asc"),
    limit: int = Query(default=50, le=200),
    offset: int = Query(default=0, ge=0),
    include_summary: bool = Query(default=False),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    """Liste des experts comptables avec filtres."""
    query = select(ExpertComptable)
    conditions = []

    if numero_ordre:
        conditions.append(ExpertComptable.numero_ordre == numero_ordre.strip())
    if nom:
        conditions.append(ExpertComptable.nom_denomination.ilike(f"%{nom}%"))
    if q:
        q_value = f"%{q.strip()}%"
        conditions.append(
            or_(
                ExpertComptable.numero_ordre.ilike(q_value),
                ExpertComptable.nom_denomination.ilike(q_value),
                ExpertComptable.email.ilike(q_value),
                ExpertComptable.cabinet_attache.ilike(q_value),
            )
        )
    if type_ec:
        conditions.append(ExpertComptable.type_ec == type_ec)
    if statut_professionnel:
        conditions.append(ExpertComptable.statut_professionnel == statut_professionnel)
    if not include_inactive and active is not None:
        conditions.append(ExpertComptable.active == active)

    if conditions:
        query = query.where(*conditions)

    query = query.order_by(_parse_order(order)).offset(offset).limit(limit)

    result = await db.execute(query)
    experts = result.scalars().all()

    items = [_expert_to_response(e) for e in experts]

    if not include_summary:
        return items

    count_query = select(func.count()).select_from(ExpertComptable)
    if conditions:
        count_query = count_query.where(*conditions)
    total_count = int((await db.execute(count_query)).scalar_one() or 0)

    return ExpertsListResponse(items=items, total=total_count)


@router.post("", response_model=ExpertComptableResponse, status_code=status.HTTP_201_CREATED)
async def create_expert(
    payload: ExpertComptableCreate,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Crée un expert comptable."""
    numero_ordre = payload.numero_ordre.strip()
    if not numero_ordre:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="numero_ordre requis")

    res = await db.execute(select(ExpertComptable).where(ExpertComptable.numero_ordre == numero_ordre))
    existing = res.scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="numero_ordre déjà existant")

    expert = ExpertComptable(
        numero_ordre=numero_ordre,
        nom_denomination=payload.nom_denomination.strip(),
        type_ec=payload.type_ec or "EC",
        categorie_personne=payload.categorie_personne,
        statut_professionnel=payload.statut_professionnel,
        sexe=payload.sexe,
        telephone=payload.telephone,
        email=payload.email,
        nif=payload.nif,
        cabinet_attache=payload.cabinet_attache,
        nom_employeur=payload.nom_employeur,
        raison_sociale=payload.raison_sociale,
        associe_gerant=payload.associe_gerant,
        active=True,
    )
    db.add(expert)
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        logger.exception("Create expert failed: numero_ordre already exists (%s)", numero_ordre)
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="numero_ordre déjà existant")
    await db.refresh(expert)

    return _expert_to_response(expert)


@router.get("/{expert_id}", response_model=ExpertComptableResponse)
async def get_expert(
    expert_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Récupère un expert comptable par son ID."""
    try:
        uid = uuid.UUID(expert_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid UUID")

    result = await db.execute(select(ExpertComptable).where(ExpertComptable.id == uid))
    expert = result.scalar_one_or_none()

    if not expert:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Expert non trouvé")

    return _expert_to_response(expert)


@router.patch("/{expert_id}", response_model=ExpertComptableResponse)
async def update_expert(
    expert_id: str,
    payload: ExpertComptableUpdate,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Met à jour un expert comptable (PATCH partiel)."""
    try:
        uid = uuid.UUID(expert_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid UUID")

    result = await db.execute(select(ExpertComptable).where(ExpertComptable.id == uid))
    expert = result.scalar_one_or_none()

    if not expert:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Expert non trouvé")

    # Mise à jour partielle
    update_data = payload.model_dump(exclude_unset=True)
    if update_data:
        for key, value in update_data.items():
            setattr(expert, key, value)
        await db.commit()
        await db.refresh(expert)

    return _expert_to_response(expert)


@router.put("/{expert_id}", response_model=ExpertComptableResponse)
async def replace_expert(
    expert_id: str,
    payload: ExpertComptableUpdate,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Met à jour un expert comptable (PUT)."""
    try:
        uid = uuid.UUID(expert_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid UUID")

    result = await db.execute(select(ExpertComptable).where(ExpertComptable.id == uid))
    expert = result.scalar_one_or_none()

    if not expert:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Expert non trouvé")

    update_data = payload.model_dump(exclude_unset=True)
    if not update_data:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Aucune donnée à mettre à jour")

    for key, value in update_data.items():
        setattr(expert, key, value)
    await db.commit()
    await db.refresh(expert)

    return _expert_to_response(expert)


@router.delete("/{expert_id}")
async def delete_expert(
    expert_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Archive un expert comptable (suppression logique)."""
    try:
        uid = uuid.UUID(expert_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid UUID")

    result = await db.execute(select(ExpertComptable).where(ExpertComptable.id == uid))
    expert = result.scalar_one_or_none()

    if not expert:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Expert non trouvé")

    expert.active = False
    await db.commit()

    return {"message": "Expert archivé avec succès. L'historique est préservé."}


async def _import_experts_payload(
    payload: ExpertImportRequest,
    user: User,
    db: AsyncSession,
) -> ExpertImportResponse:
    if not payload.rows:
        return ExpertImportResponse(
            success=False,
            imported=0,
            message="Aucune ligne à importer"
        )

    # Créer l'enregistrement d'import (optionnel si table absente ou colonne trop courte)
    import_record: ImportsHistory | None = None
    try:
        safe_filename = (payload.filename or "").strip()[:300]
        safe_file_data = (
            _coerce_json_value(payload.file_data) if payload.file_data is not None else None
        )
        import_record = ImportsHistory(
            filename=safe_filename or "import.xlsx",
            category=str(payload.category)[:50],
            imported_by=user.id,
            rows_imported=0,
            status="success",
            file_data=safe_file_data,
        )
        db.add(import_record)
        await db.flush()
    except (ProgrammingError, DataError) as exc:
        await db.rollback()
        logger.warning(
            "Import history skipped (table missing or data too large): %s",
            exc,
        )
        import_record = None

    imported_count = 0
    created_count = 0
    updated_count = 0
    skipped_count = 0
    errors: list[dict] = []
    phone_warnings: list[str] = []
    total_rows = len(payload.rows)
    for idx, row in enumerate(payload.rows):
        row_data = {k: _normalize_value(v) for k, v in row.model_dump().items()}
        numero_ordre = row_data.get("numero_ordre", "").strip()
        if not numero_ordre:
            skipped_count += 1
            errors.append({
                "ligne": idx + 2,
                "champ": "numero_ordre",
                "message": "N° d'ordre manquant",
            })
            continue
        normalized_phone = _normalize_phone(row_data.get("telephone", ""))
        if row_data.get("telephone") and not normalized_phone:
            phone_warnings.append(numero_ordre or row_data.get("nom_denomination", "inconnu"))
            errors.append({
                "ligne": idx + 2,
                "champ": "telephone",
                "message": "Téléphone invalide (ignoré)",
            })
            logger.warning("Import experts: invalid phone at line %s (numero_ordre=%s)", idx + 2, numero_ordre)
        row_data["telephone"] = normalized_phone or ""
        email_value = _normalize_email(row_data.get("email", "")) or ""
        if email_value and not _is_valid_email(email_value):
            errors.append({
                "ligne": idx + 2,
                "champ": "email",
                "message": "Format e-mail invalide",
            })
            email_value = ""
            logger.warning("Import experts: invalid email at line %s (numero_ordre=%s)", idx + 2, numero_ordre)
        row_data["email"] = email_value
        # Vérifier si l'expert existe déjà (upsert)
        result = await db.execute(
            select(ExpertComptable).where(ExpertComptable.numero_ordre == numero_ordre)
        )
        existing = result.scalar_one_or_none()

        if existing:
            # Update
            for key, value in row_data.items():
                if key == "numero_ordre":
                    continue
                if value != "":
                    setattr(existing, key, value)
            if import_record:
                existing.import_id = import_record.id
            updated_count += 1
        else:
            # Insert
            new_expert = ExpertComptable(
                numero_ordre=row_data.get("numero_ordre", ""),
                nom_denomination=row_data.get("nom_denomination", ""),
                type_ec=row_data.get("type_ec", "EC") or "EC",
                categorie_personne=row_data.get("categorie_personne") or None,
                statut_professionnel=row_data.get("statut_professionnel") or None,
                sexe=row_data.get("sexe") or None,
                telephone=row_data.get("telephone") or None,
                email=row_data.get("email") or None,
                nif=row_data.get("nif") or None,
                cabinet_attache=row_data.get("cabinet_attache") or None,
                nom_employeur=row_data.get("nom_employeur") or None,
                raison_sociale=row_data.get("raison_sociale") or None,
                associe_gerant=row_data.get("associe_gerant") or None,
                import_id=import_record.id if import_record else None,
            )
            db.add(new_expert)
            created_count += 1

        imported_count += 1

    # Mettre à jour le nombre importé
    if import_record:
        import_record.rows_imported = imported_count
    await db.commit()

    logger.info(
        "Import experts: total=%s created=%s updated=%s skipped=%s errors=%s",
        total_rows,
        created_count,
        updated_count,
        skipped_count,
        len(errors),
    )

    message = f"{created_count} expert(s)-comptable(s) importé(s) avec succès"
    if phone_warnings:
        sample = ", ".join(phone_warnings[:5])
        suffix = f" | Téléphones invalides ignorés: {len(phone_warnings)}"
        if sample:
            suffix += f" (ex: {sample})"
        message += suffix

    return ExpertImportResponse(
        success=True,
        imported=created_count,
        updated=updated_count,
        skipped=skipped_count,
        total_lignes=total_rows,
        errors=errors,
        import_id=str(import_record.id) if import_record else None,
        message=message
    )


@router.post("/import", response_model=ExpertImportResponse)
async def import_experts(
    request: Request,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ExpertImportResponse:
    """Import batch d'experts comptables depuis Excel (JSON ou multipart)."""
    content_type = (request.headers.get("content-type") or "").lower()
    try:
        if "multipart/form-data" in content_type:
            form = await request.form()
            upload = form.get("file")
            category = form.get("category")
            filename = form.get("filename")
            if upload is None or category is None:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="file et category requis")
            file_bytes = await upload.read()
            rows = _read_excel_rows(file_bytes)
            mapped_rows = [_row_to_import_row(str(category), r) for r in rows]
            payload = ExpertImportRequest(
                category=str(category),
                filename=filename or getattr(upload, "filename", "import.xlsx"),
                rows=mapped_rows,
                file_data=rows,
            )
            return await _import_experts_payload(payload, user, db)

        data = await request.json()
        payload = ExpertImportRequest.model_validate(data)
        return await _import_experts_payload(payload, user, db)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Import experts failed")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Import failed: {exc}") from exc


@router.post("/category-change", response_model=CategoryChangeResponse)
async def change_category(
    payload: CategoryChangeRequest,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Change la catégorie d'un expert comptable avec historisation."""
    try:
        expert_uid = uuid.UUID(payload.expert_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid expert_id UUID")

    result = await db.execute(select(ExpertComptable).where(ExpertComptable.id == expert_uid))
    expert = result.scalar_one_or_none()

    if not expert:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Expert non trouvé")

    # Catégorie actuelle
    old_category = _get_category_from_expert(expert)

    # Vérifier que la nouvelle catégorie est différente
    if old_category == payload.new_category:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La nouvelle catégorie doit être différente de l'actuelle"
        )

    # Sauvegarder les anciennes données
    old_data = {
        "type_ec": expert.type_ec,
        "statut_professionnel": expert.statut_professionnel,
        "nif": expert.nif,
        "cabinet_attache": expert.cabinet_attache,
        "nom_employeur": expert.nom_employeur,
        "raison_sociale": expert.raison_sociale,
        "associe_gerant": expert.associe_gerant,
    }

    # Préparer les nouvelles données selon la catégorie
    new_data: dict[str, Any] = {
        "nif": None,
        "cabinet_attache": None,
        "nom_employeur": None,
        "raison_sociale": None,
        "associe_gerant": None,
    }

    if payload.new_category == "sec":
        new_data["type_ec"] = "SEC"
        new_data["categorie_personne"] = "Personne Morale"
        new_data["statut_professionnel"] = None
        new_data["raison_sociale"] = payload.raison_sociale
        new_data["associe_gerant"] = payload.associe_gerant
    elif payload.new_category == "en_cabinet":
        new_data["type_ec"] = "EC"
        new_data["categorie_personne"] = "Personne Physique"
        new_data["statut_professionnel"] = "En Cabinet"
        new_data["cabinet_attache"] = payload.cabinet_attache
    elif payload.new_category == "independant":
        new_data["type_ec"] = "EC"
        new_data["categorie_personne"] = "Personne Physique"
        new_data["statut_professionnel"] = "Indépendant"
        new_data["nif"] = payload.nif
    elif payload.new_category == "salarie":
        new_data["type_ec"] = "EC"
        new_data["categorie_personne"] = "Personne Physique"
        new_data["statut_professionnel"] = "Salarié"
        new_data["nom_employeur"] = payload.nom_employeur

    # Appliquer les changements à l'expert
    for key, value in new_data.items():
        setattr(expert, key, value)

    # Créer l'historique
    history = CategoryChangesHistory(
        expert_id=expert.id,
        numero_ordre=expert.numero_ordre,
        old_category=old_category,
        new_category=payload.new_category,
        changed_by=user.id,
        reason=payload.reason,
        old_data=old_data,
        new_data=new_data,
    )
    db.add(history)

    await db.commit()
    await db.refresh(history)

    return {
        "id": str(history.id),
        "expert_id": str(history.expert_id),
        "numero_ordre": history.numero_ordre,
        "old_category": history.old_category,
        "new_category": history.new_category,
        "changed_by": str(history.changed_by) if history.changed_by else None,
        "reason": history.reason,
        "old_data": history.old_data,
        "new_data": history.new_data,
        "created_at": history.created_at,
    }


@router.get("/category-changes/{expert_id}", response_model=list[CategoryChangeResponse])
async def get_category_changes(
    expert_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    """Historique des changements de catégorie pour un expert."""
    try:
        uid = uuid.UUID(expert_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid UUID")

    result = await db.execute(
        select(CategoryChangesHistory)
        .where(CategoryChangesHistory.expert_id == uid)
        .order_by(CategoryChangesHistory.created_at.desc())
    )
    changes = result.scalars().all()

    return [
        {
            "id": str(c.id),
            "expert_id": str(c.expert_id),
            "numero_ordre": c.numero_ordre,
            "old_category": c.old_category,
            "new_category": c.new_category,
            "changed_by": str(c.changed_by) if c.changed_by else None,
            "reason": c.reason,
            "old_data": c.old_data,
            "new_data": c.new_data,
            "created_at": c.created_at,
        }
        for c in changes
    ]
