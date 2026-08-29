from __future__ import annotations

from datetime import datetime, timezone
import uuid
import os
import io
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, Query, Request, File, UploadFile, status
from fastapi.responses import FileResponse
from fastapi.responses import StreamingResponse
from sqlalchemy import case, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font as XlFont

from app.api.deps import get_current_tenant_id, get_current_user, has_permission, get_public_tenant_id
from app.db.session import get_db
from app.utils.excel_io import save_workbook
from app.models.cloture_caisse import ClotureCaisse
from app.models.encaissement import Encaissement
from app.models.caisse_centrale import CaisseCentrale
from app.models.organisation import Organisation
from app.models.ouverture_caisse import OuvertureCaisse
from app.models.print_settings import PrintSettings
from app.models.regularisation_caisse import RegularisationCaisse
from app.models.retour_caisse import RetourCaisse
from app.models.sortie_fonds import SortieFonds
from app.models.transfert_interne import TransfertInterne
from app.models.user import User
from app.services.regularisation_caisse import (
    SOURCE_CLOTURE,
    SOURCE_OUVERTURE,
    EcartCaisse,
    RegularisationImpossible,
    regulariser_ecart,
)
from app.schemas.cloture import ClotureBalanceResponse, ClotureCreateRequest, ClotureOut, CloturePdfData, CloturePdfDetail, EcartRegularisationRequest, OuvertureCreateRequest, OuvertureOut
from app.core.config import settings
from app.services.audit_service import get_request_ip, log_action
from app.services.document_sequences import generate_document_number
from app.services.entrees_caisse import list_entrees_internes_caisse
from app.services.tenant_identity import tenant_display_name

router = APIRouter()

PDF_ALLOWED_TYPES = {"application/pdf"}
PDF_ALLOWED_EXT = {".pdf"}
CLTURE_PDF_DIR = os.path.abspath(os.path.join(settings.upload_dir, "clotures"))


def _decimal(value: Decimal | int | float | None) -> Decimal:
    return Decimal(value or 0).quantize(Decimal("0.01"))


async def _appliquer_regularisations(
    db: AsyncSession,
    *,
    tenant_id: int,
    user_id,
    demandee: bool,
    motif: str | None,
    source_type: str,
    source_id: int,
    source_reference: str | None,
    ecarts: list[EcartCaisse],
) -> tuple[list[dict], list[str]]:
    """Régularise les écarts non nuls, sans jamais bloquer la caisse.

    Renvoie (régularisations créées, messages d'échec). Un échec — paramétrage
    manquant, motif absent — n'interrompt ni l'ouverture ni la clôture : l'écart
    reste simplement ouvert et le message est remonté à l'utilisateur.
    """
    creees: list[dict] = []
    erreurs: list[str] = []
    if not demandee:
        return creees, erreurs

    for ecart in ecarts:
        if ecart.sens is None:
            continue
        try:
            regularisation = await regulariser_ecart(
                db,
                tenant_id=tenant_id,
                user_id=user_id,
                ecart=ecart,
                motif=motif or "",
                source_type=source_type,
                source_id=source_id,
                source_reference=source_reference,
            )
        except RegularisationImpossible as exc:
            erreurs.append(f"{ecart.devise} : {exc.message}")
            continue
        creees.append(
            {
                "devise": regularisation.devise,
                "sens": regularisation.sens,
                "montant": str(regularisation.montant),
                "encaissement_id": str(regularisation.encaissement_id)
                if regularisation.encaissement_id
                else None,
                "sortie_fonds_id": str(regularisation.sortie_fonds_id)
                if regularisation.sortie_fonds_id
                else None,
            }
        )
    return creees, erreurs


async def _compute_balance(db: AsyncSession, tenant_id: int) -> ClotureBalanceResponse:
    """Balance de caisse de l'organisation `tenant_id`.

    Le filtrage sur `organisation_id` est explicite ici, et non délégué au hook
    ORM de `app/db/session.py` : ce hook est inerte quand le contexte tenant vaut
    `None` (cas d'un super-admin sur l'hôte d'administration, cf.
    `deps.py:322-325` et `deps.py:446`). Sans ces filtres, la balance agrégeait
    alors les mouvements de TOUS les tenants — et ces montants sont persistés en
    base par `create_cloture`.
    """
    last_res = await db.execute(
        select(ClotureCaisse)
        .where(ClotureCaisse.organisation_id == tenant_id)
        .order_by(ClotureCaisse.date_cloture.desc())
        .limit(1)
    )
    last = last_res.scalar_one_or_none()

    date_debut = last.date_cloture if last else None
    date_fin = datetime.now(timezone.utc)

    # Report de la clôture précédente : on repart du solde THÉORIQUE, corrigé des
    # seules régularisations réellement enregistrées. Repartir du solde physique
    # compté absorberait en silence un écart que l'utilisateur a justement
    # choisi de ne pas régulariser — exactement ce que la règle métier interdit.
    # Les opérations de régularisation portent l'horodatage de la clôture et sont
    # donc exclues des flux de la période suivante (bornes en « > » strict).
    async def _report(devise: str, theorique) -> Decimal:
        if last is None:
            return _decimal(0)
        base = _decimal(theorique)
        res_reg = await db.execute(
            select(
                func.coalesce(
                    func.sum(
                        case(
                            (RegularisationCaisse.sens == "EXCEDENT", RegularisationCaisse.montant),
                            else_=-RegularisationCaisse.montant,
                        )
                    ),
                    0,
                )
            ).where(
                RegularisationCaisse.organisation_id == tenant_id,
                RegularisationCaisse.source_type == "CLOTURE",
                RegularisationCaisse.source_id == last.id,
                RegularisationCaisse.devise == devise,
            )
        )
        return base + _decimal(res_reg.scalar_one() or 0)

    solde_initial_usd = await _report("USD", last.solde_theorique_usd if last else 0)
    solde_initial_cdf = await _report("CDF", last.solde_theorique_cdf if last else 0)

    settings_res = await db.execute(
        select(PrintSettings).where(PrintSettings.organisation_id == tenant_id).limit(1)
    )
    ps = settings_res.scalar_one_or_none()
    try:
        if ps and ps.exchange_rate_cdf:
            taux_change = Decimal(ps.exchange_rate_cdf or 1)
        else:
            taux_change = Decimal(ps.exchange_rate or 1) if ps else Decimal("1")
    except Exception:
        taux_change = Decimal("1")

    # Les encaissements ANNULÉS ou supprimés ne sont plus en caisse : leur montant
    # a été redébité. Sans ces deux filtres, le solde théorique de clôture les
    # recompte et diverge du solde de trésorerie (qui, lui, les exclut).
    enc_query = select(func.coalesce(func.sum(Encaissement.montant_paye), 0)).where(
        Encaissement.organisation_id == tenant_id,
        Encaissement.canal == "CAISSE",
        Encaissement.devise_perception == "USD",
        Encaissement.est_proforma.is_(False),
        Encaissement.is_deleted.is_(False),
        (Encaissement.statut_operation.is_(None)) | (Encaissement.statut_operation == "ACTIVE"),
    )
    if date_debut:
        enc_query = enc_query.where(Encaissement.date_encaissement > date_debut)
    enc_query = enc_query.where(Encaissement.date_encaissement <= date_fin)
    enc_total_usd = _decimal((await db.execute(enc_query)).scalar_one() or 0)

    enc_cdf_query = select(func.coalesce(func.sum(Encaissement.montant_percu), 0)).where(
        Encaissement.organisation_id == tenant_id,
        Encaissement.canal == "CAISSE",
        Encaissement.devise_perception == "CDF",
        Encaissement.est_proforma.is_(False),
        Encaissement.is_deleted.is_(False),
        (Encaissement.statut_operation.is_(None)) | (Encaissement.statut_operation == "ACTIVE"),
    )
    if date_debut:
        enc_cdf_query = enc_cdf_query.where(Encaissement.date_encaissement > date_debut)
    enc_cdf_query = enc_cdf_query.where(Encaissement.date_encaissement <= date_fin)
    enc_total_cdf = _decimal((await db.execute(enc_cdf_query)).scalar_one() or 0)

    paiement_ts = func.coalesce(SortieFonds.date_paiement, SortieFonds.created_at)
    sort_query = select(func.coalesce(func.sum(SortieFonds.montant_paye), 0)).where(
        SortieFonds.organisation_id == tenant_id,
        (SortieFonds.statut.is_(None)) | (SortieFonds.statut == "VALIDE"),
        SortieFonds.canal == "CAISSE",
        SortieFonds.devise == "USD",
    )
    if date_debut:
        sort_query = sort_query.where(paiement_ts > date_debut)
    sort_query = sort_query.where(paiement_ts <= date_fin)
    sort_total_usd = _decimal((await db.execute(sort_query)).scalar_one() or 0)

    sort_cdf_query = select(func.coalesce(func.sum(SortieFonds.montant_paye), 0)).where(
        SortieFonds.organisation_id == tenant_id,
        (SortieFonds.statut.is_(None)) | (SortieFonds.statut == "VALIDE"),
        SortieFonds.canal == "CAISSE",
        SortieFonds.devise == "CDF",
    )
    if date_debut:
        sort_cdf_query = sort_cdf_query.where(paiement_ts > date_debut)
    sort_cdf_query = sort_cdf_query.where(paiement_ts <= date_fin)
    sort_total_cdf = _decimal((await db.execute(sort_cdf_query)).scalar_one() or 0)

    # --- Approvisionnements de la caisse (banque -> caisse) : ce sont des
    # ENTRÉES de caisse. Leur canal est BANQUE, donc ils n'apparaissent pas
    # dans les sorties caisse ci-dessus ; il faut les ajouter aux entrées,
    # sinon le total des entrées est sous-évalué et le solde ne « tombe » pas.
    async def _appro_sum(devise: str) -> Decimal:
        q = select(func.coalesce(func.sum(SortieFonds.montant_paye), 0)).where(
            SortieFonds.organisation_id == tenant_id,
            (SortieFonds.statut.is_(None)) | (SortieFonds.statut == "VALIDE"),
            SortieFonds.type_sortie == "approvisionnement_caisse",
            SortieFonds.devise == devise,
        )
        if date_debut:
            q = q.where(paiement_ts > date_debut)
        q = q.where(paiement_ts <= date_fin)
        return _decimal((await db.execute(q)).scalar_one() or 0)

    # --- Transferts internes (module dédié) : entrée si destination = CAISSE,
    # sortie si source = CAISSE.
    async def _transf_sum(devise: str, *, as_destination: bool) -> Decimal:
        # NE PAS filtrer sur `TransfertInterne.statut`. La correction d'un transfert
        # est additive : l'original (CONTREPASSE) et sa ligne inverse (EXECUTE)
        # coexistent et s'annulent arithmétiquement. Exclure l'original en gardant
        # l'inverse produirait un net inversé, c'est-à-dire de l'argent créé de rien.
        col = TransfertInterne.destination_type if as_destination else TransfertInterne.source_type
        q = select(func.coalesce(func.sum(TransfertInterne.montant), 0)).where(
            TransfertInterne.organisation_id == tenant_id,
            col == "CAISSE",
            TransfertInterne.devise == devise,
        )
        if date_debut:
            q = q.where(TransfertInterne.date_transfert > date_debut)
        q = q.where(TransfertInterne.date_transfert <= date_fin)
        return _decimal((await db.execute(q)).scalar_one() or 0)

    # --- Retours en caisse (reliquats d'avances rendus) : physiquement de
    # l'argent qui RENTRE dans le tiroir, donc une entrée. Sans ce terme le
    # solde théorique est sous-évalué du total des retours et la clôture affiche
    # un écart de caisse qui n'existe pas.
    async def _retour_sum(devise: str) -> Decimal:
        q = select(func.coalesce(func.sum(RetourCaisse.montant), 0)).where(
            RetourCaisse.organisation_id == tenant_id,
            RetourCaisse.statut == "VALIDE",
            RetourCaisse.canal == "CAISSE",
            RetourCaisse.devise == devise,
        )
        if date_debut:
            q = q.where(RetourCaisse.date_retour > date_debut)
        q = q.where(RetourCaisse.date_retour <= date_fin)
        return _decimal((await db.execute(q)).scalar_one() or 0)

    appro_usd = await _appro_sum("USD")
    appro_cdf = await _appro_sum("CDF")
    transf_in_usd = await _transf_sum("USD", as_destination=True)
    transf_in_cdf = await _transf_sum("CDF", as_destination=True)
    transf_out_usd = await _transf_sum("USD", as_destination=False)
    transf_out_cdf = await _transf_sum("CDF", as_destination=False)
    retours_usd = await _retour_sum("USD")
    retours_cdf = await _retour_sum("CDF")

    total_entrees_usd = enc_total_usd + appro_usd + transf_in_usd + retours_usd
    total_entrees_cdf = enc_total_cdf + appro_cdf + transf_in_cdf + retours_cdf
    total_sorties_usd = sort_total_usd + transf_out_usd
    total_sorties_cdf = sort_total_cdf + transf_out_cdf

    # Solde théorique cohérent avec ce qui est affiché : il DOIT être égal à
    # solde initial + entrées − sorties (self-consistant), et non plus lu depuis
    # une autre source recalculée à part (qui divergeait des entrées/sorties).
    solde_theorique_usd = solde_initial_usd + total_entrees_usd - total_sorties_usd
    solde_theorique_cdf = solde_initial_cdf + total_entrees_cdf - total_sorties_cdf

    # Détail des entrées internes de la période : approvisionnements du chemin
    # historique ET transferts du moteur dédié, mêmes bornes et mêmes filtres
    # que `_appro_sum` + `_transf_sum`, pour que la liste affichée additionne
    # exactement le montant repris dans les entrées. La borne stricte vaut pour
    # les deux sources.
    lignes_entrees_internes = await list_entrees_internes_caisse(
        db,
        tenant_id=tenant_id,
        date_debut=date_debut,
        date_fin=date_fin,
        strict_debut=True,
    )

    return ClotureBalanceResponse(
        date_debut=date_debut,
        date_fin=date_fin,
        taux_change=taux_change,
        solde_initial_usd=solde_initial_usd,
        solde_initial_cdf=solde_initial_cdf,
        total_entrees_usd=total_entrees_usd,
        total_entrees_cdf=total_entrees_cdf,
        total_sorties_usd=total_sorties_usd,
        total_sorties_cdf=total_sorties_cdf,
        solde_theorique_usd=solde_theorique_usd,
        solde_theorique_cdf=solde_theorique_cdf,
        entrees_encaissements_usd=enc_total_usd,
        entrees_encaissements_cdf=enc_total_cdf,
        entrees_approvisionnements_usd=appro_usd,
        entrees_approvisionnements_cdf=appro_cdf,
        entrees_transferts_usd=transf_in_usd,
        entrees_transferts_cdf=transf_in_cdf,
        entrees_retours_usd=retours_usd,
        entrees_retours_cdf=retours_cdf,
        # Le champ garde son nom : c'est un contrat déjà consommé par le
        # frontend. Il porte désormais les deux sources — `type_operation`
        # distingue APPROVISIONNEMENT de TRANSFERT_INTERNE pour qui en a besoin.
        approvisionnements=[
            {
                **ligne,
                "montant": str(ligne["montant"]),
                "date": ligne["date"].isoformat() if ligne["date"] else None,
            }
            for ligne in lignes_entrees_internes
        ],
    )


def _cloture_out(c: ClotureCaisse) -> ClotureOut:
    return ClotureOut(
        id=c.id,
        reference_numero=c.reference_numero,
        date_cloture=c.date_cloture,
        date_debut=c.date_debut,
        caissier_id=str(c.caissier_id) if c.caissier_id else None,
        solde_initial_usd=c.solde_initial_usd,
        solde_initial_cdf=c.solde_initial_cdf,
        total_entrees_usd=c.total_entrees_usd,
        total_entrees_cdf=c.total_entrees_cdf,
        total_sorties_usd=c.total_sorties_usd,
        total_sorties_cdf=c.total_sorties_cdf,
        solde_theorique_usd=c.solde_theorique_usd,
        solde_theorique_cdf=c.solde_theorique_cdf,
        solde_physique_usd=c.solde_physique_usd,
        solde_physique_cdf=c.solde_physique_cdf,
        ecart_usd=c.ecart_usd,
        ecart_cdf=c.ecart_cdf,
        taux_change_applique=c.taux_change_applique,
        billetage_usd=c.billetage_usd,
        billetage_cdf=c.billetage_cdf,
        observation=c.observation,
        pdf_path=c.pdf_path,
        statut=c.statut,
    )


def _ensure_cloture_pdf_dir() -> None:
    os.makedirs(CLTURE_PDF_DIR, exist_ok=True)


def _safe_ref(value: str) -> str:
    cleaned = "".join(ch for ch in value if ch.isalnum() or ch in ("-", "_"))
    return cleaned or "CLOTURE"


async def _resolve_org_uuid(db: AsyncSession, organisation_id: int) -> str:
    """UUID public de l'organisation, pour ranger les fichiers par tenant."""
    res = await db.execute(
        select(Organisation.uuid).where(Organisation.id == organisation_id).limit(1)
    )
    value = res.scalar_one_or_none()
    return str(value) if value else str(organisation_id)


def _cloture_tenant_dir(tenant_uuid: str) -> str:
    """Dossier des PV de clôture propre à un tenant : tenants/<uuid>/clotures/."""
    return os.path.abspath(os.path.join(settings.upload_dir, "tenants", tenant_uuid, "clotures"))


def _resolve_cloture_file(stored_path: str) -> str | None:
    """Chemin absolu du PV à partir de la valeur stockée en base.

    Gère les deux formats :
    - nouveau : chemin relatif « tenants/<uuid>/clotures/<fichier>.pdf » (scopé tenant) ;
    - ancien  : simple nom de fichier rangé dans le dossier plat historique.
    Refuse toute sortie du répertoire d'uploads (protection path-traversal).
    """
    root = os.path.abspath(settings.upload_dir)
    normalized = (stored_path or "").replace("\\", "/").lstrip("/")
    if not normalized:
        return None
    if "/" in normalized:
        candidate = os.path.abspath(os.path.join(root, normalized))
    else:
        # Rétrocompatibilité : anciens PV stockés à plat dans uploads/clotures/.
        candidate = os.path.abspath(os.path.join(CLTURE_PDF_DIR, normalized))
    try:
        if os.path.commonpath([root, candidate]) != root:
            return None
    except ValueError:
        return None
    return candidate


@router.get("/balance-check", response_model=ClotureBalanceResponse, dependencies=[Depends(has_permission("cloture_caisse"))])
async def get_balance_check(
    db: AsyncSession = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
) -> ClotureBalanceResponse:
    return await _compute_balance(db, tenant_id)


@router.get("/status-today")
async def get_cloture_status_today(
    tenant_id: int = Depends(get_public_tenant_id),
    db: AsyncSession = Depends(get_db),
) -> dict:
    res = await db.execute(
        select(ClotureCaisse)
        .where(ClotureCaisse.organisation_id == tenant_id)
        .order_by(ClotureCaisse.date_cloture.desc())
        .limit(1)
    )
    last = res.scalar_one_or_none()
    if last is None or not last.date_cloture:
        return {"is_closed": False, "date": None}
    last_dt = last.date_cloture
    if last_dt.tzinfo is None:
        last_dt = last_dt.replace(tzinfo=timezone.utc)
    today = datetime.now(timezone.utc).date()
    return {"is_closed": last_dt.date() == today, "date": last_dt.date().isoformat()}


@router.get("", response_model=list[ClotureOut], dependencies=[Depends(has_permission("cloture_caisse"))])
async def list_clotures(
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    date_debut: str | None = Query(default=None),
    date_fin: str | None = Query(default=None),
    caissier_id: str | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
) -> list[ClotureOut]:
    stmt = select(ClotureCaisse)
    if date_debut:
        try:
            start_dt = datetime.fromisoformat(date_debut)
        except ValueError:
            start_dt = None
        if start_dt:
            if start_dt.tzinfo is None:
                start_dt = start_dt.replace(tzinfo=timezone.utc)
            stmt = stmt.where(ClotureCaisse.date_cloture >= start_dt)
    if date_fin:
        try:
            end_dt = datetime.fromisoformat(date_fin)
        except ValueError:
            end_dt = None
        if end_dt:
            if end_dt.tzinfo is None:
                end_dt = end_dt.replace(tzinfo=timezone.utc)
            stmt = stmt.where(ClotureCaisse.date_cloture <= end_dt)
    if caissier_id:
        try:
            caissier_uid = uuid.UUID(caissier_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="caissier_id invalide")
        stmt = stmt.where(ClotureCaisse.caissier_id == caissier_uid)

    res = await db.execute(
        stmt.order_by(ClotureCaisse.date_cloture.desc()).limit(limit).offset(offset)
    )
    return [_cloture_out(c) for c in res.scalars().all()]


@router.get("/caissiers", dependencies=[Depends(has_permission("cloture_caisse"))])
async def list_cloture_caissiers(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[dict]:
    stmt = select(User.id, User.email, User.nom, User.prenom).join(
        ClotureCaisse, ClotureCaisse.caissier_id == User.id
    )
    if (user.role or "").lower() != "super_admin":
        stmt = stmt.where(User.role != "super_admin")
    stmt = stmt.distinct().order_by(User.email.asc())
    res = await db.execute(stmt)
    users = []
    for uid, email, nom, prenom in res.all():
        label = " ".join(filter(None, [prenom, nom])) or email or str(uid)
        users.append({"id": str(uid), "label": label, "email": email})
    return users


@router.get("/export-xlsx", dependencies=[Depends(has_permission("cloture_caisse"))])
async def export_clotures_xlsx(
    limit: int = Query(default=5000, ge=1, le=50000),
    offset: int = Query(default=0, ge=0),
    db: AsyncSession = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    # Filtre explicite plutôt que délégué au hook ORM : celui-ci est inerte
    # quand le contexte tenant vaut None (super-admin sur l'hôte d'admin).
    res = await db.execute(
        select(ClotureCaisse)
        .where(ClotureCaisse.organisation_id == tenant_id)
        .order_by(ClotureCaisse.date_cloture.desc())
        .limit(limit)
        .offset(offset)
    )
    clotures = res.scalars().all()
    organisation = await tenant_display_name(db, tenant_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Clotures"
    # Émetteur en tête : aucun document ne sort sans identifier son tenant.
    ws.append(["CLÔTURES DE CAISSE"])
    ws.append([organisation])
    ws.append([])
    ws["A1"].font = XlFont(bold=True, size=14)
    ws["A2"].font = XlFont(bold=True, size=11)
    ws.append(
        [
            "id",
            "reference_numero",
            "date_cloture",
            "date_debut",
            "caissier_id",
            "solde_initial_usd",
            "total_entrees_usd",
            "total_sorties_usd",
            "solde_theorique_usd",
            "solde_physique_usd",
            "ecart_usd",
            "solde_initial_cdf",
            "total_entrees_cdf",
            "total_sorties_cdf",
            "solde_theorique_cdf",
            "solde_physique_cdf",
            "ecart_cdf",
            "statut",
            "observation",
        ]
    )
    for c in clotures:
        ws.append(
            [
                c.id,
                c.reference_numero,
                c.date_cloture.isoformat() if c.date_cloture else "",
                c.date_debut.isoformat() if c.date_debut else "",
                str(c.caissier_id) if c.caissier_id else "",
                float(c.solde_initial_usd or 0),
                float(c.total_entrees_usd or 0),
                float(c.total_sorties_usd or 0),
                float(c.solde_theorique_usd or 0),
                float(c.solde_physique_usd or 0),
                float(c.ecart_usd or 0),
                float(c.solde_initial_cdf or 0),
                float(c.total_entrees_cdf or 0),
                float(c.total_sorties_cdf or 0),
                float(c.solde_theorique_cdf or 0),
                float(c.solde_physique_cdf or 0),
                float(c.ecart_cdf or 0),
                c.statut,
                c.observation or "",
            ]
        )
    # Sérialisation en thread : elle domine le coût CPU de l'export et figerait
    # sinon la boucle d'événements pour toutes les requêtes en cours.
    output = await save_workbook(wb)

    filename = f"clotures_{datetime.now(timezone.utc).date().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.post("", response_model=ClotureOut, dependencies=[Depends(has_permission("can_execute_payment"))])
async def create_cloture(
    payload: ClotureCreateRequest,
    request: Request,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
) -> ClotureOut:
    balance = await _compute_balance(db, tenant_id)

    # On ne clôture qu'une caisse ouverte (Modèle B) : une caisse déjà fermée
    # doit d'abord être rouverte.
    caisse_guard = (await db.execute(
        select(CaisseCentrale).where(CaisseCentrale.organisation_id == tenant_id).limit(1)
    )).scalar_one_or_none()
    if caisse_guard is not None and not caisse_guard.est_ouverte:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La caisse est déjà fermée : ouvrez-la avant de la clôturer.",
        )

    solde_physique_usd = _decimal(payload.solde_physique_usd)
    solde_physique_cdf = _decimal(payload.solde_physique_cdf)
    taux_change = _decimal(balance.taux_change or 1)
    if taux_change <= 0:
        taux_change = Decimal("1")
    ecart_usd = _decimal(solde_physique_usd - balance.solde_theorique_usd)
    ecart_cdf = _decimal(solde_physique_cdf - balance.solde_theorique_cdf)

    reference_numero = await generate_document_number(db, "CLO", tenant_id)
    cloture = ClotureCaisse(
        organisation_id=user.organisation_id,
        reference_numero=reference_numero,
        date_cloture=balance.date_fin,
        date_debut=balance.date_debut,
        caissier_id=user.id,
        solde_initial_usd=balance.solde_initial_usd,
        solde_initial_cdf=balance.solde_initial_cdf,
        total_entrees_usd=balance.total_entrees_usd,
        total_entrees_cdf=balance.total_entrees_cdf,
        total_sorties_usd=balance.total_sorties_usd,
        total_sorties_cdf=balance.total_sorties_cdf,
        solde_theorique_usd=balance.solde_theorique_usd,
        solde_theorique_cdf=balance.solde_theorique_cdf,
        solde_physique_usd=solde_physique_usd,
        solde_physique_cdf=solde_physique_cdf,
        ecart_usd=ecart_usd,
        ecart_cdf=ecart_cdf,
        taux_change_applique=taux_change,
        billetage_usd=payload.billetage_usd,
        billetage_cdf=payload.billetage_cdf,
        observation=(payload.observation or "").strip() or None,
        statut="VALIDEE",
    )
    db.add(cloture)

    await db.flush()

    # Régularisation de l'écart : le comptage physique ne REMPLACE PAS le solde
    # théorique — cela détruirait la traçabilité. Si l'utilisateur la demande,
    # l'écart donne lieu à une opération identifiable (encaissement si excédent,
    # sortie si déficit) et c'est elle qui déplace le solde. Sinon le solde reste
    # au théorique et l'écart demeure ouvert, régularisable plus tard.
    regularisations, regularisation_erreurs = await _appliquer_regularisations(
        db,
        tenant_id=tenant_id,
        user_id=user.id,
        demandee=payload.regulariser_ecart,
        motif=payload.motif_regularisation,
        source_type=SOURCE_CLOTURE,
        source_id=cloture.id,
        source_reference=reference_numero,
        ecarts=[
            EcartCaisse("USD", balance.solde_theorique_usd, solde_physique_usd),
            EcartCaisse("CDF", balance.solde_theorique_cdf, solde_physique_cdf),
        ],
    )

    # Fin de session : on ferme la caisse. Le solde n'est PAS réaligné sur le
    # comptage — seules les régularisations ci-dessus peuvent le déplacer.
    caisse_res = await db.execute(
        select(CaisseCentrale).where(CaisseCentrale.organisation_id == tenant_id).limit(1)
    )
    caisse = caisse_res.scalar_one_or_none()
    if caisse is not None:
        caisse.est_ouverte = False
        caisse.ouverte_le = None
        caisse.ouverte_par_id = None
        caisse.derniere_maj = datetime.now(timezone.utc)

    await log_action(
        db,
        user_id=user.id,
        action="CAISSE_CLOTURE_JOURNALIERE",
        target_table="clotures",
        target_id=reference_numero,
        new_value={
            "solde_theorique_usd": str(balance.solde_theorique_usd),
            "solde_physique_usd": str(solde_physique_usd),
            "ecart_usd": str(ecart_usd),
            "solde_theorique_cdf": str(balance.solde_theorique_cdf),
            "solde_physique_cdf": str(solde_physique_cdf),
            "ecart_cdf": str(ecart_cdf),
            "regularisations": regularisations,
            "regularisation_erreurs": regularisation_erreurs,
        },
        ip_address=get_request_ip(request),
    )
    await db.commit()
    await db.refresh(cloture)
    out = _cloture_out(cloture)
    out.regularisations = regularisations
    out.regularisation_erreurs = regularisation_erreurs
    return out


def _ouverture_out(o: OuvertureCaisse) -> OuvertureOut:
    return OuvertureOut(
        id=o.id,
        reference_numero=o.reference_numero,
        date_ouverture=o.date_ouverture,
        caissier_id=o.caissier_id,
        solde_ouverture_usd=o.solde_ouverture_usd,
        solde_ouverture_cdf=o.solde_ouverture_cdf,
        solde_attendu_usd=o.solde_attendu_usd,
        solde_attendu_cdf=o.solde_attendu_cdf,
        ecart_usd=o.ecart_usd,
        ecart_cdf=o.ecart_cdf,
        billetage_usd=o.billetage_usd,
        billetage_cdf=o.billetage_cdf,
        observation=o.observation,
        statut=o.statut,
    )


@router.get("/ecarts", dependencies=[Depends(has_permission("cloture_caisse"))])
async def list_ecarts_caisse(
    non_regularises_seulement: bool = Query(default=True),
    limit: int = Query(default=100, ge=1, le=500),
    tenant_id: int = Depends(get_current_tenant_id),
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    """Écarts de caisse constatés aux comptages, avec leur état de régularisation.

    Un écart non régularisé n'est pas une anomalie technique : l'utilisateur a pu
    refuser la régularisation. Il reste listé ici pour être traité plus tard.
    """
    regs_res = await db.execute(
        select(RegularisationCaisse).where(RegularisationCaisse.organisation_id == tenant_id)
    )
    regs_par_source: dict[tuple[str, int], list[RegularisationCaisse]] = {}
    for reg in regs_res.scalars().all():
        regs_par_source.setdefault((reg.source_type, reg.source_id), []).append(reg)

    lignes: list[dict] = []

    def _ajouter(source_type: str, row, date_value, ecart_usd, ecart_cdf) -> None:
        for devise, ecart in (("USD", _decimal(ecart_usd)), ("CDF", _decimal(ecart_cdf))):
            if ecart == 0:
                continue
            regularise = any(
                r.devise == devise
                for r in regs_par_source.get((source_type, row.id), [])
            )
            if non_regularises_seulement and regularise:
                continue
            lignes.append(
                {
                    "source_type": source_type,
                    "source_id": row.id,
                    "reference_numero": row.reference_numero,
                    "date": date_value.isoformat() if date_value else None,
                    "devise": devise,
                    "ecart": str(ecart),
                    "sens": "EXCEDENT" if ecart > 0 else "DEFICIT",
                    "regularise": regularise,
                }
            )

    ouvertures = (
        await db.execute(
            select(OuvertureCaisse)
            .where(OuvertureCaisse.organisation_id == tenant_id)
            .order_by(OuvertureCaisse.date_ouverture.desc())
            .limit(limit)
        )
    ).scalars().all()
    for o in ouvertures:
        _ajouter(SOURCE_OUVERTURE, o, o.date_ouverture, o.ecart_usd, o.ecart_cdf)

    clotures = (
        await db.execute(
            select(ClotureCaisse)
            .where(ClotureCaisse.organisation_id == tenant_id)
            .order_by(ClotureCaisse.date_cloture.desc())
            .limit(limit)
        )
    ).scalars().all()
    for c in clotures:
        _ajouter(SOURCE_CLOTURE, c, c.date_cloture, c.ecart_usd, c.ecart_cdf)

    lignes.sort(key=lambda x: x["date"] or "", reverse=True)
    return lignes


@router.post("/ecarts/{source_type}/{source_id}/regulariser", dependencies=[Depends(has_permission("can_execute_payment"))])
async def regulariser_ecart_a_posteriori(
    source_type: str,
    source_id: int,
    payload: EcartRegularisationRequest,
    user: User = Depends(get_current_user),
    tenant_id: int = Depends(get_current_tenant_id),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Régularise après coup un écart laissé ouvert à l'ouverture ou à la clôture."""
    source = source_type.strip().upper()
    if source not in (SOURCE_OUVERTURE, SOURCE_CLOTURE):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="source_type invalide")

    if source == SOURCE_OUVERTURE:
        row = (
            await db.execute(
                select(OuvertureCaisse).where(
                    OuvertureCaisse.id == source_id,
                    OuvertureCaisse.organisation_id == tenant_id,
                )
            )
        ).scalar_one_or_none()
        theorique_usd, physique_usd = (
            row.solde_attendu_usd,
            row.solde_ouverture_usd,
        ) if row else (0, 0)
        theorique_cdf, physique_cdf = (
            row.solde_attendu_cdf,
            row.solde_ouverture_cdf,
        ) if row else (0, 0)
    else:
        row = (
            await db.execute(
                select(ClotureCaisse).where(
                    ClotureCaisse.id == source_id,
                    ClotureCaisse.organisation_id == tenant_id,
                )
            )
        ).scalar_one_or_none()
        theorique_usd, physique_usd = (
            row.solde_theorique_usd,
            row.solde_physique_usd,
        ) if row else (0, 0)
        theorique_cdf, physique_cdf = (
            row.solde_theorique_cdf,
            row.solde_physique_cdf,
        ) if row else (0, 0)

    if row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Écart introuvable")

    deja = (
        await db.execute(
            select(RegularisationCaisse).where(
                RegularisationCaisse.organisation_id == tenant_id,
                RegularisationCaisse.source_type == source,
                RegularisationCaisse.source_id == source_id,
            )
        )
    ).scalars().all()
    devises_deja = {r.devise for r in deja}

    ecarts = [
        EcartCaisse("USD", _decimal(theorique_usd), _decimal(physique_usd)),
        EcartCaisse("CDF", _decimal(theorique_cdf), _decimal(physique_cdf)),
    ]
    if payload.devise:
        ecarts = [e for e in ecarts if e.devise == payload.devise.upper()]
    ecarts = [e for e in ecarts if e.devise not in devises_deja]

    creees: list[dict] = []
    erreurs: list[str] = []
    for ecart in ecarts:
        if ecart.sens is None:
            continue
        try:
            reg = await regulariser_ecart(
                db,
                tenant_id=tenant_id,
                user_id=user.id,
                ecart=ecart,
                motif=payload.motif or "",
                source_type=source,
                source_id=source_id,
                source_reference=row.reference_numero,
            )
        except RegularisationImpossible as exc:
            erreurs.append(f"{ecart.devise} : {exc.message}")
            continue
        creees.append(
            {"devise": reg.devise, "sens": reg.sens, "montant": str(reg.montant)}
        )

    if not creees and erreurs:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=" / ".join(erreurs))

    await db.commit()
    return {"ok": bool(creees), "regularisations": creees, "erreurs": erreurs}


@router.get("/ouvertures", response_model=list[OuvertureOut], dependencies=[Depends(has_permission("cloture_caisse"))])
async def list_ouvertures(
    limit: int = Query(default=50, le=200),
    tenant_id: int = Depends(get_current_tenant_id),
    db: AsyncSession = Depends(get_db),
) -> list[OuvertureOut]:
    res = await db.execute(
        select(OuvertureCaisse)
        .where(OuvertureCaisse.organisation_id == tenant_id)
        .order_by(OuvertureCaisse.date_ouverture.desc())
        .limit(limit)
    )
    return [_ouverture_out(o) for o in res.scalars().all()]


@router.get("/caisse-status")
async def get_caisse_status(
    tenant_id: int = Depends(get_current_tenant_id),
    db: AsyncSession = Depends(get_db),
) -> dict:
    res = await db.execute(
        select(CaisseCentrale).where(CaisseCentrale.organisation_id == tenant_id).limit(1)
    )
    caisse = res.scalar_one_or_none()
    return {
        "est_ouverte": bool(caisse.est_ouverte) if caisse is not None else False,
        "ouverte_le": caisse.ouverte_le if caisse is not None else None,
        "solde_usd": str(caisse.solde_usd) if caisse is not None else "0",
        "solde_cdf": str(caisse.solde_cdf) if caisse is not None else "0",
    }


@router.post("/ouverture", response_model=OuvertureOut, dependencies=[Depends(has_permission("can_execute_payment"))])
async def open_caisse(
    payload: OuvertureCreateRequest,
    request: Request,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
) -> OuvertureOut:
    res = await db.execute(
        select(CaisseCentrale).where(CaisseCentrale.organisation_id == tenant_id).limit(1)
    )
    caisse = res.scalar_one_or_none()
    if caisse is None:
        caisse = CaisseCentrale(organisation_id=tenant_id, solde_usd=0, solde_cdf=0)
        db.add(caisse)
        await db.flush()
    if caisse.est_ouverte:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La caisse est déjà ouverte.")

    now = datetime.now(timezone.utc)
    solde_usd = _decimal(payload.solde_ouverture_usd)
    solde_cdf = _decimal(payload.solde_ouverture_cdf)
    # Solde attendu = report de la dernière clôture (= solde courant avant ouverture).
    attendu_usd = _decimal(caisse.solde_usd or 0)
    attendu_cdf = _decimal(caisse.solde_cdf or 0)
    ecart_usd = _decimal(solde_usd - attendu_usd)
    ecart_cdf = _decimal(solde_cdf - attendu_cdf)
    reference_numero = await generate_document_number(db, "OUV", tenant_id)
    ouverture = OuvertureCaisse(
        organisation_id=tenant_id,
        reference_numero=reference_numero,
        date_ouverture=now,
        caissier_id=user.id,
        solde_ouverture_usd=solde_usd,
        solde_ouverture_cdf=solde_cdf,
        solde_attendu_usd=attendu_usd,
        solde_attendu_cdf=attendu_cdf,
        ecart_usd=ecart_usd,
        ecart_cdf=ecart_cdf,
        billetage_usd=payload.billetage_usd,
        billetage_cdf=payload.billetage_cdf,
        observation=(payload.observation or "").strip() or None,
        statut="OUVERTE",
        created_at=now,
    )
    db.add(ouverture)
    await db.flush()

    # Le fond compté ne DEVIENT PAS le solde courant : cela écraserait le solde
    # théorique et supprimerait toute trace de l'écart. Si l'utilisateur demande
    # la régularisation, l'écart donne lieu à une opération identifiable qui,
    # elle, déplace le solde. Sinon la caisse s'ouvre sur le solde théorique et
    # l'écart reste ouvert.
    regularisations, regularisation_erreurs = await _appliquer_regularisations(
        db,
        tenant_id=tenant_id,
        user_id=user.id,
        demandee=payload.regulariser_ecart,
        motif=payload.motif_regularisation,
        source_type=SOURCE_OUVERTURE,
        source_id=ouverture.id,
        source_reference=reference_numero,
        ecarts=[
            EcartCaisse("USD", attendu_usd, solde_usd),
            EcartCaisse("CDF", attendu_cdf, solde_cdf),
        ],
    )

    caisse.est_ouverte = True
    caisse.ouverte_le = now
    caisse.ouverte_par_id = user.id
    caisse.derniere_maj = now

    await log_action(
        db,
        user_id=user.id,
        action="CAISSE_OUVERTURE",
        target_table="ouvertures_caisse",
        target_id=reference_numero,
        new_value={
            "solde_ouverture_usd": str(solde_usd),
            "solde_ouverture_cdf": str(solde_cdf),
            "regularisations": regularisations,
            "regularisation_erreurs": regularisation_erreurs,
        },
        ip_address=get_request_ip(request),
    )
    await db.commit()
    await db.refresh(ouverture)
    out = _ouverture_out(ouverture)
    out.regularisations = regularisations
    out.regularisation_erreurs = regularisation_erreurs
    return out


@router.post("/{cloture_id}/pdf", dependencies=[Depends(has_permission("cloture_caisse"))])
async def upload_cloture_pdf(
    cloture_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
) -> dict:
    res = await db.execute(select(ClotureCaisse).where(ClotureCaisse.id == cloture_id))
    cloture = res.scalar_one_or_none()
    if cloture is None:
        raise HTTPException(status_code=404, detail="Clôture introuvable")

    content_type = (file.content_type or "").lower()
    if content_type not in PDF_ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail="Format de fichier non autorisé")

    original_name = file.filename or "cloture.pdf"
    ext = os.path.splitext(original_name)[1].lower()
    if ext not in PDF_ALLOWED_EXT:
        raise HTTPException(status_code=400, detail="Extension de fichier non autorisée")

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Fichier vide")

    tenant_uuid = await _resolve_org_uuid(db, cloture.organisation_id)
    target_dir = _cloture_tenant_dir(tenant_uuid)
    os.makedirs(target_dir, exist_ok=True)
    safe_ref = _safe_ref(cloture.reference_numero or f"CLOTURE-{cloture.id}")
    filename = f"{safe_ref}-pv.pdf"
    dest_path = os.path.join(target_dir, filename)
    with open(dest_path, "wb") as f:
        f.write(contents)

    # Chemin relatif scopé tenant, cohérent avec les réquisitions/remboursements.
    rel_path = f"tenants/{tenant_uuid}/clotures/{filename}"
    cloture.pdf_path = rel_path
    await db.commit()
    return {"ok": True, "pdf_path": rel_path}


@router.get("/{cloture_id}/pdf", dependencies=[Depends(has_permission("cloture_caisse"))])
async def download_cloture_pdf(
    cloture_id: int,
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(ClotureCaisse).where(ClotureCaisse.id == cloture_id))
    cloture = res.scalar_one_or_none()
    if cloture is None:
        raise HTTPException(status_code=404, detail="Clôture introuvable")
    if not cloture.pdf_path:
        raise HTTPException(status_code=404, detail="PV non archivé")
    file_path = _resolve_cloture_file(cloture.pdf_path)
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Fichier PV introuvable")
    return FileResponse(
        file_path, media_type="application/pdf", filename=os.path.basename(cloture.pdf_path)
    )


@router.get("/{cloture_id}/pdf-data", response_model=CloturePdfData, dependencies=[Depends(has_permission("cloture_caisse"))])
async def get_cloture_pdf_data(
    cloture_id: int,
    db: AsyncSession = Depends(get_db),
) -> CloturePdfData:
    res = await db.execute(select(ClotureCaisse).where(ClotureCaisse.id == cloture_id))
    cloture = res.scalar_one_or_none()
    if cloture is None:
        raise HTTPException(status_code=404, detail="Clôture introuvable")

    start_dt = cloture.date_debut
    end_dt = cloture.date_cloture
    paiement_ts = func.coalesce(SortieFonds.date_paiement, SortieFonds.created_at)
    query = select(SortieFonds).where(
        (SortieFonds.statut.is_(None)) | (SortieFonds.statut == "VALIDE"),
        SortieFonds.canal == "CAISSE",
    )
    if start_dt:
        query = query.where(paiement_ts >= start_dt)
    query = query.where(paiement_ts <= end_dt).order_by(paiement_ts.asc())
    sort_res = await db.execute(query)
    sorties = sort_res.scalars().all()

    details = [
        CloturePdfDetail(
            reference_numero=s.reference_numero or s.reference,
            beneficiaire=s.beneficiaire,
            motif=s.motif,
            montant_paye=s.montant_paye,
        )
        for s in sorties
    ]
    return CloturePdfData(cloture=_cloture_out(cloture), details=details)
