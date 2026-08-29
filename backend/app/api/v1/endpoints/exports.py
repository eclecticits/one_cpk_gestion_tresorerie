from __future__ import annotations

from collections import defaultdict
from copy import copy as _copier_style
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from decimal import Decimal
from io import BytesIO
from typing import Any
from collections.abc import Iterator, Sequence
import logging
import re
import textwrap
import unicodedata

import uuid
import anyio
from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import Select, func, or_, select
from sqlalchemy.orm import aliased, joinedload
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import (
    _normalize_plan_status,
    get_cached_saas_status,
    get_current_user,
    has_permission,
)
from app.core.config import settings
from app.services.export_jobs import serialiser_job, soumettre, types_asynchrones
from app.services.export_queue import publier
from app.db.session import get_db
from app.utils.excel_io import save_workbook
from app.models.encaissement import Encaissement
from app.models.expert_comptable import ExpertComptable
from app.models.organisation import Organisation
from app.models.print_settings import PrintSettings
from app.services.entrees_caisse import list_entrees_internes_caisse, list_entrees_internes_banque
from app.services.tenant_identity import tenant_display_name
from app.models.budget import BudgetExercice, BudgetPoste
from app.models.budget_commentaire import BudgetPosteCommentaire
from app.models.compte_bancaire import CompteBancaire
from app.models.ligne_requisition import LigneRequisition
from app.models.requisition import Requisition
from app.models.service import Service
from app.models.service_rubrique import ServiceRubrique
from app.models.sortie_fonds import SortieFonds
from app.services import transferts_delegues
from app.models.retour_caisse import RetourCaisse
from app.models.user import User
from app.utils.budget_code import cle_tri_code_budget

logger = logging.getLogger("onec_cpk_api.exports")

router = APIRouter()


async def require_expert_admin(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> User:
    role = (user.role or "").lower()
    if role == "super_admin":
        return user
    if role == "admin" and user.organisation_id:
        org_res = await db.execute(select(Organisation.slug).where(Organisation.id == user.organisation_id))
        slug = (org_res.scalar_one_or_none() or "").lower()
        if slug == "cn":
            return user
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail="Action réservée au Conseil National (CN).",
    )

REQUISITION_STATUTS_VALIDES = ("APPROUVEE", "PAYEE")


def _requisition_status_values_for_filter(value: str) -> list[str]:
    normalized = value.strip().upper()
    return {
        "AUTORISEE": ["AUTORISEE", "VALIDEE", "VALIDEE_TRESORERIE", "VALIDE_TECHNIQUE"],
        "PAYEE": ["PAYEE", "DECAISSE"],
        "REJETEE": ["REJETEE", "REJETTE"],
    }.get(normalized, [normalized])


def _strip_accents(value: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", value) if not unicodedata.combining(c))


def _normalize_statut_professionnel(value: str | None) -> str | None:
    if not value:
        return None
    trimmed = value.strip()
    compact = trimmed.replace("_", " ").replace("-", " ")
    compact = " ".join(compact.split())
    normalized = _strip_accents(compact).lower()
    mapping = {
        "en cabinet": "En Cabinet",
        "independant": "Indépendant",
        "salarie": "Salarié",
        "cabinet": "Cabinet",
    }
    return mapping.get(normalized, trimmed)


def _statut_professionnel_variants(value: str) -> list[str]:
    canonical = _normalize_statut_professionnel(value)
    if not canonical:
        return []
    variants = {
        "En Cabinet": [
            "En Cabinet",
            "En cabinet",
            "en cabinet",
            "EN CABINET",
            "en_cabinet",
            "En_Cabinet",
            "en-cabinet",
            "En-Cabinet",
        ],
        "Indépendant": [
            "Indépendant",
            "indépendant",
            "Independant",
            "independant",
            "INDEPENDANT",
        ],
        "Salarié": [
            "Salarié",
            "salarié",
            "Salarie",
            "salarie",
            "SALARIE",
        ],
        "Cabinet": [
            "Cabinet",
            "cabinet",
            "CABINET",
        ],
    }
    return variants.get(canonical, [canonical])

def _parse_datetime(value: str | None, end_of_day: bool = False) -> datetime | None:
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(value)
    except ValueError:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    if end_of_day and len(value) <= 10:
        dt = dt.replace(hour=23, minute=59, second=59, microsecond=999999)
    return dt


# PostgreSQL n'accepte pas plus de 32 767 parametres de bind par requete : c'est
# une limite du protocole, pas un reglage. Une clause IN alimentee par un jeu de
# resultats non borne la franchit des que le tenant grossit.
#
# Mesure sur le tenant de charge (60 000 requisitions) :
#   asyncpg.InterfaceError: the number of query arguments cannot exceed 32767
# L'export repondait 500 apres 41 s, AVEC UN SEUL UTILISATEUR. Ce n'est pas un
# probleme de performance : au-dela du seuil l'export est simplement impossible.
#
# 10 000 laisse de la marge pour les autres parametres de la requete.
_TAILLE_LOT_IN = 10_000


def _par_lots(valeurs: Sequence[Any], taille: int = _TAILLE_LOT_IN) -> Iterator[list[Any]]:
    """Decoupe une liste d'identifiants en lots utilisables dans un IN."""
    for debut in range(0, len(valeurs), taille):
        yield list(valeurs[debut : debut + taille])


def _milliers(n: int) -> str:
    """12345 -> '12 345'. Separateur francais, pour un message lisible."""
    return f"{n:,}".replace(",", " ")


class BasculeAsynchroneRequise(Exception):
    """Cet export depasse le seuil : il doit passer par la file, pas par HTTP.

    POURQUOI UNE EXCEPTION plutot qu'une valeur de retour. La decision demande
    le nombre de lignes, or ce nombre n'est connu qu'apres construction de la
    requete — laquelle vit DANS le constructeur, avec ses quinze filtres. Faire
    remonter la decision jusqu'a l'endpoint supposerait soit de dupliquer la
    construction de requete (deux endroits ou se tromper, pour cinq exports),
    soit de scinder chaque constructeur en deux. L'exception traverse ce que
    la valeur de retour ne peut pas traverser, et elle est levee AVANT le
    chargement des entites et la construction du classeur : rien de couteux
    n'a encore ete paye.

    Elle ne franchit jamais la frontiere HTTP : les cinq endpoints la
    rattrapent. Dans le worker, `seuil_bascule` vaut None et elle ne peut pas
    etre levee — c'est ce qui empeche un job de se remettre en file lui-meme,
    indefiniment.
    """

    def __init__(self, total: int) -> None:
        super().__init__(f"{total} lignes : au-dela du seuil de bascule asynchrone.")
        self.total = total


def _seuil_bascule(type_export: str) -> int | None:
    """Seuil applicable a ce type, ou None si le type n'est pas ouvert.

    None et 0 ne veulent PAS dire la meme chose : None = le type reste
    entierement synchrone (drapeau ferme), 0 = le type est ouvert et tout
    bascule, quelle que soit la taille.
    """
    if type_export not in types_asynchrones():
        return None
    return settings.export_async_row_threshold


async def _compter_lignes(
    db: AsyncSession,
    requete: Select,
    *,
    export: str,
    seuil_bascule: int | None = None,
) -> int:
    """Compte les lignes AVANT de construire quoi que ce soit, et refuse l'excessif.

    POURQUOI AVANT, et pas `len(rows)` : le comptage precede l'execution de la
    requete elle-meme. C'est ce qui evite de ramener 120 000 entites ORM en
    memoire (+310 Mo de RSS mesures le 28/08 pour UN export de requisitions,
    non rendus a la fin) avant de decouvrir que le classeur ne pourra pas etre
    construit a temps. Le cout ajoute est un agregat sur les memes filtres,
    negligeable devant la construction openpyxl qu'il protege.

    POURQUOI UN PLAFOND : un export coute ~1,25 ms de temps serveur par ligne
    (75 s pour 60 000 lignes, perf-exports-20260827.md), et l'arbitre gunicorn
    tue le worker a 120 s (`--timeout`, docker-compose*.yml) en emportant les
    requetes de ses voisins — l'UvicornWorker est partage. Au-dela du plafond,
    un refus immediat vaut mieux qu'un worker tenu deux minutes pour un fichier
    que le client ne recevra pas : nginx a rendu la main a 130 s, et
    l'utilisateur a deja reclique (motif observe dans les tirs de charge).

    Le plafond par defaut (60 000, EXPORT_MAX_ROWS) est conservateur : la mesure
    des 1,25 ms/ligne est ANTERIEURE au cache de styles de `_build_list_sheet`
    ci-dessous, qui a supprime 14,9 s sur les 18 s de construction de 4 800
    lignes. Le cout par ligne reel est donc plus bas aujourd'hui, et le plafond
    sera reevalue a la premiere mesure post-correctif plutot que devine ici.

    La trace `EXPORT_COUNT` est le pendant amont de `EXPORT_WORKBOOK`
    (utils/excel_io.py) : ensemble elles donnent, par type d'export, la
    volumetrie d'entree et de sortie. C'est ce qui permettra de placer le seuil
    de bascule en generation asynchrone sur une distribution reelle plutot que
    sur une intuition (phase 2 de
    docs/architecture-exports-asynchrones-20260828.md).

    CLOISONNEMENT — a lire avant de reutiliser ce helper. `select(func.count())
    .select_from(requete.subquery())` produit une sous-requete Core : les
    `with_loader_criteria` poses par le listener multi-tenant (db/session.py)
    ne s'y appliquent pas necessairement. Ce comptage ne doit donc rien leur
    devoir, et il ne leur doit rien : chaque requete d'export porte son propre
    filtre d'organisation explicite (`Encaissement.organisation_id ==
    user.organisation_id` et equivalents ; pour /budget, l'exercice a deja ete
    resolu sous filtre d'organisation), ou ne porte sur aucune table
    multi-tenant (`ExpertComptable` est un registre national, il n'a pas de
    colonne `organisation_id`). Un export ajoute plus tard sans filtre explicite
    compterait large : le classeur resterait cloisonne — c'est le listener qui
    le garantit — mais le refus se declencherait sur le volume des autres
    organisations. Le correctif est alors d'ajouter le filtre, pas de retirer ce
    comptage.
    """
    total = int(
        (
            await db.execute(
                select(func.count()).select_from(requete.order_by(None).subquery())
            )
        ).scalar_one()
    )
    plafond = settings.export_max_rows
    logger.info("EXPORT_COUNT export=%s lignes=%d plafond=%d", export, total, plafond)
    if plafond > 0 and total > plafond:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=(
                f"Cet export porte sur {_milliers(total)} lignes, au-delà de la "
                f"limite de {_milliers(plafond)} lignes qu'un export direct peut "
                "produire sans être interrompu. Restreignez la période ou les "
                "filtres, puis relancez l'export."
            ),
        )
    # ORDRE VOULU : le plafond d'abord, la bascule ensuite. Un export au-dela du
    # plafond est refuse ICI, a la soumission, plutot qu'accepte en 202 puis
    # echoue par le worker vingt minutes plus tard pour la meme raison. C'est
    # ce qui rend coherents les deux reglages : seuil < plafond.
    if seuil_bascule is not None and total > seuil_bascule:
        raise BasculeAsynchroneRequise(total)
    return total


async def _relacher_connexion(db: AsyncSession) -> None:
    """Rend la connexion au pool avant le travail CPU de generation du classeur.

    Mesure sur /exports/encaissements (4 800 lignes), avant ce relachement :

        duree 33 616 ms | SQL 730 ms (6 requetes) | connexion retenue 33 204 ms

    La connexion restait donc prise 45 fois plus longtemps que le temps SQL
    reel, pendant que le thread construisait le fichier. Avec pool_size=5 par
    worker, quelques exports simultanes suffisent a vider le pool — et les
    requetes qui echouent alors n'ont aucun rapport avec l'export :

        QueuePool limit of size 5 overflow 5 reached, connection timed out

    Ce relachement est sur ici parce que deux conditions sont reunies :
      1. `expire_on_commit=False` (app/db/session.py:99) : les attributs deja
         charges restent lisibles apres la fin de la transaction ;
      2. tout ce que la closure de construction lit est charge en amont
         (joinedload explicites, entites ramenees par la requete elle-meme).

    Si un export venait a lire un attribut NON precharge apres cet appel,
    SQLAlchemy leverait une erreur explicite plutot que de recharger en douce :
    la panne serait visible, pas silencieuse.

    POURQUOI `commit()` ET NON `rollback()` — la nuance a coute un 500 :
    `expire_on_commit=False` ne concerne que `commit()`. `rollback()`, lui,
    EXPIRE systematiquement tous les objets de la session pour la ramener a un
    etat propre. La premiere lecture d'attribut dans le thread declenchait donc
    un rechargement, hors contexte greenlet :

        sqlalchemy.exc.MissingGreenlet: greenlet_spawn has not been called;
        can't call await_only() here. Was IO attempted in an unexpected place?

    ...et ce sur une colonne ordinaire (`enc.montant_total`), pas une relation.
    `commit()` rend la connexion sans expirer : les valeurs deja chargees
    restent lisibles. L'export est en lecture seule, il n'y a rien a ecrire.
    """
    await db.commit()


async def _refuser_si_abonnement_suspendu(db: AsyncSession, organisation_id: int) -> None:
    """Applique au passage en file la meme regle qu'aux ecritures.

    POURQUOI CE CONTROLE EXISTE ICI. Le garde-fou d'abonnement de `deps.py` ne
    s'applique qu'aux methodes d'ecriture (`if request.method in {POST, PUT,
    PATCH, DELETE}`). Or `GET /exports/<type>` ECRIT desormais une ligne
    `export_jobs` des qu'un type est bascule : une organisation passee en
    lecture seule pouvait continuer a mettre des exports en file, donc a
    consommer le worker. C'est un contournement du passage en lecture seule,
    ouvert par la bascule asynchrone elle-meme.

    Le chemin synchrone n'est PAS concerne et ne doit pas l'etre : lire ses
    propres donnees reste permis en lecture seule. C'est bien la creation du
    job qui est une ecriture.

    La resolution du statut reprend celle de `deps.py`, dans le meme ordre :
    la console SaaS fait autorite quand elle repond, l'organisation sinon. Deux
    resolutions differentes finiraient par diverger, et un tenant serait
    suspendu d'un cote et actif de l'autre. D'ou aussi la reutilisation de
    `_normalize_plan_status` plutot qu'un second `.strip().upper()`.
    """
    statut = await get_cached_saas_status(organisation_id)
    if not statut:
        statut = (
            await db.execute(
                select(Organisation.status_abonnement).where(Organisation.id == organisation_id)
            )
        ).scalar_one_or_none()
    if _normalize_plan_status(statut) not in {"ACTIVE", "TRIAL"}:
        raise HTTPException(
            status_code=status.HTTP_402_PAYMENT_REQUIRED,
            detail="Abonnement expiré. Passage en lecture seule. Veuillez régulariser via ePaieLink.",
        )


async def _soumettre_export_asynchrone(
    db: AsyncSession,
    user: User,
    *,
    type_export: str,
    params: dict[str, Any],
    row_count: int | None = None,
) -> JSONResponse:
    """Cree (ou reutilise) un job et repond 202 avec de quoi le suivre.

    Toujours 202, meme quand un artefact identique et recent est reutilise et
    que le job est deja `DONE`. Deux codes differents obligeraient le client a
    deux chemins ; un seul code et un `status` dans le corps lui en laissent un.
    Un job deja termine porte son `download_path` : le client telecharge sans
    attendre.

    ORDRE CRITIQUE — `commit()` PUIS publication. Publier avant de committer
    ouvre une fenetre pendant laquelle le worker recoit l'identifiant d'une
    ligne que sa propre transaction ne voit pas encore : il conclurait « job
    introuvable » et l'export serait perdu alors que la demande, elle, existe.
    L'ordre inverse ne perd rien — si la publication echoue, le job reste
    `QUEUED` et le balayage de reconciliation le reprend.

    `expire_on_commit=False` (db/session.py:99) rend les attributs de `job`
    encore lisibles apres le commit : la serialisation qui suit ne declenche
    aucun rechargement.
    """
    if user.organisation_id is None:
        # Un super-admin hors organisation n'a pas de tenant a exporter, et un
        # job sans organisation serait un job dont le worker ne saurait pas quel
        # contexte poser.
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Aucune organisation selectionnee pour cet export.",
        )

    await _refuser_si_abonnement_suspendu(db, user.organisation_id)

    # Les filtres non renseignes sont retires AVANT l'empreinte : sinon
    # `{"annee": 2026}` et `{"annee": 2026, "type": None}` produiraient deux
    # empreintes differentes pour le meme export, et la deduplication ne
    # dedupliquerait rien.
    params_nets = {cle: valeur for cle, valeur in params.items() if valeur is not None}

    job, reutilise = await soumettre(
        db,
        organisation_id=user.organisation_id,
        requested_by=user.id,
        type_export=type_export,
        params=params_nets,
        row_count=row_count,
    )
    await db.commit()
    if not reutilise and not await publier(str(job.id), tentative=job.attempts + 1):
        # Le job N'EST PAS perdu : il reste `QUEUED`, et le balayage du worker
        # republie les orphelins. Mais il attendra EXPORT_JOB_LEASE_SECONDS
        # avant de demarrer. Sans cette trace, un Redis absent se lit, cote
        # utilisateur, comme « l'export met cinq minutes a demarrer », et rien
        # dans les journaux de l'API ne dit pourquoi — `export_queue` a bien
        # trace la panne Redis, mais pas quel job en a fait les frais.
        logger.warning(
            "EXPORT_JOB_NON_PUBLIE job=%s type=%s org=%s : cree mais non publie, "
            "reprise au prochain balayage (jusqu'a %s s d'attente).",
            job.id,
            type_export,
            user.organisation_id,
            settings.export_job_lease_seconds,
        )
    return JSONResponse(status_code=status.HTTP_202_ACCEPTED, content=serialiser_job(job))


# Liste BLANCHE, et non liste noire des caracteres genants : le nom de fichier
# des exports est bati a partir de parametres de requete bruts
# (`f"requisitions_{date_debut}_{date_fin}.xlsx"`), et `_parse_datetime` rend
# `None` sur une date invalide SANS refuser la requete — la chaine arbitraire
# arrive donc telle quelle dans l'en-tete.
_NOM_FICHIER_AUTORISE = re.compile(r"[^A-Za-z0-9._-]")


def entete_piece_jointe(filename: str) -> str:
    """Valeur de `Content-Disposition` sure pour un nom de fichier quelconque.

    Deux pieges, tous deux payes d'un 500 AVANT que la reponse n'existe, donc
    sans classeur pour l'utilisateur :

    1. Starlette encode les en-tetes en latin-1. Un caractere hors de ce jeu
       (`?date_debut=€`) leve `UnicodeEncodeError` a la construction de la
       reponse.
    2. Starlette laisse passer `\\r\\n` dans une valeur d'en-tete ; c'est h11
       qui refuse ensuite la reponse. Un `%0d%0a` dans un parametre suffisait
       donc a casser l'export, et sur un serveur plus permissif ce serait une
       scission de reponse.

    Le nom est aussi mis entre guillemets, comme dans `secure_uploads.py` : sans
    eux, tout ce qui suit un espace est ignore par le navigateur.
    """
    propre = _NOM_FICHIER_AUTORISE.sub("_", filename).strip("._")
    return f'attachment; filename="{propre or "export.xlsx"}"'


async def _excel_response(filename: str, wb: Workbook) -> StreamingResponse:
    return StreamingResponse(
        await save_workbook(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": entete_piece_jointe(filename)},
    )


def _round_money(value: Decimal | float | int | None) -> Decimal:
    if value is None:
        return Decimal("0.00")
    return Decimal(str(value)).quantize(Decimal("0.01"))


def _person_name(u) -> str:
    """Nom affichable d'un utilisateur, avec repli sur l'email puis l'identifiant."""
    if not u:
        return ""
    full = f"{u.prenom or ''} {u.nom or ''}".strip()
    return full or u.email or str(u.id)


def _format_mode_paiement(value: str | None) -> str:
    if not value:
        return ""
    val = value.strip().lower()
    mapping = {
        "cash": "Cash",
        "mobile_money": "Mobile Money",
        "virement": "Opération bancaire",
        "card": "Carte (Visa)",
        "cheque": "Chèque",
    }
    return mapping.get(val, value)


def _source_columns(
    operation_type: str, canal: str | None, banque_nom: str | None, compte_numero: str | None
) -> list[str]:
    """Les quatre colonnes de source, à partir de libellés déjà résolus.

    Une source dont les libellés viennent d'ailleurs qu'un objet `CompteBancaire`
    chargé — un transfert du moteur dédié, par exemple — passe par ici plutôt que
    de reconstruire ces colonnes de son côté.
    """
    source = "Banque" if (canal or "").upper() == "BANQUE" else "Caisse"
    if source != "Banque" or not (banque_nom or compte_numero):
        return [operation_type, source, "—", "—"]
    return [operation_type, source, banque_nom or "—", compte_numero or "—"]


def _financial_source_columns(operation_type: str, canal: str | None, compte_bancaire: Any | None) -> list[str]:
    return _source_columns(
        operation_type,
        canal,
        getattr(getattr(compte_bancaire, "banque", None), "nom", None) if compte_bancaire else None,
        getattr(compte_bancaire, "numero_compte", None) if compte_bancaire else None,
    )


def _sort_key_datetime(value: datetime | None) -> datetime:
    """Clé de tri robuste : les dates viennent de tables différentes et peuvent
    être naïves ou aware. Les comparer telles quelles lèverait un TypeError et
    ferait échouer l'export entier — on les ramène toutes en UTC aware."""
    if value is None:
        return datetime(1970, 1, 1, tzinfo=timezone.utc)
    return value if value.tzinfo is not None else value.replace(tzinfo=timezone.utc)


def _format_operation_time(operation_dt: datetime | None, created_at: datetime | None) -> str:
    """Date métier + input HTML date => heure 00:00. Dans ce cas, afficher
    l'heure réelle de création de l'opération."""
    if operation_dt and (
        operation_dt.hour != 0
        or operation_dt.minute != 0
        or operation_dt.second != 0
        or operation_dt.microsecond != 0
    ):
        return operation_dt.strftime("%H:%M")
    return created_at.strftime("%H:%M") if created_at else ""


def _autosize_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


# ── Palette & styles partagés (repris du modèle « budget ») ────────────────────
# Promus au niveau module pour être réutilisés par tous les exports sans être
# redéfinis (une seule source de vérité pour la charte visuelle).
GREEN = "FF065F46"
GREEN_DARK = "FF064E3B"
GREEN_LIGHT = "FFD1FAE5"
TEAL_SOFT = "FFCCFBF1"
AMBER_SOFT = "FFFEF3C7"
RED_SOFT = "FFFEE2E2"
SLATE = "FF334155"
SLATE_LIGHT = "FFF8FAFC"
LEVEL_FILLS = ["FF6EE7B7", "FFA7F3D0", "FFC6F6DF", "FFD1FAE5"]
header_font = Font(bold=True, color="FFFFFFFF", size=10)
header_fill = PatternFill(fill_type="solid", fgColor=GREEN)
subheader_fill = PatternFill(fill_type="solid", fgColor=GREEN_LIGHT)
muted_fill = PatternFill(fill_type="solid", fgColor=SLATE_LIGHT)
# Transferts internes caisse <-> banque : ni une dépense, ni une recette. Teinte
# franche mais douce, distincte du zébrage gris et du bandeau vert.
transfert_fill = PatternFill(fill_type="solid", fgColor=TEAL_SOFT)
# Types de sortie qui ne font pas sortir l'argent de l'organisation.
TRANSFERT_TYPES = ("versement_banque", "approvisionnement_caisse")
white_fill = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")
thin = Side(style="thin", color="FFD1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = "#,##0.00"
PCT = '0.0"%"'


async def _tenant_display_name(db: AsyncSession, organisation_id: int) -> str:
    return await tenant_display_name(db, organisation_id)


# Marqueur de l'horodatage de generation. Il sert deux fois : ici pour ecrire
# la mention, et dans observe/comparer_classeurs.py pour la NEUTRALISER lors de
# la comparaison de deux classeurs. Les deux doivent rester identiques — d'ou
# une constante et non deux litteraux.
MENTION_GENERATION = "Généré le"


def _fuseau_documents() -> ZoneInfo:
    """Fuseau de l'horodatage des documents, avec repli sur celui des rapports."""
    nom = (settings.document_timezone or settings.weekly_report_timezone or "UTC").strip() or "UTC"
    try:
        return ZoneInfo(nom)
    except Exception:
        logger.warning("DOCUMENT_TIMEZONE=%s invalide, repli sur UTC", nom)
        return ZoneInfo("UTC")


def horodatage_generation(maintenant: datetime | None = None) -> str:
    """« Généré le 29/08/2026 à 14:32 » — l'instant ou le classeur est construit.

    POURQUOI CETTE MENTION EXISTE. Un export asynchrone reflete les donnees au
    moment de sa GENERATION, pas du clic : le job peut demarrer plusieurs
    minutes apres la demande, et la deduplication peut rendre un artefact
    produit une demi-heure plus tot. Sans cette mention, un classeur imprime ne
    dit pas a quel instant ses chiffres etaient vrais — sur des pieces
    comptables, c'est une ambiguite que la bascule asynchrone introduirait
    sans le dire.

    La mention vaut aussi pour le chemin synchrone : deux regimes qui
    horodatent differemment seraient pires que deux regimes qui n'horodatent
    pas du tout.
    """
    instant = (maintenant or datetime.now(timezone.utc)).astimezone(_fuseau_documents())
    return f"{MENTION_GENERATION} {instant.strftime('%d/%m/%Y à %H:%M')} ({instant.tzname()})"


def _write_banner(ws, title: str, subtitle: str | None, ncols: int, organisation: str) -> None:
    """Bandeau titre vert (ligne 1), organisation émettrice (ligne 2) et
    sous-titre italique (ligne 3), fusionnés sur les ``ncols`` premières
    colonnes. À appeler APRÈS ``_autosize_columns`` pour ne pas gonfler la
    largeur de la colonne A avec le texte du titre.

    ``organisation`` est OBLIGATOIRE : aucun document ne doit sortir de
    l'application sans identifier le tenant qui l'émet. Les lignes 1 à 3 sont
    réservées à ce bandeau, l'en-tête des données commençant en ligne 4.
    """
    last_col = get_column_letter(max(ncols, 1))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = organisation
    ws["A2"].font = Font(bold=True, size=11, color=GREEN_DARK)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # L'horodatage est ajoute au sous-titre plutot que sur une ligne a lui :
    # les lignes 1 a 3 sont reservees au bandeau et l'en-tete des donnees
    # commence en ligne 4, sur les cinq exports et toutes leurs feuilles. Une
    # quatrieme ligne aurait decale toutes les references de plage.
    ligne_generation = horodatage_generation()
    subtitle = f"{subtitle}  |  {ligne_generation}" if subtitle else ligne_generation

    if subtitle:
        ws.merge_cells(f"A3:{last_col}3")
        ws["A3"] = subtitle
        ws["A3"].font = Font(italic=True, color=SLATE)
        ws["A3"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[3].height = 18


def _build_list_sheet(
    ws,
    *,
    title: str,
    subtitle: str | None,
    headers: list[str],
    data_rows: list[list[Any]],
    organisation: str,
    money_cols: tuple[int, ...] = (),
    total_values: dict[int, Any] | None = None,
    ordinal: bool = True,
    highlight_rows: frozenset[int] | set[int] = frozenset(),
    highlight_fill: PatternFill | None = None,
    highlight_row_fills: dict[int, PatternFill] | None = None,
) -> int:
    """Construit une feuille « liste » au style budget : bandeau titre, en-tête
    vert, lignes zébrées, formats monétaires, ligne TOTAL, en-tête figé, filtre
    automatique et largeurs auto. Renvoie l'index de la ligne TOTAL.

    Si ``ordinal`` (défaut), une colonne « N° » (1, 2, 3…) est ajoutée en tête pour
    pouvoir référencer chaque ligne dans les rapports ; les index monétaires et de
    total sont décalés automatiquement (+1).

    ``highlight_rows`` (index dans ``data_rows``) reçoit ``highlight_fill`` à la
    place du zébrage : sert à distinguer d'un coup d'œil une catégorie de lignes
    mêlée aux autres. La couleur doit toujours DOUBLER une information écrite
    (une colonne qui nomme la catégorie), jamais la remplacer."""
    if ordinal:
        headers = ["N°", *headers]
        data_rows = [[idx + 1, *row] for idx, row in enumerate(data_rows)]
        money_cols = tuple(c + 1 for c in money_cols)
        if total_values is not None:
            total_values = {k + 1: v for k, v in total_values.items()}
    ncols = len(headers)
    last_col_letter = get_column_letter(max(ncols, 1))
    money = set(money_cols)
    HEADER_ROW = 4
    first_data = HEADER_ROW + 1

    # En-tête stylé (vert + blanc gras)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=HEADER_ROW, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[HEADER_ROW].height = 24

    # openpyxl re-serialise et re-hache l'objet de style COMPLET a chaque
    # affectation (`cell.border = ...`), meme lorsque c'est le meme objet
    # partage. Profil mesure sur 4 800 lignes (observe/profil_export.py) :
    #
    #   _build_list_sheet ................. 18,0 s
    #     styleable.__set__ (276 358 app.)  14,9 s
    #       Serialisable.__hash__ (2,5 M)   13,4 s
    #
    # Une cellule stylee ne porte en interne qu'un `StyleArray` : des index vers
    # les tables de styles du classeur. Deux cellules d'apparence identique
    # partagent exactement le meme tuple d'index. On paie donc le cout openpyxl
    # UNE FOIS par combinaison distincte, puis on recopie le StyleArray obtenu.
    #
    # Le rendu est identique : memes index, donc memes bordures, alignements,
    # formats et remplissages. Aucune regle d'affichage n'est modifiee ici.
    modeles_style: dict[tuple, Any] = {}

    def _styler(cellule, cle, appliquer) -> None:
        modele = modeles_style.get(cle)
        if modele is None:
            appliquer(cellule)
            modeles_style[cle] = _copier_style(cellule._style)
        else:
            cellule._style = _copier_style(modele)

    # Lignes de données zébrées, sauf les lignes mises en évidence.
    for ridx, row_values in enumerate(data_rows):
        r = first_data + ridx
        row_fill = (highlight_row_fills or {}).get(ridx)
        highlighted = row_fill is not None or (highlight_fill is not None and ridx in highlight_rows)
        zebra = ridx % 2 == 1
        for i, val in enumerate(row_values, start=1):
            c = ws.cell(row=r, column=i, value=val)
            est_money = i in money
            est_ordinal = ordinal and i == 1
            if highlighted:
                # Les fills de mise en evidence peuvent differer d'une ligne a
                # l'autre (highlight_row_fills) : l'identite de l'objet suffit a
                # les distinguer, et deux lignes partageant le meme fill
                # partagent la meme entree de cache.
                cle_remplissage = id(row_fill or highlight_fill)
            elif zebra:
                cle_remplissage = "zebre"
            else:
                cle_remplissage = None

            def appliquer(
                cellule,
                _money=est_money,
                _ordinal=est_ordinal,
                _remplissage=cle_remplissage,
                _row_fill=row_fill,
            ) -> None:
                cellule.border = border
                if _money:
                    cellule.number_format = MONEY
                    cellule.alignment = right
                elif _ordinal:
                    cellule.alignment = center
                else:
                    cellule.alignment = left
                if _remplissage == "zebre":
                    cellule.fill = muted_fill
                elif _remplissage is not None:
                    cellule.fill = _row_fill or highlight_fill

            _styler(c, (est_money, est_ordinal, cle_remplissage), appliquer)

    last_data = HEADER_ROW + len(data_rows)
    total_row = last_data + 1

    # Ligne TOTAL (bandeau vert, gras blanc)
    if total_values is not None:
        ws.cell(row=total_row, column=1, value="TOTAL")
        for col_idx, value in total_values.items():
            ws.cell(row=total_row, column=col_idx, value=value)
        for col in range(1, ncols + 1):
            c = ws.cell(row=total_row, column=col)
            c.font = Font(bold=True, color="FFFFFFFF")
            c.fill = header_fill
            c.border = border
            c.alignment = right if col in money else center
            if col in money:
                c.number_format = MONEY

    # Largeurs auto AVANT le bandeau titre (le titre ne doit pas élargir A)
    _autosize_columns(ws)
    _write_banner(ws, title, subtitle, ncols, organisation)

    ws.freeze_panes = f"A{first_data}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col_letter}{max(last_data, HEADER_ROW)}"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GREEN
    return total_row


def _write_synthese_block(
    ws,
    start_row: int,
    block_title: str,
    col_headers: list[str],
    rows: list[list[Any]],
    money_cols: tuple[int, ...] = (),
) -> tuple[int, int, int, int]:
    """Écrit un bloc « titre + en-tête + lignes » stylé sur la feuille Synthèse.
    Renvoie ``(header_row, first_data_row, last_data_row, next_free_row)``."""
    money = set(money_cols)
    ncols = len(col_headers)
    tcell = ws.cell(row=start_row, column=1, value=block_title)
    tcell.font = Font(bold=True, color=GREEN_DARK, size=11)
    for col in range(1, ncols + 1):
        c = ws.cell(row=start_row, column=col)
        c.fill = subheader_fill
        c.border = border
    header_row = start_row + 1
    for i, h in enumerate(col_headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    first = header_row + 1
    for ridx, row_values in enumerate(rows):
        r = first + ridx
        for i, val in enumerate(row_values, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.border = border
            if i in money:
                c.number_format = MONEY
                c.alignment = right
            else:
                c.alignment = left if i == 1 else center
            c.fill = muted_fill if ridx % 2 == 1 else white_fill
    last = header_row + len(rows)
    return header_row, first, last, last + 2


def _build_synthese_sheet(
    wb: Workbook,
    banner_title: str,
    blocks: list[dict[str, Any]],
    *,
    chart_title: str | None = None,
    chart_block_index: int = 0,
    chart_value_col: int = 2,
    organisation: str,
) -> None:
    """Crée la feuille « Synthèse » (blocs de totaux + un BarChart), dans le
    même esprit que la feuille de synthèse du modèle budget."""
    ws = wb.create_sheet("Synthèse")
    ws.sheet_properties.tabColor = "FF0F766E"
    ws.sheet_view.showGridLines = False
    row = 3
    refs: list[tuple[int, int, int]] = []
    for b in blocks:
        header_row, first, last, nxt = _write_synthese_block(
            ws, row, b["title"], b["headers"], b["rows"], b.get("money_cols", ())
        )
        refs.append((header_row, first, last))
        row = nxt

    if chart_title and blocks and blocks[chart_block_index]["rows"]:
        header_row, first, last = refs[chart_block_index]
        bar = BarChart()
        bar.type = "col"
        bar.title = chart_title
        bar.height = 8
        bar.width = 16
        bar.add_data(
            Reference(ws, min_col=chart_value_col, min_row=header_row, max_row=last),
            titles_from_data=True,
        )
        bar.set_categories(Reference(ws, min_col=1, min_row=first, max_row=last))
        ws.add_chart(bar, "F3")

    _autosize_columns(ws)
    _write_banner(ws, banner_title, None, 8, organisation)


def _budget_code_key(value: str | None) -> str:
    """Clé d'appariement d'un code de poste budgétaire.

    Mêmes règles que `_normalize_budget_code` côté API budget, casse repliée en
    plus : un commentaire ancré sur « I.7.1 » doit retrouver le poste « i.7.1 ».
    """
    if not value:
        return ""
    code = re.sub(r"\s+", "", value.strip())
    code = re.sub(r"\.+", ".", code).strip(".")
    return code.lower()


async def construire_classeur_budget(
    db: AsyncSession,
    organisation_id: int,
    *,
    annee: int | None = None,
    type: str | None = None,
    service_id: int | None = None,
    seuil_bascule: int | None = None,
) -> tuple[Workbook, str]:
    """Construit le classeur budgetaire et rend `(classeur, nom de fichier)`.

    EXTRAIT DE L'ENDPOINT, sans autre changement que le remplacement de
    `user.organisation_id` par `organisation_id` : le corps est celui qui
    tournait jusqu'ici derriere `GET /exports/budget`.

    Pourquoi cette extraction : le worker d'exports doit produire EXACTEMENT le
    meme fichier que le chemin synchrone. Deux implementations, meme guidees par
    la meme specification, divergent — c'est une question de temps, pas de soin.
    Une seule fonction appelee par les deux chemins rend la divergence
    impossible plutot qu'improbable, et c'est ce que verifie
    `observe/comparer_classeurs.py` cellule par cellule.

    La fonction ne prend PAS de `User` : elle n'en lisait que
    `organisation_id`. Un worker n'a pas de requete, donc pas d'utilisateur
    courant ; lui en fabriquer un serait inventer un contexte pour satisfaire
    une signature.
    """
    if annee is None:
        result = await db.execute(select(func.max(BudgetExercice.annee)).where(BudgetExercice.organisation_id == organisation_id))
        annee = result.scalar_one_or_none()

    if annee is None:
        raise HTTPException(status_code=404, detail="Aucun exercice budgétaire disponible")

    exercice_res = await db.execute(select(BudgetExercice).where(
        BudgetExercice.annee == annee,
        BudgetExercice.organisation_id == organisation_id
    ))
    exercice = exercice_res.scalar_one_or_none()
    if exercice is None:
        raise HTTPException(status_code=404, detail="Exercice introuvable")

    query = select(BudgetPoste).where(
        BudgetPoste.exercice_id == exercice.id,
        BudgetPoste.is_deleted.is_(False),
    )
    filtre_type = type.upper() if type else None
    if filtre_type and filtre_type != "TOUT":
        query = query.where(BudgetPoste.type == filtre_type)
    query = query.order_by(BudgetPoste.code)
    await _compter_lignes(db, query, export="budget", seuil_bascule=seuil_bascule)
    lignes = list((await db.execute(query)).scalars().all())
    lignes.sort(key=lambda poste: cle_tri_code_budget(poste.code))
    # Identification du tenant émetteur : obligatoire sur tout document exporté.
    organisation = await _tenant_display_name(db, organisation_id)

    service_label: str | None = None
    if service_id is not None:
        service_res = await db.execute(
            select(Service).where(
                Service.id == service_id,
                Service.organisation_id == organisation_id,
            )
        )
        service = service_res.scalar_one_or_none()
        if service is not None:
            service_label = f"{service.code} - {service.libelle}"

    # Filtre par service : on ne garde que les rubriques rattachées au service
    # (via ServiceRubrique) et leurs postes parents, pour préserver la hiérarchie.
    if service_id is not None:
        rub_res = await db.execute(
            select(ServiceRubrique.budget_poste_id).where(ServiceRubrique.service_id == service_id)
        )
        allowed = set(rub_res.scalars().all())
        by_id_all = {p.id: p for p in lignes}
        keep: set[int] = set()
        for pid in allowed:
            cur = by_id_all.get(pid)
            while cur is not None and cur.id not in keep:
                keep.add(cur.id)
                cur = by_id_all.get(cur.parent_id) if cur.parent_id else None
        lignes = [p for p in lignes if p.id in keep]

    recette_ids = [p.id for p in lignes if (p.type or "").upper() == "RECETTE"]
    recettes_affichees: dict[int, Decimal] = {}
    if recette_ids:
        conditions = [
            Encaissement.organisation_id == organisation_id,
            Encaissement.budget_poste_id.in_(recette_ids),
            Encaissement.est_proforma.is_(False),
            Encaissement.is_deleted.is_(False),
            (Encaissement.statut_operation.is_(None)) | (Encaissement.statut_operation == "ACTIVE"),
        ]
        if service_id is not None:
            conditions.append(Encaissement.service_id == service_id)
        recettes_res = await db.execute(
            select(
                Encaissement.budget_poste_id,
                func.coalesce(func.sum(func.coalesce(Encaissement.montant_paye, 0)), 0),
            )
            .where(*conditions)
            .group_by(Encaissement.budget_poste_id)
        )
        recettes_actives = {int(row[0]): Decimal(row[1] or 0) for row in recettes_res.all() if row[0]}
        # Surcharge d'AFFICHAGE, pas d'écriture. La version précédente affectait
        # `poste.montant_engage` / `poste.montant_paye` sur les entités de la
        # session : les postes devenaient sales, et le premier `db.execute`
        # suivant (commentaires, exercice précédent) déclenchait l'autoflush,
        # donc un `UPDATE budget_postes` par poste recette au milieu d'un GET.
        # Mesuré le 27/08 sous charge : `UPDATE budget_postes SET montant_engage,
        # montant_paye WHERE id = $3` relevé à 11,5 s pour UN seul poste, dans un
        # export lui-même tracé à 168 s — donc des verrous de ligne tenus tout ce
        # temps contre les écritures réelles du tenant. Rien n'est persisté (la
        # transaction est annulée en fin de requête) : seul le verrou coûte.
        # Compté hors charge : 76 lignes écrites par appel, 0 après ce correctif
        # (`pg_stat_user_tables.n_tup_upd`), pour un classeur identique.
        # Un dictionnaire local produit les mêmes valeurs sans toucher la base.
        recettes_affichees = {
            poste.id: recettes_actives.get(poste.id, Decimal("0"))
            for poste in lignes
            if (poste.type or "").upper() == "RECETTE"
        }

    # ── Arbre hiérarchique : un poste parent = somme de ses sous-postes ────────
    by_id = {p.id: p for p in lignes}
    children_map: dict[int | None, list] = {}
    for p in lignes:
        pid = p.parent_id if (p.parent_id in by_id) else None
        children_map.setdefault(pid, []).append(p)
    for kids in children_map.values():
        kids.sort(key=lambda x: cle_tri_code_budget(x.code))

    totals_cache: dict[int, tuple[Decimal, Decimal, Decimal]] = {}

    def est_inclus(p) -> bool:
        """Ligne comptée dans les totaux. Le drapeau est propagé à la branche
        entière côté API, un test par ligne suffit donc ici."""
        return bool(getattr(p, "inclure_dans_calculs", True))

    def node_totals(p) -> tuple[Decimal, Decimal, Decimal]:
        if p.id in totals_cache:
            return totals_cache[p.id]
        kids = children_map.get(p.id, [])
        if kids:
            prevu = engage = paye = Decimal(0)
            for k in kids:
                kp, ke, kpy = node_totals(k)
                if not est_inclus(k):
                    continue
                prevu += kp
                engage += ke
                paye += kpy
        else:
            prevu = Decimal(p.montant_prevu or 0)
            surcharge = recettes_affichees.get(p.id)
            if surcharge is not None:
                # Poste de recette : le cumul des encaissements actifs remplace
                # les deux montants, comme le faisait l'affectation d'origine.
                engage = paye = surcharge
            else:
                engage = Decimal(p.montant_engage or 0)
                paye = Decimal(p.montant_paye or 0)
        totals_cache[p.id] = (prevu, engage, paye)
        return totals_cache[p.id]

    ordered: list[tuple[Any, int, bool]] = []

    def walk(nodes, depth: int) -> None:
        for n in nodes:
            kids = children_map.get(n.id, [])
            ordered.append((n, depth, bool(kids)))
            if kids:
                walk(kids, depth + 1)

    walk(children_map.get(None, []), 0)

    # ── Fil de commentaires : même donnée que le PDF annoté ────────────────────
    # Restitué en commentaires Excel natifs, posés sur la ligne du poste. Le
    # texte seul : l'auteur et la date restent dans l'application, où le fil se
    # discute. Le classeur circule à l'extérieur, il porte la justification du
    # montant, pas la signature de qui l'a écrite.
    comm_res = await db.execute(
        select(BudgetPosteCommentaire)
        .where(
            BudgetPosteCommentaire.organisation_id == organisation_id,
            BudgetPosteCommentaire.exercice_id == exercice.id,
        )
        .order_by(BudgetPosteCommentaire.created_at.asc())
    )
    # Seuls les commentaires des postes réellement exportés sont retenus : un
    # fil attaché à une recette n'a rien à faire dans l'export des dépenses,
    # ni dans un export filtré par service.
    codes_exportes = {_budget_code_key(p.code) for (p, _, _) in ordered}
    commentaires_par_code: dict[str, list[str]] = defaultdict(list)
    for comm in comm_res.scalars().all():
        cle = _budget_code_key(comm.code)
        if not cle or cle not in codes_exportes:
            continue
        texte = (comm.texte or "").strip()
        if texte:
            commentaires_par_code[cle].append(texte)

    # ── Exercice N-1 : même comparaison qu'à l'écran ───────────────────────────
    # L'écran rapproche chaque poste de son homologue de l'exercice précédent,
    # apparié sur le code — jamais sur l'identifiant, qui change d'un exercice à
    # l'autre. L'export reprend le même appariement, sans quoi les deux vues
    # afficheraient des écarts différents pour la même ligne.
    prev_par_code: dict[str, Decimal] = {}
    prev_annee = annee - 1
    prev_ex_res = await db.execute(
        select(BudgetExercice.id).where(
            BudgetExercice.annee == prev_annee,
            BudgetExercice.organisation_id == organisation_id,
        )
    )
    prev_exercice_id = prev_ex_res.scalar_one_or_none()
    if prev_exercice_id is not None:
        prev_query = select(BudgetPoste).where(
            BudgetPoste.exercice_id == prev_exercice_id,
            BudgetPoste.is_deleted.is_(False),
        )
        if filtre_type and filtre_type != "TOUT":
            prev_query = prev_query.where(BudgetPoste.type == filtre_type)
        prev_lignes = list((await db.execute(prev_query)).scalars().all())

        prev_by_id = {p.id: p for p in prev_lignes}
        prev_children: dict[int | None, list] = {}
        for p in prev_lignes:
            pid = p.parent_id if (p.parent_id in prev_by_id) else None
            prev_children.setdefault(pid, []).append(p)

        prev_cache: dict[int, Decimal] = {}

        def prev_total(p) -> Decimal:
            """Prévu roulé : un parent vaut la somme de ses enfants inclus."""
            if p.id in prev_cache:
                return prev_cache[p.id]
            kids = prev_children.get(p.id, [])
            if kids:
                montant = sum(
                    (prev_total(k) for k in kids if getattr(k, "inclure_dans_calculs", True)),
                    Decimal(0),
                )
            else:
                montant = Decimal(p.montant_prevu or 0)
            prev_cache[p.id] = montant
            return montant

        for poste_prev in prev_lignes:
            cle = _budget_code_key(poste_prev.code)
            if cle:
                prev_par_code[cle] = prev_total(poste_prev)

    # ── Styles ─── constantes & objets de style promus au niveau module ────────
    def _pct(num: Decimal, den: Decimal) -> Decimal:
        return (num / den * Decimal(100)) if den > 0 else Decimal(0)

    def _build_workbook() -> tuple[Workbook, str]:
        wb = Workbook()
        ws = wb.active
        ws.title = f"Budget {annee}"
        ws.sheet_properties.tabColor = GREEN
        ws.sheet_view.showGridLines = False
        if wb.calculation is not None:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True

        vue_label = (
            "RECETTES" if filtre_type == "RECETTE"
            else "DÉPENSES" if filtre_type == "DEPENSE"
            else "GLOBAL"
        )
        # Un budget de recettes ne se « paie » pas et n'a pas de « disponible » :
        # on reprend le vocabulaire du PDF pour que les deux restitutions nomment
        # la même donnée pareil.
        is_recette = filtre_type == "RECETTE"
        LBL_PREVISION = "Prévision"
        LBL_REALISE = "Réalisation"
        LBL_SOLDE = "Solde budgétaire"
        LBL_TAUX = "Taux de réalisation"
        ws.merge_cells("A1:N1")
        ws["A1"] = f"BUDGET {vue_label} {annee}"
        ws["A1"].font = Font(bold=True, size=16, color="FFFFFFFF")
        ws["A1"].fill = header_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        # Ligne 2 : organisation émettrice, comme sur les autres feuilles exportées.
        # Aucun document ne sort sans identifier son tenant. Ligne 3 : sous-titre.
        # Les cartes de synthèse commencent en ligne 4, rien n'est décalé.
        ws.merge_cells("A2:N2")
        ws["A2"] = organisation
        ws["A2"].font = Font(bold=True, size=11, color=GREEN_DARK)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        ws.merge_cells("A3:N3")
        subtitle_parts = ["Suivi de l'exécution budgétaire par poste et sous-poste", "Montants en USD"]
        if service_label:
            subtitle_parts.insert(1, f"Service : {service_label}")
        ws["A3"] = " | ".join(subtitle_parts)
        ws["A3"].font = Font(italic=True, color=SLATE)
        ws["A3"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[3].height = 18

        # Les lignes hors calcul restent exportées, mais ne pèsent ni dans les
        # cartes de synthèse ni dans le total : c'est tout l'objet du drapeau.
        leaves = [(p, d, ip) for (p, d, ip) in ordered if not ip and est_inclus(p)]
        tot_prevu = sum((node_totals(p)[0] for p, _, _ in leaves), Decimal(0))
        tot_engage = sum((node_totals(p)[1] for p, _, _ in leaves), Decimal(0))
        tot_paye = sum((node_totals(p)[2] for p, _, _ in leaves), Decimal(0))
        # Recettes : le solde budgétaire se lit « réalisation − prévision »
        # (positif = objectif dépassé), comme dans le PDF. Dépenses : le
        # disponible se lit « prévision − payé ».
        tot_disp = (tot_paye - tot_prevu) if is_recette else (tot_prevu - tot_paye)
        tot_reste_engager = tot_prevu - tot_engage

        # Total N-1 sur le meme perimetre que tot_prevu : les feuilles hors
        # calcul en sont exclues, sinon la carte comparerait deux assiettes.
        tot_prevu_prev = sum(
            (prev_par_code.get(_budget_code_key(p.code), Decimal(0)) for p, _, _ in leaves),
            Decimal(0),
        )

        SUMMARY_LABEL_ROW = 4
        SUMMARY_VALUE_ROW = 5
        summary_cards = [
            (LBL_PREVISION, float(tot_prevu), MONEY, GREEN_LIGHT),
            ("Engagé", float(tot_engage), MONEY, TEAL_SOFT),
            (LBL_REALISE, float(tot_paye), MONEY, AMBER_SOFT),
            (LBL_SOLDE, float(tot_disp), MONEY, RED_SOFT if tot_disp < 0 else GREEN_LIGHT),
            (LBL_TAUX, float(_pct(tot_paye, tot_prevu)), PCT, SLATE_LIGHT),
            ("Sous-postes", len(leaves), None, SLATE_LIGHT),
            ("Budget N-1", float(tot_prevu_prev), MONEY, SLATE_LIGHT),
        ]
        for idx, (label, value, fmt, fill_color) in enumerate(summary_cards):
            start_col = 1 + idx * 2
            end_col = start_col + 1
            ws.merge_cells(start_row=SUMMARY_LABEL_ROW, start_column=start_col, end_row=SUMMARY_LABEL_ROW, end_column=end_col)
            ws.merge_cells(start_row=SUMMARY_VALUE_ROW, start_column=start_col, end_row=SUMMARY_VALUE_ROW, end_column=end_col)
            label_cell = ws.cell(row=SUMMARY_LABEL_ROW, column=start_col, value=label)
            value_cell = ws.cell(row=SUMMARY_VALUE_ROW, column=start_col, value=value)
            label_cell.font = Font(bold=True, color=SLATE, size=9)
            value_cell.font = Font(bold=True, color=GREEN_DARK, size=12)
            label_cell.alignment = center
            value_cell.alignment = center
            if fmt:
                value_cell.number_format = fmt
            for row in (SUMMARY_LABEL_ROW, SUMMARY_VALUE_ROW):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
                    cell.border = border

        ws.row_dimensions[4].height = 20
        ws.row_dimensions[5].height = 26

        HEADER_ROW = 7
        FIRST = 8
        headers = [
            "Code", "Nature", "Niveau", "Poste budgétaire", "Type",
            f"{LBL_PREVISION} (USD)", "Budget N-1 (USD)", "Solde budgétaire N/N-1 (USD)",
            "Engagé (USD)", f"{LBL_REALISE} (USD)", f"{LBL_SOLDE} (USD)",
            "Reste à engager (USD)", "Taux d'engagement %", f"{LBL_TAUX} %",
        ]
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=HEADER_ROW, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        ws.row_dimensions[HEADER_ROW].height = 24

        # Rangée Excel de chaque poste : un parent référence ses enfants, et le total
        # référence seulement les sous-postes pour éviter tout double comptage.
        row_of = {poste.id: FIRST + idx for idx, (poste, _, _) in enumerate(ordered)}

        for idx, (poste, depth, is_parent) in enumerate(ordered):
            r = FIRST + idx
            prevu, _engage, paye = node_totals(poste)
            taux_exec = _pct(paye, prevu)
            marker = ("»" * min(depth + 1, 3) + " ") if is_parent else ""
            libelle = ("    " * depth) + (poste.libelle or "")
            ws.cell(row=r, column=1, value=f"{marker}{poste.code or ''}")
            # Le SUMIF du total filtre sur ce libellé exact : une ligne hors calcul
            # porte une autre nature et sort donc du total sans formule spéciale.
            nature = "Poste parent" if is_parent else "Sous-poste"
            if not est_inclus(poste):
                nature = f"{nature} (hors calcul)"
            ws.cell(row=r, column=2, value=nature)
            ws.cell(row=r, column=3, value=depth)
            ws.cell(row=r, column=4, value=libelle)
            ws.cell(row=r, column=5, value=poste.type or "")
            if is_parent:
                krows = [row_of[k.id] for k in children_map.get(poste.id, []) if est_inclus(k)]
                for col_letter, col_idx in (("F", 6), ("I", 9), ("J", 10)):
                    ws.cell(
                        row=r,
                        column=col_idx,
                        # Aucun enfant inclus : le parent vaut zéro. Une formule
                        # « =+ » vide casserait l'ouverture du classeur.
                        value=("=" + "+".join(f"{col_letter}{k}" for k in krows)) if krows else 0,
                    )
            else:
                ws.cell(row=r, column=6, value=float(prevu))
                ws.cell(row=r, column=9, value=float(_engage))
                ws.cell(row=r, column=10, value=float(paye))
            # Budget N-1 : valeur litterale appariee sur le code, parents compris,
            # exactement comme a l'ecran. Une cellule vide dit « pas d'homologue
            # l'an dernier » ; un zero dirait « budgete a zero », ce qui est faux.
            montant_prev = prev_par_code.get(_budget_code_key(poste.code))
            if montant_prev is not None:
                ws.cell(row=r, column=7, value=float(montant_prev))
            ws.cell(row=r, column=8, value=f'=IF(G{r}="","",F{r}-G{r})')
            ws.cell(row=r, column=11, value=(f"=J{r}-F{r}" if is_recette else f"=F{r}-J{r}"))
            ws.cell(row=r, column=12, value=f"=F{r}-I{r}")
            ws.cell(row=r, column=13, value=f"=IF(F{r}>0,I{r}/F{r}*100,0)")
            ws.cell(row=r, column=14, value=f"=IF(F{r}>0,J{r}/F{r}*100,0)")
            # Fil de la ligne : commentaire Excel natif (la bulle qui s'ouvre au
            # survol, marquée d'un coin rouge), posé sur le libellé du poste. Une
            # colonne de texte déformerait la grille de chiffres et casserait
            # l'impression ; l'annotation, elle, reste sur la ligne qu'elle explique.
            fil = commentaires_par_code.get(_budget_code_key(poste.code), [])
            if fil:
                note = Comment("\n\n".join(fil), "")
                # Bulle dimensionnée sur le contenu, sinon un long fil s'ouvre sur
                # un cadre de trois lignes qu'il faut redimensionner à la main.
                lignes_note = sum(max(1, len(t) // 45 + 1) for t in fil) + len(fil)
                note.width = 320
                note.height = min(60 + 14 * lignes_note, 400)
                ws.cell(row=r, column=4).comment = note
            for col in range(1, 15):
                cell = ws.cell(row=r, column=col)
                cell.border = border
                cell.alignment = left if col in (1, 4) else center
                if col >= 6:
                    cell.alignment = right
            ws.cell(row=r, column=3).font = Font(color="FF64748B")
            for col in range(6, 13):
                ws.cell(row=r, column=col).number_format = MONEY
            ws.cell(row=r, column=13).number_format = PCT
            ws.cell(row=r, column=14).number_format = PCT
            # Regroupement pliable : chaque sous-poste est indenté sous son parent.
            if depth > 0:
                ws.row_dimensions[r].outline_level = min(depth, 7)
            if is_parent:
                fill = PatternFill(fill_type="solid", fgColor=LEVEL_FILLS[min(depth, len(LEVEL_FILLS) - 1)])
                for col in range(1, 15):
                    cell = ws.cell(row=r, column=col)
                    cell.fill = fill
                    cell.font = Font(bold=True, color=GREEN_DARK)
            elif idx % 2 == 1:
                for col in range(1, 15):
                    ws.cell(row=r, column=col).fill = muted_fill
            # Ligne hors calcul : grisée et en italique. Elle reste lisible et à sa
            # place dans la hiérarchie, mais l'oeil voit tout de suite qu'elle
            # n'entre dans aucun total — sans quoi le classeur semblerait faux.
            if not est_inclus(poste):
                for col in range(1, 15):
                    cellule_hc = ws.cell(row=r, column=col)
                    cellule_hc.font = Font(italic=True, color="FF94A3B8")
            exec_cell = ws.cell(row=r, column=14)
            if taux_exec >= Decimal(100):
                exec_cell.font = Font(bold=True, color="FFDC2626")
            elif taux_exec >= Decimal(90):
                exec_cell.font = Font(bold=True, color="FFB45309")

        last_data = FIRST + len(ordered) - 1

        # ── Ligne TOTAL : somme des sous-postes uniquement, donc sans double comptage.
        total_row = last_data + 1
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=2, value="Synthèse")
        ws.cell(row=total_row, column=4, value="Total sans double comptage")
        for col_letter, col in (("F", 6), ("G", 7), ("I", 9), ("J", 10)):
            formula = (
                f'=SUMIF($B${FIRST}:$B${last_data},"Sous-poste",{col_letter}${FIRST}:{col_letter}${last_data})'
                if last_data >= FIRST
                else 0
            )
            ws.cell(row=total_row, column=col, value=formula)
        ws.cell(row=total_row, column=8, value=f"=F{total_row}-G{total_row}")
        ws.cell(
            row=total_row,
            column=11,
            value=(f"=J{total_row}-F{total_row}" if is_recette else f"=F{total_row}-J{total_row}"),
        )
        ws.cell(row=total_row, column=12, value=f"=F{total_row}-I{total_row}")
        ws.cell(row=total_row, column=13, value=f"=IF(F{total_row}>0,I{total_row}/F{total_row}*100,0)")
        ws.cell(row=total_row, column=14, value=f"=IF(F{total_row}>0,J{total_row}/F{total_row}*100,0)")
        for col in range(1, 15):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = right if col >= 6 else center
        for col in range(6, 13):
            ws.cell(row=total_row, column=col).number_format = MONEY
        ws.cell(row=total_row, column=13).number_format = PCT
        ws.cell(row=total_row, column=14).number_format = PCT

        # ── Commentaire général, sous le tableau ──────────────────────────────────
        # Le chapeau du document : ce qui justifie l'ensemble du budget, là où les
        # commentaires de ligne justifient un montant. Un export « global » porte
        # les deux vues, chacune étiquetée ; une vue seule n'a pas à se nommer.
        blocs_generaux: list[str] = []
        global_view = filtre_type not in ("DEPENSE", "RECETTE")
        if filtre_type != "RECETTE":
            texte_dep = (exercice.commentaire_general_depense or "").strip()
            if texte_dep:
                blocs_generaux.append(f"Dépenses — {texte_dep}" if global_view else texte_dep)
        if filtre_type != "DEPENSE":
            texte_rec = (exercice.commentaire_general_recette or "").strip()
            if texte_rec:
                blocs_generaux.append(f"Recettes — {texte_rec}" if global_view else texte_rec)

        if blocs_generaux:
            bloc_row = total_row + 2
            ws.merge_cells(start_row=bloc_row, start_column=1, end_row=bloc_row, end_column=14)
            titre = ws.cell(row=bloc_row, column=1, value="COMMENTAIRE GÉNÉRAL")
            titre.font = Font(bold=True, color="FFFFFFFF")
            titre.fill = header_fill
            titre.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for col in range(1, 15):
                ws.cell(row=bloc_row, column=col).border = border
            ws.row_dimensions[bloc_row].height = 18

            # Le texte est découpé ici, une ligne par rangée, plutôt que confié au
            # retour à la ligne automatique d'une cellule fusionnée : Excel n'ajuste
            # jamais la hauteur d'une fusion, il faudrait donc l'estimer — et une
            # estimation trop large dessine un cadre bien plus haut que le texte.
            # Des rangées de hauteur par défaut collent au contenu, toujours.
            # 170 < 231 (largeur cumulée des quatorze colonnes) : marge suffisante pour
            # qu'aucune ligne ne déborde, quelle que soit la police du poste client.
            LARGEUR_BLOC = 170
            for texte in blocs_generaux:
                for paragraphe in texte.split("\n"):
                    for ligne_txt in textwrap.wrap(paragraphe, width=LARGEUR_BLOC) or [""]:
                        bloc_row += 1
                        ws.merge_cells(
                            start_row=bloc_row, start_column=1, end_row=bloc_row, end_column=14
                        )
                        cellule = ws.cell(row=bloc_row, column=1, value=ligne_txt)
                        cellule.alignment = Alignment(
                            horizontal="left", vertical="center", indent=1
                        )
                        cellule.font = Font(color=SLATE)
                        for col in range(1, 15):
                            ws.cell(row=bloc_row, column=col).border = border

        # Barres de données sur le taux d'exécution (jauge visuelle par ligne).
        if last_data >= FIRST:
            ws.conditional_formatting.add(
                f"N{FIRST}:N{last_data}",
                DataBarRule(start_type="num", start_value=0, end_type="num",
                            end_value=100, color="FF10B981"),
            )

        ws.freeze_panes = f"A{FIRST}"
        ws.auto_filter.ref = f"A{HEADER_ROW}:N{last_data}"
        if ws.sheet_properties.outlinePr is not None:
            ws.sheet_properties.outlinePr.summaryBelow = False
        for col_letter, w in (("A", 16), ("B", 13), ("C", 8), ("D", 46), ("E", 12),
                              ("F", 15), ("G", 17), ("H", 17), ("I", 14), ("J", 14),
                              ("K", 15), ("L", 16), ("M", 16), ("N", 16)):
            ws.column_dimensions[col_letter].width = w

        # ── Feuille « Synthèse » (indicateurs + graphiques) ────────────────────────
        leaf_stats = []
        for p, _, _ in leaves:
            pv, en, py = node_totals(p)
            leaf_stats.append((p, pv, en, py, _pct(py, pv)))
        nb_postes = len(leaf_stats)
        nb_entames = sum(1 for _, _, _, py, _ in leaf_stats if py > 0)
        nb_proches = sum(1 for *_, pct in leaf_stats if Decimal(90) <= pct < Decimal(100))
        nb_depass = sum(1 for *_, pct in leaf_stats if pct >= Decimal(100))

        ws2 = wb.create_sheet("Synthèse")
        ws2.sheet_properties.tabColor = "FF0F766E"
        ws2.sheet_view.showGridLines = False
        ws2.append(["Indicateur", "Valeur"])
        for col_idx in (1, 2):
            c = ws2.cell(row=1, column=col_idx)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        synth_rows = [
            # Émetteur en tête : cette feuille est consultable indépendamment de la
            # première, elle doit donc porter elle aussi l'identification du tenant.
            ("Organisation", organisation, None),
            ("Exercice", annee, None),
            ("Type", (filtre_type or "TOUT"), None),
            ("Nombre de sous-postes", nb_postes, None),
            ("Sous-postes entamés", nb_entames, None),
            (f"Proches de la prévision (90-99%)", nb_proches, None),
            ("En dépassement (>=100%)", nb_depass, None),
            ("Total budget en prévision (USD)", float(tot_prevu), MONEY),
            ("Total engagé (USD)", float(tot_engage), MONEY),
            (
                "Total recettes réalisées (USD)" if filtre_type == "RECETTE"
                else "Total dépenses réalisées (USD)",
                float(tot_paye), MONEY,
            ),
            (f"{LBL_SOLDE} (USD)", float(tot_disp), MONEY),
            ("Reste à engager (USD)", float(tot_reste_engager), MONEY),
            ("Taux d'engagement global %", float(_pct(tot_engage, tot_prevu)), PCT),
            (f"{LBL_TAUX} global %", float(_pct(tot_paye, tot_prevu)), PCT),
        ]
        for label, val, fmt in synth_rows:
            ws2.append([label, val])
            row_idx = ws2.max_row
            fill = muted_fill if row_idx % 2 == 0 else PatternFill(fill_type="solid", fgColor="FFFFFFFF")
            if label.startswith("Total") or label.startswith("Solde") or label.startswith("Reste"):
                fill = subheader_fill
            if label.startswith("En dépassement"):
                fill = PatternFill(fill_type="solid", fgColor=RED_SOFT)
            ws2.cell(row=row_idx, column=1).font = Font(bold=True, color=SLATE)
            ws2.cell(row=row_idx, column=1).border = border
            ws2.cell(row=row_idx, column=2).border = border
            ws2.cell(row=row_idx, column=2).alignment = right
            for col_idx in (1, 2):
                ws2.cell(row=row_idx, column=col_idx).fill = fill
            if fmt:
                ws2.cell(row=row_idx, column=2).number_format = fmt

        # ── Graphiques ─────────────────────────────────────────────────────────────
        top = sorted(leaf_stats, key=lambda t: t[3], reverse=True)[:10]
        if top:
            gh = ws2.max_row + 2  # en-tête du bloc de données servant aux graphiques
            ws2.cell(row=gh, column=1, value="Poste")
            ws2.cell(row=gh, column=2, value=f"{LBL_REALISE} (USD)")
            ws2.cell(row=gh, column=3, value=f"{LBL_PREVISION} (USD)")
            for col_idx in range(1, 4):
                cell = ws2.cell(row=gh, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center
            for i, (p, pv, en, py, pct) in enumerate(top, start=1):
                rr = gh + i
                ws2.cell(row=rr, column=1, value=(p.libelle or p.code or "")[:30])
                ws2.cell(row=rr, column=2, value=float(py)).number_format = MONEY
                ws2.cell(row=rr, column=3, value=float(pv)).number_format = MONEY
                for col_idx in range(1, 4):
                    cell = ws2.cell(row=rr, column=col_idx)
                    cell.fill = muted_fill if i % 2 == 0 else PatternFill(fill_type="solid", fgColor="FFFFFFFF")
                    cell.border = border
            bar = BarChart()
            bar.type = "bar"
            bar.title = f"Top postes — {LBL_REALISE} vs {LBL_PREVISION}"
            bar.height = 9
            bar.width = 22
            bar.add_data(
                Reference(ws2, min_col=2, max_col=3, min_row=gh, max_row=gh + len(top)),
                titles_from_data=True,
            )
            bar.set_categories(Reference(ws2, min_col=1, min_row=gh + 1, max_row=gh + len(top)))
            ws2.add_chart(bar, "E2")

            dg = gh + len(top) + 2
            ws2.cell(row=dg, column=1, value=LBL_REALISE)
            ws2.cell(row=dg, column=2, value=float(tot_paye)).number_format = MONEY
            ws2.cell(row=dg + 1, column=1, value=LBL_SOLDE)
            ws2.cell(row=dg + 1, column=2, value=float(tot_disp)).number_format = MONEY
            for row_idx, fill_color in ((dg, AMBER_SOFT), (dg + 1, GREEN_LIGHT if tot_disp >= 0 else RED_SOFT)):
                for col_idx in (1, 2):
                    cell = ws2.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
                    cell.border = border
                    if col_idx == 1:
                        cell.font = Font(bold=True, color=SLATE)
            doughnut = DoughnutChart()
            doughnut.title = f"{LBL_TAUX} globale ({LBL_REALISE} vs {LBL_SOLDE})"
            doughnut.height = 7
            doughnut.width = 9
            doughnut.add_data(Reference(ws2, min_col=2, min_row=dg, max_row=dg + 1))
            doughnut.set_categories(Reference(ws2, min_col=1, min_row=dg, max_row=dg + 1))
            ws2.add_chart(doughnut, "E20")

        # Liste des dépassements (postes au plafond ou au-delà).
        depassements = sorted(
            [(p, pv, py, pct) for (p, pv, en, py, pct) in leaf_stats if pct >= Decimal(100)],
            key=lambda t: t[3],
            reverse=True,
        )
        if depassements:
            ws2.append([])
            ws2.append(["Postes en dépassement", ""])
            dep_title_row = ws2.max_row
            for col_idx in range(1, 6):
                cell = ws2.cell(row=dep_title_row, column=col_idx)
                cell.fill = PatternFill(fill_type="solid", fgColor=RED_SOFT)
                cell.border = border
            ws2.cell(row=dep_title_row, column=1).font = Font(bold=True, color="FFDC2626")
            ws2.append(["Code", "Poste", LBL_PREVISION, LBL_REALISE, f"{LBL_TAUX} %"])
            for col_idx in range(1, 6):
                ws2.cell(row=ws2.max_row, column=col_idx).font = Font(bold=True, color="FFFFFFFF")
                ws2.cell(row=ws2.max_row, column=col_idx).fill = PatternFill(fill_type="solid", fgColor="FFDC2626")
                ws2.cell(row=ws2.max_row, column=col_idx).border = border
            for p, pv, py, pct in depassements:
                ws2.append([p.code or "", p.libelle or "", float(pv), float(py), float(pct)])
                row_idx = ws2.max_row
                for col_idx in range(1, 6):
                    ws2.cell(row=row_idx, column=col_idx).fill = PatternFill(fill_type="solid", fgColor=RED_SOFT)
                    ws2.cell(row=row_idx, column=col_idx).border = border
                ws2.cell(row=row_idx, column=3).number_format = MONEY
                ws2.cell(row=row_idx, column=4).number_format = MONEY
                ws2.cell(row=row_idx, column=5).number_format = PCT
        _autosize_columns(ws2)

        suffix = filtre_type or "TOUT"
        if service_id is not None:
            suffix = f"{suffix}_service{service_id}"
        filename = f"budget_{annee}_{suffix}.xlsx"
        return wb, filename

    # Connexion rendue au pool avant le travail CPU, comme sur /encaissements et
    # /requisitions. Verifie et non suppose : la closure ci-dessus ne lit que des
    # colonnes deja chargees — `code`, `libelle`, `id`, `type` sur BudgetPoste,
    # `commentaire_general_recette` / `_depense` sur BudgetExercice — et aucune
    # relation (`exercice`, `parent`, `children` ne sont jamais touchees). Avec
    # `expire_on_commit=False` (db/session.py:99) ces valeurs restent lisibles
    # apres la fin de la transaction.
    #
    # `commit()` et non `rollback()` : rollback expire systematiquement les
    # objets de la session, et la premiere lecture d'attribut dans le thread
    # declencherait un rechargement hors contexte greenlet (MissingGreenlet).
    # Cet export ne modifie plus rien depuis que la surcharge d'affichage des
    # recettes passe par un dictionnaire local : il n'y a rien a ecrire.
    await _relacher_connexion(db)
    return await anyio.to_thread.run_sync(_build_workbook)


@router.get("/budget")
async def export_budget(
    annee: int | None = Query(default=None),
    type: str | None = Query(default=None),
    service_id: int | None = Query(default=None),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> Response:
    """Rend le classeur directement, ou 202 et un job au-dessus du seuil.

    L'aiguillage est cote SERVEUR : c'est lui qui decide du regime, le client
    s'adapte a ce qu'il recoit (frontend/src/utils/download.ts traite 200 et
    202). C'est ce qui rend la bascule reversible type par type sans
    redeploiement du frontend — EXPORT_ASYNC_TYPES suffit, dans les deux sens.
    """
    params = {"annee": annee, "type": type, "service_id": service_id}
    try:
        wb, filename = await construire_classeur_budget(
            db,
            user.organisation_id,
            annee=annee,
            type=type,
            service_id=service_id,
            seuil_bascule=_seuil_bascule("budget"),
        )
    except BasculeAsynchroneRequise as bascule:
        return await _soumettre_export_asynchrone(
            db,
            user,
            type_export="budget",
            params=params,
            row_count=bascule.total,
        )
    return await _excel_response(filename, wb)


async def construire_classeur_encaissements(
    db: AsyncSession,
    organisation_id: int,
    *,
    date_debut: str | None = None,
    date_fin: str | None = None,
    statut_paiement: str | None = None,
    numero_recu: str | None = None,
    client: str | None = None,
    budget_poste_id: int | None = None,
    type_client: str | None = None,
    mode_paiement: str | None = None,
    expert_comptable_id: str | None = None,
    deleted_status: str | None = "all",
    est_proforma: bool | None = False,
    seuil_bascule: int | None = None,
) -> tuple[Workbook, str]:
    """Construit le classeur `encaissements` et rend `(classeur, nom de fichier)`.

    EXTRAIT DE L'ENDPOINT, sans autre changement que `user.organisation_id`
    remplace par `organisation_id`. Le worker et la route HTTP appellent
    cette meme fonction : c'est ce qui rend leur divergence impossible
    plutot qu'improbable.
    """
    # `created_by` porte l'utilisateur qui a enregistré l'encaissement. Ce n'est
    # pas une clé étrangère déclarée : la jointure est explicite.
    deleted_filter = (deleted_status or "all").strip().lower()
    if deleted_filter not in {"all", "active", "deleted"}:
        raise HTTPException(status_code=400, detail="deleted_status invalide (all, active, deleted)")
    Encaisseur = aliased(User)
    query = (
        select(Encaissement, ExpertComptable, Encaisseur)
        .options(joinedload(Encaissement.compte_bancaire).joinedload(CompteBancaire.banque))
        .outerjoin(ExpertComptable, Encaissement.expert_comptable_id == ExpertComptable.id)
        .outerjoin(Encaisseur, Encaissement.created_by == Encaisseur.id)
        .where(Encaissement.organisation_id == organisation_id)
    )

    start_dt = _parse_datetime(date_debut)
    end_dt = _parse_datetime(date_fin, end_of_day=True)
    if start_dt:
        query = query.where(Encaissement.date_encaissement >= start_dt)
    if end_dt:
        query = query.where(Encaissement.date_encaissement <= end_dt)

    if statut_paiement:
        query = query.where(Encaissement.statut_paiement == statut_paiement)
    if numero_recu:
        query = query.where(Encaissement.numero_recu.ilike(f"%{numero_recu}%"))
    if budget_poste_id:
        query = query.where(Encaissement.budget_poste_id == budget_poste_id)
    if type_client:
        query = query.where(Encaissement.type_client == type_client)
    if mode_paiement:
        query = query.where(Encaissement.mode_paiement == mode_paiement)
    if deleted_filter == "active":
        query = query.where(Encaissement.is_deleted.is_(False))
    elif deleted_filter == "deleted":
        query = query.where(Encaissement.is_deleted.is_(True))
    if est_proforma is not None:
        query = query.where(Encaissement.est_proforma.is_(est_proforma))
    if est_proforma is False:
        query = query.where((Encaissement.statut_operation.is_(None)) | (Encaissement.statut_operation == "ACTIVE"))
    if expert_comptable_id:
        try:
            exp_uid = uuid.UUID(expert_comptable_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid expert_comptable_id UUID")
        query = query.where(Encaissement.expert_comptable_id == exp_uid)

    if client:
        query = query.where(
            or_(
                Encaissement.client_nom.ilike(f"%{client}%"),
                ExpertComptable.nom_denomination.ilike(f"%{client}%"),
                ExpertComptable.numero_ordre.ilike(f"%{client}%"),
            )
        )

    query = query.order_by(Encaissement.date_encaissement.desc())

    await _compter_lignes(db, query, export="encaissements", seuil_bascule=seuil_bascule)
    rows = (await db.execute(query)).all()
    # Identification du tenant émetteur : obligatoire sur tout document exporté.
    organisation = await _tenant_display_name(db, organisation_id)

    # Entrées internes (approvisionnements banque -> caisse et versements
    # caisse -> banque) : préchargées ici
    # (avant la construction, purement synchrone, du classeur) pour ne pas
    # mêler d'await à ce bloc CPU.
    filtres_clients = any(
        [statut_paiement, numero_recu, client, budget_poste_id, type_client, expert_comptable_id, mode_paiement]
    )
    approvisionnements: list = []
    versements_banque: list = []
    if not filtres_clients and est_proforma is False:
        approvisionnements = await list_entrees_internes_caisse(
            db,
            tenant_id=organisation_id,
            date_debut=start_dt,
            date_fin=end_dt,
        )
        versements_banque = await list_entrees_internes_banque(
            db,
            tenant_id=organisation_id,
            date_debut=start_dt,
            date_fin=end_dt,
        )

    def _build_workbook() -> tuple[Workbook, str]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Encaissements"

        headers = [
            "Type d'opération",
            "Source / Mode",
            "Banque source",
            "Compte bancaire",
            "Date",
            "Heure",
            "N° Note de débit",
            "Type de client",
            "Client",
            "Libellé",
            "Poste budgétaire",
            "Description",
            "Devise perçue",
            "Montant perçu",
            "Montant total (USD)",
            "Montant payé (USD)",
            "Reste à payer (USD)",
            "Mode de paiement",
            "Référence",
            "Statut paiement",
            "Encaissé par",
            "Statut",
        ]

        # (clé de tri, ligne, entrée interne ?) : notes de débit et entrées de caisse
        # sont mêlées puis retriées par date. Le drapeau suit la ligne à travers le
        # tri, seul moyen de retrouver les entrées internes une fois l'ordre changé.
        entries: list[tuple[Any, list[Any], bool, bool]] = []
        total_notes_debit = Decimal("0")
        total_paye = Decimal("0")
        totals_by_mode: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
        totals_by_type_client: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))

        for enc, expert, encaisseur in rows:
            client_label = (
                f"{expert.numero_ordre} - {expert.nom_denomination}"
                if expert is not None
                else (enc.client_nom or "")
            )
            montant_total = _round_money(enc.montant_total or enc.montant or Decimal("0"))
            montant_paye = _round_money(enc.montant_percu or enc.montant_paye or Decimal("0"))
            reste = _round_money(montant_total - montant_paye)
            if abs(reste) < Decimal("0.05"):
                reste = Decimal("0.00")
                montant_paye = montant_total
            is_deleted = bool(enc.is_deleted)
            if not is_deleted:
                total_notes_debit += Decimal(montant_total or 0)
                total_paye += Decimal(montant_paye or 0)

            mode_label = _format_mode_paiement(enc.mode_paiement)
            if not is_deleted:
                totals_by_mode[mode_label or "Non précisé"] += Decimal(montant_paye or 0)
                totals_by_type_client[enc.type_client or "Non précisé"] += Decimal(montant_total or 0)

            poste_label = (
                f"{enc.budget_poste_code} - {enc.budget_poste_libelle}"
                if enc.budget_poste_code and enc.budget_poste_libelle
                else (enc.budget_poste_code or enc.budget_poste_libelle or "")
            )
            entries.append((
                enc.date_encaissement or enc.created_at,
                _financial_source_columns("Encaissement", enc.canal, enc.compte_bancaire)
                + [
                    enc.date_encaissement.strftime("%d/%m/%Y") if enc.date_encaissement else "",
                    _format_operation_time(enc.date_encaissement, enc.created_at),
                    enc.numero_recu,
                    enc.type_client,
                    client_label,
                    enc.libelle or "",
                    poste_label,
                    enc.description or "",
                    enc.devise_perception or "USD",
                    float(enc.montant_percu or 0),
                    float(montant_total or 0),
                    float(montant_paye or 0),
                    float(reste or 0),
                    mode_label,
                    enc.reference or "",
                    enc.statut_paiement,
                    _person_name(encaisseur),
                    "SUPPRIMÉ" if is_deleted else "Actif",
                ],
                False,
                is_deleted,
            ))

        # --- Entrées de caisse hors notes de débit : les approvisionnements
        # banque -> caisse. Ce sont des sorties du compte bancaire, mais de l'argent
        # qui ENTRE en caisse — sans eux, l'export laisse croire que seules les notes
        # de débit alimentent le tiroir. Ils ne sont PAS des recettes clients : ils
        # remplissent uniquement les colonnes qu'ils renseignent (montant perçu,
        # devise, source bancaire) et restent hors des totaux USD.
        # Un filtre propre aux notes de débit (client, poste, statut de paiement…)
        # n'a pas de sens pour eux : on les omet alors, plutôt que de livrer une
        # liste qui contredit les filtres affichés.
        # (préchargées en amont dans approvisionnements, cf. plus haut)
        for ligne in approvisionnements:
            entries.append((
                ligne["date"] or ligne["created_at"],
                [
                    "Approvisionnement caisse",
                    # Origine portée par la ligne : un transfert du moteur dédié
                    # peut venir d'un compte précis, et un virement de banque à
                    # banque n'a pas la caisse pour origine.
                    ligne["provenance"],
                    ligne["banque"] or "—",
                    ligne["compte_numero"] or "—",
                    ligne["date"].strftime("%d/%m/%Y") if ligne["date"] else "",
                    _format_operation_time(ligne["date"], ligne["created_at"]),
                    "—",  # pas de note de débit : ce n'est pas une recette client
                    "—",
                    "—",
                    ligne["libelle"],
                    "—",
                    "Transfert interne banque → caisse (entrée en caisse)",
                    ligne["devise"],
                    float(ligne["montant"] or 0),
                    # Colonnes de la note de débit (total / payé / reste) : sans
                    # objet ici. Les laisser vides plutôt qu'à 0 évite qu'elles
                    # soient lues comme un montant réel ou intégrées aux totaux.
                    "",
                    "",
                    "",
                    "Transfert interne",
                    ligne["reference"] or "",
                    "—",
                    ligne["auteur"],
                    "Actif",
                ],
                True,
                False,
            ))

        # --- Entrées bancaires hors notes de débit : les versements caisse ->
        # banque. Ce sont des sorties de caisse, mais de l'argent qui ENTRE sur
        # un compte bancaire. Même traitement que les approvisionnements :
        # visible pour rapprochement, hors totaux économiques d'encaissements.
        for ligne in versements_banque:
            entries.append((
                ligne["date"] or ligne["created_at"],
                [
                    "Versement banque",
                    ligne["provenance"],
                    ligne["banque"] or "—",
                    ligne["compte_numero"] or "—",
                    ligne["date"].strftime("%d/%m/%Y") if ligne["date"] else "",
                    _format_operation_time(ligne["date"], ligne["created_at"]),
                    "—",
                    "—",
                    "—",
                    ligne["libelle"],
                    "—",
                    "Transfert interne caisse → banque (entrée bancaire)",
                    ligne["devise"],
                    float(ligne["montant"] or 0),
                    "",
                    "",
                    "",
                    "Transfert interne",
                    ligne["reference"] or "",
                    "—",
                    ligne["auteur"],
                    "Actif",
                ],
                True,
                False,
            ))

        # Tri décroissant par date : notes de débit et entrées de caisse entremêlées.
        entries.sort(key=lambda e: _sort_key_datetime(e[0]), reverse=True)
        data_rows = [row for _, row, _, _ in entries]
        entrees_internes_rows = {idx for idx, (_, _, interne, _) in enumerate(entries) if interne}
        deleted_rows = {idx for idx, (_, _, _, deleted) in enumerate(entries) if deleted}
        row_fills = {idx: transfert_fill for idx in entrees_internes_rows}
        row_fills.update({idx: PatternFill(fill_type="solid", fgColor=RED_SOFT) for idx in deleted_rows})

        periode = f"{date_debut or 'début'} → {date_fin or 'fin'}"
        legende_entrees = (
            "  |  Lignes turquoise = transferts internes caisse ↔ banque "
            "(hors totaux notes de débit)"
            if entrees_internes_rows
            else ""
        )
        legende_supprimes = (
            "  |  Lignes rouges = encaissements supprimés (hors totaux)"
            if deleted_rows
            else ""
        )
        _build_list_sheet(
            ws,
            title="ENCAISSEMENTS",
            subtitle=f"Période : {periode}  |  Montants en USD{legende_entrees}{legende_supprimes}",
            headers=headers,
            data_rows=data_rows,
            money_cols=(14, 15, 16, 17),
            total_values={
                15: float(total_notes_debit),
                16: float(total_paye),
                17: float(total_notes_debit - total_paye),
            },
            organisation=organisation,
            highlight_rows=entrees_internes_rows,
            highlight_fill=transfert_fill,
            highlight_row_fills=row_fills,
        )

        mode_rows = [
            [mode, float(amount)]
            for mode, amount in sorted(totals_by_mode.items(), key=lambda kv: kv[1], reverse=True)
        ]
        type_rows = [
            [tc, float(amount)]
            for tc, amount in sorted(totals_by_type_client.items(), key=lambda kv: kv[1], reverse=True)
        ]
        _build_synthese_sheet(
            wb,
            "Synthèse — Encaissements",
            [
                {
                    "title": "Répartition par mode de paiement",
                    "headers": ["Mode de paiement", "Montant payé (USD)"],
                    "rows": mode_rows,
                    "money_cols": (2,),
                },
                {
                    "title": "Répartition par type de client",
                    "headers": ["Type de client", "Montant total (USD)"],
                    "rows": type_rows,
                    "money_cols": (2,),
                },
            ],
            chart_title="Encaissements par mode de paiement",
            chart_block_index=0,
            chart_value_col=2,
            organisation=organisation,
        )

        suffix = f"{date_debut or 'debut'}_{date_fin or 'fin'}"
        filename = f"encaissements_{suffix}.xlsx"
        return wb, filename

    await _relacher_connexion(db)
    return await anyio.to_thread.run_sync(_build_workbook)


@router.get("/encaissements", dependencies=[Depends(has_permission("menu_encaissements"))])
async def export_encaissements(
    date_debut: str | None = Query(default=None),
    date_fin: str | None = Query(default=None),
    statut_paiement: str | None = Query(default=None),
    numero_recu: str | None = Query(default=None),
    client: str | None = Query(default=None),
    budget_poste_id: int | None = Query(default=None),
    type_client: str | None = Query(default=None),
    mode_paiement: str | None = Query(default=None),
    expert_comptable_id: str | None = Query(default=None),
    deleted_status: str | None = Query(default="all"),
    est_proforma: bool | None = Query(default=False),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> Response:
    """Rend le classeur directement, ou 202 et un job au-dessus du seuil.

    L'aiguillage est cote SERVEUR : c'est lui qui decide du regime, le client
    s'adapte a ce qu'il recoit (frontend/src/utils/download.ts traite 200 et
    202). C'est ce qui rend la bascule reversible type par type sans
    redeploiement du frontend — EXPORT_ASYNC_TYPES suffit, dans les deux sens.
    """
    params = {
            "date_debut": date_debut,
            "date_fin": date_fin,
            "statut_paiement": statut_paiement,
            "numero_recu": numero_recu,
            "client": client,
            "budget_poste_id": budget_poste_id,
            "type_client": type_client,
            "mode_paiement": mode_paiement,
            "expert_comptable_id": expert_comptable_id,
            "deleted_status": deleted_status,
            "est_proforma": est_proforma,
    }
    try:
        wb, filename = await construire_classeur_encaissements(
            db,
            user.organisation_id,
            date_debut=date_debut,
            date_fin=date_fin,
            statut_paiement=statut_paiement,
            numero_recu=numero_recu,
            client=client,
            budget_poste_id=budget_poste_id,
            type_client=type_client,
            mode_paiement=mode_paiement,
            expert_comptable_id=expert_comptable_id,
            deleted_status=deleted_status,
            est_proforma=est_proforma,
            seuil_bascule=_seuil_bascule("encaissements"),
        )
    except BasculeAsynchroneRequise as bascule:
        return await _soumettre_export_asynchrone(
            db,
            user,
            type_export="encaissements",
            params=params,
            row_count=bascule.total,
        )
    return await _excel_response(filename, wb)


async def construire_classeur_sorties_fonds(
    db: AsyncSession,
    organisation_id: int,
    *,
    date_debut: str | None = None,
    date_fin: str | None = None,
    type_sortie: str | None = None,
    mode_paiement: str | None = None,
    statut: str | None = None,
    requisition_numero: str | None = None,
    reference: str | None = None,
    seuil_bascule: int | None = None,
) -> tuple[Workbook, str]:
    """Construit le classeur `sorties-fonds` et rend `(classeur, nom de fichier)`.

    EXTRAIT DE L'ENDPOINT, sans autre changement que `user.organisation_id`
    remplace par `organisation_id`. Le worker et la route HTTP appellent
    cette meme fonction : c'est ce qui rend leur divergence impossible
    plutot qu'improbable.
    """
    Auteur = aliased(User)
    Programmeur = aliased(User)
    query = (
        select(SortieFonds, Requisition, Auteur, Programmeur)
        .options(joinedload(SortieFonds.compte_bancaire).joinedload(CompteBancaire.banque))
        .outerjoin(Requisition, SortieFonds.requisition_id == Requisition.id)
        .outerjoin(Auteur, SortieFonds.created_by == Auteur.id)
        .outerjoin(Programmeur, SortieFonds.programme_par_id == Programmeur.id)
        .where(SortieFonds.organisation_id == organisation_id)
    )

    query = query.where(
        or_(
            SortieFonds.requisition_id.is_(None),
            Requisition.status.in_(REQUISITION_STATUTS_VALIDES),
        )
    )

    start_dt = _parse_datetime(date_debut)
    end_dt = _parse_datetime(date_fin, end_of_day=True)
    if start_dt:
        query = query.where(SortieFonds.date_paiement >= start_dt)
    if end_dt:
        query = query.where(SortieFonds.date_paiement <= end_dt)

    if type_sortie:
        query = query.where(SortieFonds.type_sortie == type_sortie)
    if mode_paiement:
        query = query.where(SortieFonds.mode_paiement == mode_paiement)
    if statut:
        statut_value = statut.strip().upper()
        if statut_value == "ALL":
            query = query
        elif statut_value == "VALIDE":
            query = query.where(
                (SortieFonds.statut.is_(None)) | (SortieFonds.statut == "VALIDE")
            )
        else:
            query = query.where(SortieFonds.statut == statut_value)
    else:
        query = query.where(
            (SortieFonds.statut.is_(None)) | (SortieFonds.statut == "VALIDE")
        )
    if reference:
        query = query.where(SortieFonds.reference.ilike(f"%{reference}%"))
    if requisition_numero:
        query = query.where(Requisition.numero_requisition.ilike(f"%{requisition_numero}%"))

    query = query.order_by(SortieFonds.created_at.desc())

    await _compter_lignes(db, query, export="sorties-fonds", seuil_bascule=seuil_bascule)
    rows = (await db.execute(query)).all()
    # Identification du tenant émetteur : obligatoire sur tout document exporté.
    organisation = await _tenant_display_name(db, organisation_id)

    req_ids = [req.id for _, req, _, _ in rows if req is not None]
    rubriques_map: dict[str, str] = {}
    if req_ids:
        lignes = []
        for lot in _par_lots(req_ids):
            lignes.extend(
                (
                    await db.execute(
                        select(LigneRequisition).where(LigneRequisition.requisition_id.in_(lot))
                    )
                ).scalars().all()
            )
        grouped: dict[str, set[str]] = {}
        for ligne in lignes:
            key = str(ligne.requisition_id)
            grouped.setdefault(key, set()).add(ligne.rubrique)
        rubriques_map = {k: ", ".join(sorted(v)) for k, v in grouped.items()}

    # Retours en caisse de la période : préchargés ici (avant la construction,
    # purement synchrone, du classeur) pour ne pas mêler d'await à ce bloc CPU.
    include_retours = (
        (not statut or statut.strip().upper() in ("VALIDE", "ALL"))
        and not type_sortie
        and not mode_paiement
    )
    # Les transferts délégués au moteur dédié : même écran, mêmes filtres, donc
    # même classeur. Sans eux, l'export d'une période contiendrait moins de
    # lignes que la liste qu'il exporte — sur un document imprimé et signé.
    transferts_delegues_rows = await transferts_delegues.lignes_export(
        db,
        tenant_id=organisation_id,
        filtres=transferts_delegues.FiltresSorties(
            date_debut=start_dt,
            date_fin=end_dt,
            type_sortie=type_sortie,
            mode_paiement=mode_paiement,
            statut=statut,
            reference=reference,
            filtre_requisition=bool(requisition_numero),
        ),
    )

    retours_rows: list = []
    if include_retours:
        r_query = (
            select(RetourCaisse, SortieFonds, Requisition)
            .options(joinedload(RetourCaisse.compte_bancaire).joinedload(CompteBancaire.banque))
            .join(SortieFonds, RetourCaisse.sortie_fonds_id == SortieFonds.id)
            .outerjoin(Requisition, RetourCaisse.requisition_id == Requisition.id)
            .where(
                RetourCaisse.organisation_id == organisation_id,
                RetourCaisse.statut == "VALIDE",
            )
        )
        if start_dt:
            r_query = r_query.where(RetourCaisse.date_retour >= start_dt)
        if end_dt:
            r_query = r_query.where(RetourCaisse.date_retour <= end_dt)
        if requisition_numero:
            r_query = r_query.where(Requisition.numero_requisition.ilike(f"%{requisition_numero}%"))
        retours_rows = (await db.execute(r_query)).all()

    def _build_workbook() -> tuple[Workbook, str]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sorties"

        headers = [
            "Type d'opération",
            "Source / Mode",
            "Banque source",
            "Compte bancaire",
            "Créée le",
            "Date",
            "Heure",
            "Auteur de l'opération",
            "Programmé par",
            "N° Réquisition",
            "Objet",
            "Poste budgétaire",
            "Bénéficiaire",
            "Motif",
            "Montant payé (USD)",
            "Mode de paiement",
            "Référence",
            "Statut",
            "Commentaire",
        ]

        # (clé de tri, ligne, transfert interne ?) : sorties et retours sont mêlés
        # puis triés par date. Le drapeau suit la ligne à travers le tri, seul moyen
        # de retrouver les transferts une fois l'ordre changé.
        entries: list[tuple[Any, list[Any], bool]] = []
        total_paye = Decimal("0")
        totals_by_type: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
        totals_by_mode: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))

        for sortie, req, creator, programmeur in rows:
            montant = Decimal(sortie.montant_paye or 0)
            total_paye += montant
            rubrique_value = rubriques_map.get(str(req.id), "") if req else ""

            author_name = _person_name(creator)
            programmeur_name = _person_name(programmeur)

            # Un versement ou un approvisionnement n'est pas une dépense : l'argent
            # reste dans l'organisation. Le nommer « Transfert interne » dans le mode
            # de paiement le sort du lot — et le sort aussi du cumul « Cash », qui
            # comptait jusqu'ici ces mouvements comme des décaissements.
            est_transfert = (sortie.type_sortie or "").lower() in TRANSFERT_TYPES
            mode_label = (
                "Transfert interne" if est_transfert else _format_mode_paiement(sortie.mode_paiement)
            )
            totals_by_type[sortie.type_sortie or "Non précisé"] += montant
            totals_by_mode[mode_label or "Non précisé"] += montant

            # La colonne « Type d'opération » garde son rôle : nature de
            # l'enregistrement (sortie vs retour). C'est le mode de paiement qui
            # porte « Transfert interne », et la couleur ne fait que doubler ce
            # libellé écrit — jamais le remplacer (impression N&B, daltonisme).
            entries.append((
                sortie.created_at,
                _financial_source_columns("Sortie des fonds", sortie.canal, sortie.compte_bancaire)
                + [
                    sortie.created_at.strftime("%d/%m/%Y") if sortie.created_at else "",
                    sortie.date_paiement.strftime("%d/%m/%Y") if sortie.date_paiement else "",
                    _format_operation_time(sortie.date_paiement, sortie.created_at),
                    author_name,
                    programmeur_name,
                    req.numero_requisition if req else "",
                    req.objet if req else "",
                    rubrique_value,
                    sortie.beneficiaire or "",
                    sortie.motif or "",
                    float(sortie.montant_paye or 0),
                    mode_label,
                    sortie.reference or "",
                    (sortie.statut or "VALIDE"),
                    sortie.commentaire or "",
                ],
                est_transfert,
            ))

        for ligne in transferts_delegues_rows:
            montant = ligne["montant"]
            total_paye += montant
            # Même traitement que pour le chemin historique : « Transfert
            # interne » dans le mode de paiement, et hors du cumul « Cash ».
            mode_label = "Transfert interne"
            totals_by_type[ligne["type_sortie"]] += montant
            totals_by_mode[mode_label] += montant
            entries.append((
                ligne["created_at"],
                _source_columns(
                    "Sortie des fonds", ligne["canal"], ligne["banque_nom"], ligne["compte_numero"]
                )
                + [
                    ligne["created_at"].strftime("%d/%m/%Y") if ligne["created_at"] else "",
                    ligne["date"].strftime("%d/%m/%Y") if ligne["date"] else "",
                    _format_operation_time(ligne["date"], ligne["created_at"]),
                    _person_name(ligne["auteur"]),
                    "",
                    "",
                    "",
                    "",
                    ligne["beneficiaire"],
                    ligne["motif"],
                    float(montant),
                    mode_label,
                    ligne["reference"],
                    ligne["statut"],
                    "",
                ],
                True,
            ))

        # --- Retours en caisse de la période : lignes à montant NÉGATIF, intégrées
        # et triées avec les sorties. Le total de la colonne devient donc net
        # (sorties brutes − retours). L'export budget, lui, reflète déjà les retours
        # via budget_poste.montant_paye, il n'a rien à changer.
        # (préchargés en amont dans retours_rows, cf. plus haut)
        for retour, sortie_orig, req_r in retours_rows:
            montant_neg = -Decimal(retour.montant or 0)
            total_paye += montant_neg
            mode_label = _format_mode_paiement(retour.mode)
            totals_by_type["Retour en caisse"] += montant_neg
            totals_by_mode[mode_label or "Non précisé"] += montant_neg
            objet_retour = "↩ RETOUR EN CAISSE"
            if req_r and req_r.objet:
                objet_retour = f"↩ RETOUR — {req_r.objet}"
            entries.append((
                retour.created_at,
                _financial_source_columns("Sortie des fonds", retour.canal, retour.compte_bancaire)
                + [
                    retour.created_at.strftime("%d/%m/%Y") if retour.created_at else "",
                    retour.date_retour.strftime("%d/%m/%Y") if retour.date_retour else "",
                    _format_operation_time(retour.date_retour, retour.created_at),
                    "",
                    "",
                    req_r.numero_requisition if req_r else "",
                    objet_retour,
                    retour.budget_poste_libelle or "",
                    sortie_orig.beneficiaire or "",
                    retour.motif or f"Reliquat rendu ({retour.type_retour})",
                    float(montant_neg),
                    mode_label,
                    retour.reference_numero or "",
                    retour.statut or "VALIDE",
                    retour.commentaire or "",
                ],
                False,
            ))

        # Tri décroissant par date de création : sorties et retours entremêlés. Les
        # deux tables peuvent rendre des dates naïves ou aware selon le moteur, d'où
        # la clé normalisée (une comparaison mixte ferait échouer l'export entier).
        entries.sort(key=lambda e: _sort_key_datetime(e[0]), reverse=True)
        data_rows = [row for _, row, _ in entries]
        transfert_rows = {idx for idx, (_, _, est_transfert) in enumerate(entries) if est_transfert}

        periode = f"{date_debut or 'début'} → {date_fin or 'fin'}"
        legende_transferts = (
            "  |  Lignes turquoise = transferts internes caisse ↔ banque (pas des dépenses)"
            if transfert_rows
            else ""
        )
        _build_list_sheet(
            ws,
            title="SORTIES DE FONDS (retours en négatif ; total net)",
            subtitle=(
                f"Période : {periode}  |  Montants en USD  |  Total = sorties − retours"
                f"{legende_transferts}"
            ),
            headers=headers,
            data_rows=data_rows,
            money_cols=(15,),
            total_values={15: float(total_paye)},
            organisation=organisation,
            highlight_rows=transfert_rows,
            highlight_fill=transfert_fill,
        )

        type_rows = [
            [t, float(amount)]
            for t, amount in sorted(totals_by_type.items(), key=lambda kv: kv[1], reverse=True)
        ]
        mode_rows = [
            [mode, float(amount)]
            for mode, amount in sorted(totals_by_mode.items(), key=lambda kv: kv[1], reverse=True)
        ]
        _build_synthese_sheet(
            wb,
            "Synthèse — Sorties de fonds",
            [
                {
                    "title": "Répartition par type de sortie",
                    "headers": ["Type de sortie", "Montant payé (USD)"],
                    "rows": type_rows,
                    "money_cols": (2,),
                },
                {
                    "title": "Répartition par mode de paiement",
                    "headers": ["Mode de paiement", "Montant payé (USD)"],
                    "rows": mode_rows,
                    "money_cols": (2,),
                },
            ],
            chart_title="Sorties par type",
            chart_block_index=0,
            chart_value_col=2,
            organisation=organisation,
        )

        suffix = f"{date_debut or 'debut'}_{date_fin or 'fin'}"
        filename = f"sorties_fonds_{suffix}.xlsx"
        return wb, filename

    return await anyio.to_thread.run_sync(_build_workbook)


@router.get("/sorties-fonds", dependencies=[Depends(has_permission("sorties_fonds"))])
async def export_sorties_fonds(
    date_debut: str | None = Query(default=None),
    date_fin: str | None = Query(default=None),
    type_sortie: str | None = Query(default=None),
    mode_paiement: str | None = Query(default=None),
    statut: str | None = Query(default=None),
    requisition_numero: str | None = Query(default=None),
    reference: str | None = Query(default=None),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> Response:
    """Rend le classeur directement, ou 202 et un job au-dessus du seuil.

    L'aiguillage est cote SERVEUR : c'est lui qui decide du regime, le client
    s'adapte a ce qu'il recoit (frontend/src/utils/download.ts traite 200 et
    202). C'est ce qui rend la bascule reversible type par type sans
    redeploiement du frontend — EXPORT_ASYNC_TYPES suffit, dans les deux sens.
    """
    params = {
            "date_debut": date_debut,
            "date_fin": date_fin,
            "type_sortie": type_sortie,
            "mode_paiement": mode_paiement,
            "statut": statut,
            "requisition_numero": requisition_numero,
            "reference": reference,
    }
    try:
        wb, filename = await construire_classeur_sorties_fonds(
            db,
            user.organisation_id,
            date_debut=date_debut,
            date_fin=date_fin,
            type_sortie=type_sortie,
            mode_paiement=mode_paiement,
            statut=statut,
            requisition_numero=requisition_numero,
            reference=reference,
            seuil_bascule=_seuil_bascule("sorties-fonds"),
        )
    except BasculeAsynchroneRequise as bascule:
        return await _soumettre_export_asynchrone(
            db,
            user,
            type_export="sorties-fonds",
            params=params,
            row_count=bascule.total,
        )
    return await _excel_response(filename, wb)


async def construire_classeur_requisitions(
    db: AsyncSession,
    organisation_id: int,
    *,
    date_debut: str | None = None,
    date_fin: str | None = None,
    statut: str | None = None,
    service_id: int | None = None,
    type_requisition: str | None = None,
    mode_paiement: str | None = None,
    budget_poste_id: int | None = None,
    search: str | None = None,
    objet: str | None = None,
    seuil_bascule: int | None = None,
) -> tuple[Workbook, str]:
    """Construit le classeur `requisitions` et rend `(classeur, nom de fichier)`.

    EXTRAIT DE L'ENDPOINT, sans autre changement que `user.organisation_id`
    remplace par `organisation_id`. Le worker et la route HTTP appellent
    cette meme fonction : c'est ce qui rend leur divergence impossible
    plutot qu'improbable.
    """
    query = select(Requisition, Service).outerjoin(
        Service, Requisition.service_id == Service.id
    ).where(
        Requisition.organisation_id == organisation_id,
        Requisition.is_deleted.is_(False),
    )

    start_dt = _parse_datetime(date_debut)
    end_dt = _parse_datetime(date_fin, end_of_day=True)
    # Même date que la liste et les indicateurs de l'écran Réquisitions.
    req_date = Requisition.created_at
    if start_dt:
        query = query.where(req_date >= start_dt)
    if end_dt:
        query = query.where(req_date <= end_dt)

    if statut:
        statut_value = statut.strip().upper()
        if statut_value != "ALL":
            query = query.where(Requisition.status.in_(_requisition_status_values_for_filter(statut_value)))
    if service_id is not None:
        query = query.where(Requisition.service_id == service_id)
    if type_requisition:
        query = query.where(Requisition.type_requisition == type_requisition)
    if mode_paiement:
        query = query.where(Requisition.mode_paiement == mode_paiement)
    if budget_poste_id is not None:
        query = query.where(
            Requisition.id.in_(
                select(LigneRequisition.requisition_id).where(
                    LigneRequisition.budget_poste_id == budget_poste_id,
                    LigneRequisition.organisation_id == organisation_id,
                )
            )
        )
    if search:
        search_pattern = f"%{search.strip()}%"
        query = query.where(
            or_(
                Requisition.numero_requisition.ilike(search_pattern),
                Requisition.objet.ilike(search_pattern),
                Requisition.created_by.in_(
                    select(User.id).where(
                        or_(User.prenom.ilike(search_pattern), User.nom.ilike(search_pattern)),
                        User.organisation_id == organisation_id,
                    )
                ),
            )
        )
    if objet:
        query = query.where(Requisition.objet.ilike(f"%{objet.strip()}%"))

    query = query.order_by(Requisition.created_at.desc())

    await _compter_lignes(db, query, export="requisitions", seuil_bascule=seuil_bascule)
    rows = (await db.execute(query)).all()
    # Identification du tenant émetteur : obligatoire sur tout document exporté.
    organisation = await _tenant_display_name(db, organisation_id)

    # Montant déjà payé par réquisition = somme des sorties de fonds validées
    # (même règle que la liste des réquisitions).
    montant_paye_map: dict[Any, Decimal] = {}
    req_ids = [req.id for req, _ in rows]
    if req_ids:
        # Le regroupement est fait par requisition_id : decouper en lots ne
        # change aucun total, chaque cle appartenant a un seul lot.
        for lot in _par_lots(req_ids):
            sortie_res = await db.execute(
                select(
                    SortieFonds.requisition_id,
                    func.coalesce(func.sum(SortieFonds.montant_paye), 0),
                )
                .where(SortieFonds.requisition_id.in_(lot))
                .where((SortieFonds.statut.is_(None)) | (SortieFonds.statut == "VALIDE"))
                .group_by(SortieFonds.requisition_id)
            )
            montant_paye_map.update({row[0]: Decimal(row[1] or 0) for row in sortie_res.all()})

    # Les commentaires d'examen et les étapes de validation partagent une
    # annotation Excel par réquisition. Les utilisateurs sont chargés en une
    # seule requête pour afficher un nom plutôt qu'un identifiant technique.
    validation_user_ids = {
        user_id
        for req, _ in rows
        for user_id in (req.examen_par, req.validee_par, req.approuvee_par)
        if user_id
    }
    validation_users: dict[Any, User] = {}
    if validation_user_ids:
        for lot in _par_lots(sorted(validation_user_ids, key=str)):
            users_res = await db.execute(
                select(User).where(
                    User.id.in_(lot),
                    User.organisation_id == organisation_id,
                )
            )
            validation_users.update({item.id: item for item in users_res.scalars().all()})

    def _validation_user_label(user_id: Any) -> str:
        validator = validation_users.get(user_id)
        if validator:
            return " ".join(filter(None, [validator.prenom, validator.nom])) or validator.email or str(user_id)
        return str(user_id) if user_id else "Non renseigné"

    def _validation_datetime(value: datetime | None) -> str:
        return value.strftime("%d/%m/%Y à %H:%M") if value else "Non renseignée"

    def _requisition_annotation(req: Requisition) -> str:
        entries: list[str] = []
        if req.examen_commentaire or req.examen_par or req.examen_le:
            entries.append(
                "Examen\n"
                f"Auteur : {_validation_user_label(req.examen_par)}\n"
                f"Date et heure : {_validation_datetime(req.examen_le)}\n"
                f"Commentaire : {(req.examen_commentaire or 'Aucun commentaire').strip()}"
            )
        if req.validee_par or req.validee_le:
            entries.append(
                "Validation technique\n"
                f"Auteur : {_validation_user_label(req.validee_par)}\n"
                f"Date et heure : {_validation_datetime(req.validee_le)}"
            )
        if req.approuvee_par or req.approuvee_le:
            entries.append(
                "Approbation / visa\n"
                f"Auteur : {_validation_user_label(req.approuvee_par)}\n"
                f"Date et heure : {_validation_datetime(req.approuvee_le)}"
            )
        return "\n\n".join(entries)

    def _build_workbook() -> tuple[Workbook, str]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Réquisitions"

        headers = [
            "N° Réquisition",
            "Date",
            "Objet",
            "Service",
            "Type",
            "Statut",
            "Montant total (USD)",
            "Montant déjà payé (USD)",
            "Reliquat (USD)",
        ]

        data_rows: list[list[Any]] = []
        total_montant = Decimal("0")
        total_paye = Decimal("0")
        by_statut_count: dict[str, int] = defaultdict(int)
        by_statut_amount: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
        by_service_amount: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))

        for req, service in rows:
            montant_total = _round_money(req.montant_total)
            paye = _round_money(montant_paye_map.get(req.id, Decimal("0")))
            reliquat = _round_money(montant_total - paye)
            total_montant += montant_total
            total_paye += paye

            service_label = f"{service.code} - {service.libelle}" if service else ""
            statut_label = req.status or "Non précisé"
            by_statut_count[statut_label] += 1
            by_statut_amount[statut_label] += montant_total
            by_service_amount[service_label or "Sans service"] += montant_total

            data_rows.append(
                [
                    req.numero_requisition or "",
                    (req.date_requisition or req.created_at).strftime("%d/%m/%Y")
                    if (req.date_requisition or req.created_at)
                    else "",
                    req.objet or "",
                    service_label,
                    req.type_requisition or "",
                    statut_label,
                    float(montant_total),
                    float(paye),
                    float(reliquat),
                ]
            )

        periode = f"{date_debut or 'début'} → {date_fin or 'fin'}"
        _build_list_sheet(
            ws,
            title="RÉQUISITIONS",
            subtitle=f"Période : {periode}  |  Montants en USD",
            headers=headers,
            data_rows=data_rows,
            money_cols=(7, 8, 9),
            total_values={
                7: float(total_montant),
                8: float(total_paye),
                9: float(total_montant - total_paye),
            },
            organisation=organisation,
        )

        # Comme les commentaires du budget, le commentaire de l'examinateur est
        # une annotation Excel native (triangle rouge), attachée à l'objet de la
        # réquisition plutôt qu'ajoutée comme colonne visible.
        for row_idx, (req, _) in enumerate(rows, start=5):
            commentaire = _requisition_annotation(req)
            if commentaire:
                note = Comment(commentaire, "ONEC")
                note.width = 320
                note.height = min(60 + 14 * (commentaire.count("\n") + len(commentaire) // 45 + 1), 400)
                ws.cell(row=row_idx, column=4).comment = note

        statut_rows = [
            [statut_label, by_statut_count[statut_label], float(by_statut_amount[statut_label])]
            for statut_label in sorted(
                by_statut_amount, key=lambda s: by_statut_amount[s], reverse=True
            )
        ]
        service_rows = [
            [svc, float(amount)]
            for svc, amount in sorted(by_service_amount.items(), key=lambda kv: kv[1], reverse=True)
        ]
        _build_synthese_sheet(
            wb,
            "Synthèse — Réquisitions",
            [
                {
                    "title": "Répartition par statut",
                    "headers": ["Statut", "Nombre", "Montant total (USD)"],
                    "rows": statut_rows,
                    "money_cols": (3,),
                },
                {
                    "title": "Répartition par service",
                    "headers": ["Service", "Montant total (USD)"],
                    "rows": service_rows,
                    "money_cols": (2,),
                },
            ],
            chart_title="Montant total par statut",
            chart_block_index=0,
            chart_value_col=3,
            organisation=organisation,
        )

        suffix = f"{date_debut or 'debut'}_{date_fin or 'fin'}"
        filename = f"requisitions_{suffix}.xlsx"
        return wb, filename

    await _relacher_connexion(db)
    return await anyio.to_thread.run_sync(_build_workbook)


@router.get("/requisitions", dependencies=[Depends(has_permission("requisitions"))])
async def export_requisitions(
    date_debut: str | None = Query(default=None),
    date_fin: str | None = Query(default=None),
    statut: str | None = Query(default=None),
    service_id: int | None = Query(default=None),
    type_requisition: str | None = Query(default=None),
    mode_paiement: str | None = Query(default=None),
    budget_poste_id: int | None = Query(default=None),
    search: str | None = Query(default=None),
    objet: str | None = Query(default=None),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> Response:
    """Rend le classeur directement, ou 202 et un job au-dessus du seuil.

    L'aiguillage est cote SERVEUR : c'est lui qui decide du regime, le client
    s'adapte a ce qu'il recoit (frontend/src/utils/download.ts traite 200 et
    202). C'est ce qui rend la bascule reversible type par type sans
    redeploiement du frontend — EXPORT_ASYNC_TYPES suffit, dans les deux sens.
    """
    params = {
            "date_debut": date_debut,
            "date_fin": date_fin,
            "statut": statut,
            "service_id": service_id,
            "type_requisition": type_requisition,
            "mode_paiement": mode_paiement,
            "budget_poste_id": budget_poste_id,
            "search": search,
            "objet": objet,
    }
    try:
        wb, filename = await construire_classeur_requisitions(
            db,
            user.organisation_id,
            date_debut=date_debut,
            date_fin=date_fin,
            statut=statut,
            service_id=service_id,
            type_requisition=type_requisition,
            mode_paiement=mode_paiement,
            budget_poste_id=budget_poste_id,
            search=search,
            objet=objet,
            seuil_bascule=_seuil_bascule("requisitions"),
        )
    except BasculeAsynchroneRequise as bascule:
        return await _soumettre_export_asynchrone(
            db,
            user,
            type_export="requisitions",
            params=params,
            row_count=bascule.total,
        )
    return await _excel_response(filename, wb)


async def construire_classeur_experts(
    db: AsyncSession,
    organisation_id: int,
    *,
    q: str | None = None,
    statut_professionnel: str | None = None,
    include_inactive: bool = False,
    active: bool | None = True,
    order: str | None = None,
    seuil_bascule: int | None = None,
) -> tuple[Workbook, str]:
    """Construit le classeur `experts-comptables` et rend `(classeur, nom de fichier)`.

    EXTRAIT DE L'ENDPOINT, sans autre changement que `user.organisation_id`
    remplace par `organisation_id`. Le worker et la route HTTP appellent
    cette meme fonction : c'est ce qui rend leur divergence impossible
    plutot qu'improbable.
    """
    query = select(ExpertComptable)

    if q:
        q_value = f"%{q.strip()}%"
        query = query.where(
            or_(
                ExpertComptable.numero_ordre.ilike(q_value),
                ExpertComptable.nom_denomination.ilike(q_value),
                ExpertComptable.email.ilike(q_value),
                ExpertComptable.cabinet_attache.ilike(q_value),
            )
        )
    if statut_professionnel:
        variants = _statut_professionnel_variants(statut_professionnel)
        if variants:
            query = query.where(func.trim(ExpertComptable.statut_professionnel).in_(variants))
    if not include_inactive and active is not None:
        query = query.where(ExpertComptable.active == active)

    if order:
        parts = order.split(".")
        field = parts[0]
        direction = parts[1] if len(parts) > 1 else "asc"
        column_map = {
            "numero_ordre": ExpertComptable.numero_ordre,
            "nom_denomination": ExpertComptable.nom_denomination,
            "created_at": ExpertComptable.created_at,
            "statut_professionnel": ExpertComptable.statut_professionnel,
        }
        col = column_map.get(field)
        if col is not None:
            query = query.order_by(col.desc() if direction.lower() == "desc" else col.asc())
        else:
            query = query.order_by(ExpertComptable.numero_ordre.asc())
    else:
        query = query.order_by(ExpertComptable.numero_ordre.asc())

    await _compter_lignes(db, query, export="experts-comptables", seuil_bascule=seuil_bascule)
    experts = (await db.execute(query)).scalars().all()
    # Identification du tenant émetteur : obligatoire sur tout document exporté.
    organisation = await _tenant_display_name(db, organisation_id)

    def _build_workbook() -> tuple[Workbook, str]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Experts"

        headers = [
            "N° Ordre",
            "Nom/Dénomination",
            "Type",
            "Catégorie Personne",
            "Statut Professionnel",
            "Cabinet Attache",
            "Email",
            "Téléphone",
            "État",
        ]
        data_rows = [
            [
                expert.numero_ordre,
                expert.nom_denomination,
                expert.type_ec,
                expert.categorie_personne or "",
                expert.statut_professionnel or "",
                expert.cabinet_attache or "",
                expert.email or "",
                expert.telephone or "",
                "Actif" if expert.active else "Archivé",
            ]
            for expert in experts
        ]
        _build_list_sheet(
            ws,
            title="Experts-comptables",
            subtitle=f"{len(experts)} expert(s)",
            headers=headers,
            data_rows=data_rows,
            organisation=organisation,
        )

        filename = "experts_comptables.xlsx"
        return wb, filename

    return await anyio.to_thread.run_sync(_build_workbook)


@router.get("/experts-comptables")
async def export_experts(
    q: str | None = Query(default=None),
    statut_professionnel: str | None = Query(default=None),
    include_inactive: bool = Query(default=False),
    active: bool | None = Query(default=True),
    order: str | None = Query(default=None),
    user: User = Depends(require_expert_admin),
    db: AsyncSession = Depends(get_db),
) -> Response:
    """Rend le classeur directement, ou 202 et un job au-dessus du seuil.

    L'aiguillage est cote SERVEUR : c'est lui qui decide du regime, le client
    s'adapte a ce qu'il recoit (frontend/src/utils/download.ts traite 200 et
    202). C'est ce qui rend la bascule reversible type par type sans
    redeploiement du frontend — EXPORT_ASYNC_TYPES suffit, dans les deux sens.
    """
    params = {
            "q": q,
            "statut_professionnel": statut_professionnel,
            "include_inactive": include_inactive,
            "active": active,
            "order": order,
    }
    try:
        wb, filename = await construire_classeur_experts(
            db,
            user.organisation_id,
            q=q,
            statut_professionnel=statut_professionnel,
            include_inactive=include_inactive,
            active=active,
            order=order,
            seuil_bascule=_seuil_bascule("experts-comptables"),
        )
    except BasculeAsynchroneRequise as bascule:
        return await _soumettre_export_asynchrone(
            db,
            user,
            type_export="experts-comptables",
            params=params,
            row_count=bascule.total,
        )
    return await _excel_response(filename, wb)
