"""Génération du tableau provincial de sortie (.xlsx).

Régénère un classeur avec une feuille par section (Société, Cabinet,
Indépendant, Salarié, Stagiaire) et la colonne « Conclusion » calculée par le
moteur de verdict, en tenant compte des réglages de l'organisation.
"""
from __future__ import annotations

import io
from typing import Any

from . import verdict as verdict_engine

# (clé logique, libellé colonne) par catégorie
_COMMON = [
    ("index", "N°"),
    ("numero_ordre", "N° d'ordre"),
    ("nom", "Nom, post-noms et prénoms"),
    ("sexe", "Sexe"),
    ("telephone", "N° téléphone"),
    ("email", "E-mail"),
    ("date_naissance", "Date naissance"),
    ("age", "Âge"),
    ("anciennete", "Ancienneté"),
]

_LAYOUT: dict[str, list[tuple[str, str]]] = {
    "EC Cabinet": _COMMON + [
        ("cabinet", "Cabinet d'attache"), ("cotisation", "Cotisation"),
        ("heures", "Heures formation"), ("conclusion", "Conclusion"), ("motif", "Observation"),
    ],
    "EC Salarié": _COMMON + [
        ("cabinet", "Employeur"), ("cotisation", "Cotisation"),
        ("heures", "Heures formation"), ("conclusion", "Conclusion"), ("motif", "Observation"),
    ],
    "EC Indépendant": _COMMON + [
        ("nif", "NIF"), ("cotisation", "Cotisation"), ("assurance", "Assurance valide"),
        ("chiffre_affaires", "Chiffre d'affaires"), ("heures", "Heures formation"),
        ("conclusion", "Conclusion"), ("motif", "Observation"),
    ],
    "Société": [
        ("index", "N°"), ("numero_ordre", "N° d'ordre"), ("nom", "Dénomination"),
        ("gerant", "Associé gérant"), ("nif", "NIF"), ("telephone", "N° téléphone"),
        ("email", "E-mail"), ("cotisation", "Cotisation"), ("assurance", "Assurance"),
        ("chiffre_affaires", "Chiffre d'affaires"), ("conclusion", "Conclusion"), ("motif", "Observation"),
    ],
    "Stagiaire": [
        ("index", "N°"), ("numero_ordre", "N° d'ordre"), ("nom", "Nom"),
        ("sexe", "Sexe"), ("telephone", "Téléphone"), ("email", "E-mail"),
    ],
}

_SHEET_TITLE = {
    "Société": "SOCIETES",
    "EC Cabinet": "EC EN CABINET",
    "EC Indépendant": "EC INDEPENDANTS",
    "EC Salarié": "EC SALARIES",
    "Stagiaire": "STAGIAIRES",
}
_SECTION_LABEL = {
    "Société": "SECTION C : SOCIETES D'EXPERTISE COMPTABLE",
    "EC Cabinet": "SECTION A1 : EXPERTS-COMPTABLES EN CABINET",
    "EC Indépendant": "SECTION A2 : EXPERTS-COMPTABLES INDEPENDANTS",
    "EC Salarié": "SECTION B : EXPERTS-COMPTABLES SALARIES",
    "Stagiaire": "STAGIAIRES",
}
_ORDER = ["Société", "EC Cabinet", "EC Indépendant", "EC Salarié", "Stagiaire"]


def _bool_fr(v: Any) -> str:
    if v is True:
        return "OUI"
    if v is False:
        return "NON"
    return ""


def _cell_value(key: str, dossier: dict[str, Any], idx: int, verdict_res: dict[str, Any]) -> Any:
    raw = dossier.get("raw_data") or {}
    if key == "index":
        return idx
    if key == "conclusion":
        return verdict_res.get("conclusion", "")
    if key == "motif":
        return verdict_res.get("motif", "")
    if key == "heures":
        return dossier.get("heures_forco")
    if key in ("cotisation", "assurance", "chiffre_affaires"):
        val = dossier.get("cotisation_payee") if key == "cotisation" else (
            dossier.get("assurance") if key == "assurance"
            else (dossier.get("chiffre_affaires") if dossier.get("chiffre_affaires") is not None
                  else raw.get("chiffre_affaires"))
        )
        return _bool_fr(val)
    if key in ("sexe", "date_naissance", "age", "anciennete", "nif"):
        return dossier.get(key) if dossier.get(key) is not None else raw.get(key)
    if key == "gerant":
        return raw.get("gerant")
    return dossier.get(key)


def build_workbook(dossiers: list[dict[str, Any]], exercice: str,
                   reglages: "verdict_engine.TableauReglages | None" = None,
                   exercice_annee: int | None = None,
                   organisation: str = "CONSEIL PROVINCIAL") -> bytes:
    """Construit le classeur de sortie et renvoie les octets .xlsx."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    reglages = reglages or verdict_engine.TableauReglages()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    groups: dict[str, list[dict]] = {}
    for d in dossiers:
        groups.setdefault(d.get("categorie") or "Inconnu", []).append(d)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(bold=True, size=12)
    fills = {
        verdict_engine.INSCRIT: PatternFill("solid", fgColor="C6EFCE"),
        verdict_engine.NON_INSCRIT: PatternFill("solid", fgColor="FFC7CE"),
        verdict_engine.A_DELIBERER: PatternFill("solid", fgColor="FFEB9C"),
    }

    for cat in _ORDER + [c for c in groups if c not in _ORDER]:
        rows_data = groups.get(cat)
        if not rows_data:
            continue
        layout = _LAYOUT.get(cat, _LAYOUT["EC Cabinet"])
        ws = wb.create_sheet(_SHEET_TITLE.get(cat, cat[:31])[:31])

        ws.cell(1, 1, organisation).font = title_font
        ws.cell(2, 1, _SECTION_LABEL.get(cat, cat)).font = title_font
        ws.cell(3, 1, f"Exercice {exercice} — seuil {int(reglages.heures_formation_min)}h").font = Font(italic=True)

        header_row = 5
        for c, (_key, label) in enumerate(layout, start=1):
            cell = ws.cell(header_row, c, label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        concl_col = next((i for i, (k, _l) in enumerate(layout, start=1) if k == "conclusion"), None)
        r = header_row + 1
        idx = 1
        for d in rows_data:
            vres = verdict_engine.evaluer(_to_eval_dict(d), exercice_annee, reglages)
            for c, (key, _label) in enumerate(layout, start=1):
                ws.cell(r, c, _cell_value(key, d, idx, vres))
            if concl_col and vres.get("conclusion") in fills:
                ws.cell(r, concl_col).fill = fills[vres["conclusion"]]
            r += 1
            idx += 1

        for c, (_key, label) in enumerate(layout, start=1):
            width = 30 if _key in ("nom", "email", "motif") else 14
            ws.column_dimensions[get_column_letter(c)].width = width
        ws.freeze_panes = ws.cell(header_row + 1, 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_eval_dict(d: dict[str, Any]) -> dict[str, Any]:
    raw = d.get("raw_data") or {}

    def pick(key):
        return d.get(key) if d.get(key) is not None else raw.get(key)

    return {
        "categorie": d.get("categorie"),
        "numero_ordre": d.get("numero_ordre"),
        "cotisation_payee": d.get("cotisation_payee"),
        "assurance": d.get("assurance"),
        "chiffre_affaires": pick("chiffre_affaires"),
        "heures_forco": d.get("heures_forco"),
        "hformation": raw.get("hformation"),
        "anciennete": pick("anciennete"),
        "age": pick("age"),
    }
