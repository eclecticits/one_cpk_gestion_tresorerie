#!/usr/bin/env python3
"""Compare deux classeurs cellule par cellule : valeurs ET rendu.

    python comparer_classeurs.py avant.xlsx apres.xlsx

Existe pour une raison precise : toute optimisation d'un export doit laisser le
fichier VISUELLEMENT identique. Comparer les tailles ou les valeurs ne suffit
pas — une regression de style (bordure perdue, format monetaire absent, zebrage
casse) ne se verrait pas.

On compare donc, pour chaque cellule : la valeur, le format de nombre, la
police, le remplissage, la bordure et l'alignement.
"""

from __future__ import annotations

import sys

from openpyxl import load_workbook


def signature_style(cellule) -> tuple:
    police = cellule.font
    remplissage = cellule.fill
    bordure = cellule.border
    alignement = cellule.alignment
    return (
        cellule.number_format,
        (police.bold, police.italic, police.size, police.color.rgb if police.color else None),
        (remplissage.fill_type, remplissage.fgColor.rgb if remplissage.fgColor else None),
        tuple(
            (getattr(bordure, cote).style, getattr(bordure, cote).color.rgb
             if getattr(bordure, cote).color else None)
            for cote in ("left", "right", "top", "bottom")
        ),
        (alignement.horizontal, alignement.vertical, alignement.wrap_text),
    )


def comparer(chemin_a: str, chemin_b: str) -> int:
    wa = load_workbook(chemin_a)
    wb = load_workbook(chemin_b)

    if wa.sheetnames != wb.sheetnames:
        print(f"ECART : feuilles differentes\n  avant {wa.sheetnames}\n  apres {wb.sheetnames}")
        return 1

    total = 0
    ecarts_valeur = 0
    ecarts_style = 0
    exemples: list[str] = []

    for nom in wa.sheetnames:
        fa, fb = wa[nom], wb[nom]
        if (fa.max_row, fa.max_column) != (fb.max_row, fb.max_column):
            print(f"ECART : dimensions de « {nom} » — "
                  f"avant {fa.max_row}x{fa.max_column}, apres {fb.max_row}x{fb.max_column}")
            return 1

        for ligne_a, ligne_b in zip(fa.iter_rows(), fb.iter_rows()):
            for ca, cb in zip(ligne_a, ligne_b):
                total += 1
                if ca.value != cb.value:
                    ecarts_valeur += 1
                    if len(exemples) < 5:
                        exemples.append(f"  valeur {nom}!{ca.coordinate} : {ca.value!r} -> {cb.value!r}")
                    continue
                sa, sb = signature_style(ca), signature_style(cb)
                if sa != sb:
                    ecarts_style += 1
                    if len(exemples) < 5:
                        differences = [
                            f"{champ}: {x!r} -> {y!r}"
                            for champ, x, y in zip(
                                ("format", "police", "remplissage", "bordure", "alignement"), sa, sb)
                            if x != y
                        ]
                        exemples.append(f"  style {nom}!{ca.coordinate} : {'; '.join(differences)}")

        # Largeurs de colonnes : elles font partie du rendu.
        for lettre in set(fa.column_dimensions) | set(fb.column_dimensions):
            la = round(fa.column_dimensions[lettre].width or 0, 3)
            lb = round(fb.column_dimensions[lettre].width or 0, 3)
            if la != lb:
                ecarts_style += 1
                if len(exemples) < 8:
                    exemples.append(f"  largeur {nom}!{lettre} : {la} -> {lb}")

    print(f"Cellules comparees : {total}")
    print(f"  ecarts de valeur : {ecarts_valeur}")
    print(f"  ecarts de style  : {ecarts_style}")
    if exemples:
        print("\nExemples :")
        print("\n".join(exemples))
        return 1
    print("\nIDENTIQUE — valeurs, styles et largeurs de colonnes.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(2)
    sys.exit(comparer(sys.argv[1], sys.argv[2]))
